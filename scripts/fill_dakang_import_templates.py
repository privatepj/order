"""
根据「大康耳机」目录截图行，填写 Downloads 下的产品导入 / 库存录入模板并另存。

用法:
  python scripts/fill_dakang_import_templates.py
  python scripts/fill_dakang_import_templates.py --product-in ... --inventory-in ... --out-dir ...
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# 品名规格, 期末结存(K), 料号（写入模板「规格」列）
DAKANG_ROWS: List[Tuple[str, str, str]] = [
    ("G98尾钉白色", "20", "310105369"),
    ("G98尾钉黑色", "28", "310105370"),
    ("G98尾钉晨光白色", "1", "310108559"),
    ("G98尾钉米白色", "2", "310109529"),
    ("G98尾钉瓦松绿", "11", "310106139"),
    ("G98尾钉淡绿色", "2", "310108748"),
    ("G125尾钉黑色", "2", "310105593"),
    ("G98尾钉肤色", "5", "310109501"),
    ("G98尾钉军绿色", "7", "310109512"),
    ("G98尾钉牙粉色", "2", "310108263"),
    ("G98尾钉烟霞粉色", "2", "310108560"),
    ("G98尾钉鼠尾草绿", "6", "310108968"),
    ("G98尾钉BOAT蓝色", "21", "310107652"),
    ("G98尾钉浅金色", "8", "310108975"),
    ("G98尾钉神秘蓝色", "1", "310108982"),
    ("G98尾钉竞速黑色", "29", "310108534"),
    ("G98尾钉冷灰6C色", "1", "310106997"),
    ("G98尾钉惠普蓝色", "3", "310108477"),
    ("G98尾钉石墨黑色", "2", "3101010224"),
    ("G98尾钉浅金色", "6", "310109931"),
    ("G98尾钉他山玉色", "6", "310109931"),
    ("G176尾钉白色", "8", "3101011273"),
    ("G176尾钉黑色", "5", "310107093"),
    ("G102尾钉白色", "15", "310105020"),
    ("G102尾钉黑色", "2", "310105168"),
    ("G171尾钉白色", "46", "3101010469"),
    ("G171尾钉黑色", "20", "3101010470"),
    ("G171尾钉深蓝色", "1", "310108562"),
]


def _norm_header(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return s.replace("\n", "").replace(" ", "")


def header_map(ws, max_scan_cols: int = 32) -> Dict[str, int]:
    """1-based column index by normalized header text."""
    m: Dict[str, int] = {}
    for c in range(1, max_scan_cols + 1):
        h = ws.cell(1, c).value
        key = _norm_header(h)
        if not key:
            continue
        m[key] = c
    return m


def require_col(hm: Dict[str, int], *candidates: str) -> int:
    for cand in candidates:
        k = _norm_header(cand)
        if k in hm:
            return hm[k]
    raise SystemExit(
        f"缺少列（尝试过: {candidates}）。当前表头: {sorted(hm.keys())}"
    )


def optional_col(hm: Dict[str, int], *candidates: str) -> Optional[int]:
    for cand in candidates:
        k = _norm_header(cand)
        if k in hm:
            return hm[k]
    return None


def qty_from_k(s: str) -> int:
    d = Decimal(str(s).strip())
    return int(d * 1000)


def fill_product_template(path_in: Path, path_out: Path) -> None:
    wb = load_workbook(path_in)
    ws = wb.active
    hm = header_map(ws)
    # Downloads 与系统一致：产品编号、产品名称、规格、基础单位、备注、系列
    c_code = optional_col(hm, "产品编号", "料号")
    c_name = require_col(hm, "产品名称", "品名")
    c_spec = require_col(hm, "规格")
    c_unit = require_col(hm, "基础单位", "单位")
    c_remark = optional_col(hm, "备注")
    c_series = optional_col(hm, "系列")

    start = 2
    for i, (name_text, k_str, material_no) in enumerate(DAKANG_ROWS):
        r = start + i
        if c_code is not None:
            ws.cell(r, c_code, None)
        # 产品名称 = 品名规格；规格列 = 料号
        ws.cell(r, c_name, name_text)
        ws.cell(r, c_spec, material_no)
        ws.cell(r, c_unit, "PCS")
        if c_remark is not None:
            ws.cell(r, c_remark, None)
        if c_series is not None:
            ws.cell(r, c_series, None)

    path_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path_out)
    wb.close()


def fill_inventory_template(path_in: Path, path_out: Path) -> None:
    wb = load_workbook(path_in)
    ws = wb.active
    hm = header_map(ws)
    c_name = require_col(hm, "品名", "产品名称")
    c_spec = require_col(hm, "规格")
    c_area = require_col(hm, "仓储区", "库区", "仓库")
    c_qty = require_col(hm, "数量")
    c_unit = optional_col(hm, "单位")
    c_remark = optional_col(hm, "备注")

    start = 2
    for i, (name_text, k_str, material_no) in enumerate(DAKANG_ROWS):
        r = start + i
        ws.cell(r, c_name, name_text)
        ws.cell(r, c_spec, material_no)
        ws.cell(r, c_area, "暂定")
        try:
            q = qty_from_k(k_str)
        except (InvalidOperation, ValueError) as exc:
            raise SystemExit(f"数量换算失败 行{i+1} k={k_str!r}") from exc
        ws.cell(r, c_qty, q)
        if c_unit is not None:
            ws.cell(r, c_unit, "PCS")
        if c_remark is not None:
            ws.cell(r, c_remark, None)

    path_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path_out)
    wb.close()


def default_paths() -> Tuple[Path, Path, Path]:
    home = Path.home()
    downloads = home / "Downloads"
    prod = downloads / "产品导入模板.xlsx"
    inv = downloads / "库存录入导入模板 (2).xlsx"
    return prod, inv, downloads


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="填写大康耳机目录行到产品/库存导入模板")
    p.add_argument("--product-in", type=Path, default=None)
    p.add_argument("--inventory-in", type=Path, default=None)
    p.add_argument("--out-dir", type=Path, default=None)
    p.add_argument(
        "--product-out",
        type=Path,
        default=None,
        help="默认: <out-dir>/产品导入模板_已填.xlsx",
    )
    p.add_argument(
        "--inventory-out",
        type=Path,
        default=None,
        help="默认: <out-dir>/库存录入导入模板_已填.xlsx",
    )
    args = p.parse_args(argv)

    d_prod, d_inv, d_out = default_paths()
    prod_in = args.product_in or d_prod
    inv_in = args.inventory_in or d_inv
    out_dir = args.out_dir or d_out
    prod_out = args.product_out or (out_dir / "产品导入模板_已填.xlsx")
    inv_out = args.inventory_out or (out_dir / "库存录入导入模板_已填.xlsx")

    if not prod_in.is_file():
        print(f"找不到产品模板: {prod_in}", file=sys.stderr)
        return 1
    if not inv_in.is_file():
        print(f"找不到库存模板: {inv_in}", file=sys.stderr)
        return 1

    fill_product_template(prod_in, prod_out)
    fill_inventory_template(inv_in, inv_out)
    print(f"已写入: {prod_out}")
    print(f"已写入: {inv_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
