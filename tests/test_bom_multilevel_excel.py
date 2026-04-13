from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from app import create_app, db
from app.config import Config
from app.models import BomHeader, BomLine, Product, SemiMaterial
from app.services import bom_multilevel_excel, bom_svc


class TestConfig(Config):
    TESTING = True
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_ENGINE_OPTIONS = {}


@pytest.fixture()
def app():
    app = create_app(TestConfig)
    with app.app_context():
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def app_ctx(app):
    with app.app_context():
        yield


def _build_sheet() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "品名"
    ws["B1"] = "P-100"
    ws.cell(3, 1, "分段")
    ws.cell(3, 2, "层级1")
    ws.cell(3, 3, "层级2")
    ws.cell(3, 4, "产品名称")
    ws.cell(3, 5, "规格")
    ws.cell(3, 6, "单个用量")
    ws.cell(3, 7, "数量")
    ws.cell(3, 8, "总用量")
    ws.cell(3, 9, "单位")
    ws.cell(3, 10, "工序")
    return wb


def test_parse_multilevel_sheet_and_topo_order(app_ctx):
    product = Product(product_code="P-100", name="成品A", base_unit="PCS")
    semi_a = SemiMaterial(kind="semi", code="S-A", name="半成品A", base_unit="PCS")
    semi_b = SemiMaterial(kind="semi", code="S-B", name="半成品B", base_unit="PCS")
    mat_a = SemiMaterial(kind="material", code="M-A", name="原料A", base_unit="KG")
    db.session.add_all([product, semi_a, semi_b, mat_a])
    db.session.commit()

    wb = _build_sheet()
    ws = wb.active
    ws.cell(4, 2, "S-A")
    ws.cell(4, 6, 1)
    ws.cell(4, 9, "PCS")
    ws.cell(5, 3, "M-A")
    ws.cell(5, 6, 2)
    ws.cell(5, 9, "KG")
    ws.cell(6, 2, "S-B")
    ws.cell(6, 6, 3)
    ws.cell(6, 9, "PCS")

    parsed = bom_multilevel_excel.parse_multilevel_sheet(ws)
    assert parsed.errors == []
    root_key = ("finished", product.id)
    assert root_key in parsed.groups
    assert ("semi", semi_a.id) in parsed.groups
    assert len(parsed.groups[root_key]) == 2

    ordered = bom_multilevel_excel.topological_parent_order(parsed.groups)
    assert ordered.index(("semi", semi_a.id)) < ordered.index(root_key)


def test_parse_multilevel_sheet_b1_full_name(app_ctx):
    product = Product(product_code="P-100", name="成品A", base_unit="PCS", spec="10mm")
    semi_a = SemiMaterial(kind="semi", code="S-A", name="半成品A", base_unit="PCS")
    db.session.add_all([product, semi_a])
    db.session.commit()

    wb = _build_sheet()
    ws = wb.active
    ws["B1"] = "P-100 - 成品A（10mm）"
    ws.cell(4, 2, "S-A")
    ws.cell(4, 6, 1)
    ws.cell(4, 9, "PCS")

    parsed = bom_multilevel_excel.parse_multilevel_sheet(ws)
    assert parsed.errors == []
    assert ("finished", product.id) in parsed.groups


def test_parse_multilevel_sheet_depth_gap_error(app_ctx):
    product = Product(product_code="P-100", name="成品A", base_unit="PCS")
    semi_a = SemiMaterial(kind="semi", code="S-A", name="半成品A", base_unit="PCS")
    db.session.add_all([product, semi_a])
    db.session.commit()

    wb = _build_sheet()
    ws = wb.active
    ws.cell(4, 3, "S-A")
    ws.cell(4, 6, 1)
    parsed = bom_multilevel_excel.parse_multilevel_sheet(ws)
    assert any("层级跳变过大" in e for e in parsed.errors)


def test_build_multilevel_workbook_for_semi_root(app_ctx):
    semi_a = SemiMaterial(kind="semi", code="S-A", name="半成品A", base_unit="PCS")
    mat_a = SemiMaterial(kind="material", code="M-A", name="原料A", base_unit="KG")
    db.session.add_all([semi_a, mat_a])
    db.session.flush()

    header = BomHeader(parent_kind="semi", parent_product_id=0, parent_material_id=semi_a.id, version_no=1, is_active=True)
    db.session.add(header)
    db.session.flush()
    db.session.add(
        BomLine(
            bom_header_id=header.id,
            line_no=1,
            child_kind="material",
            child_material_id=mat_a.id,
            quantity=Decimal("2"),
            unit="KG",
        )
    )
    db.session.commit()

    buf, _filename = bom_multilevel_excel.build_multilevel_workbook(parent_kind="semi", parent_id=semi_a.id)
    wb = load_workbook(buf, data_only=True)
    ws = wb.active
    assert ws["A1"].value == "品名"
    assert ws["B1"].value == "S-A - 半成品A"
    assert ws.cell(3, 3).value == "产品名称"
    assert ws.cell(4, 2).value == "M-A"
    assert ws.cell(4, 5).value == Decimal("2")

    buf.seek(0)
    wb_style = load_workbook(buf)
    wss = wb_style.active
    assert wss["B1"].font.bold is True
    assert wss["B1"].font.size == 12
    assert wss["A1"].font.bold is True
    assert wss["A1"].font.size == 11
    assert wss["A1"].border.left.style == "thin"
    assert wss["A1"].fill.fill_type == "solid"
    assert wss["B1"].fill.fill_type == "solid"
    assert len(list(wss.merged_cells.ranges)) >= 1


def test_build_multilevel_template_workbook_title_and_style(app_ctx):
    buf = bom_multilevel_excel.build_multilevel_template_workbook(level_columns=3)
    wb = load_workbook(buf)
    ws = wb.active
    assert ws["A1"].value == "品名"
    assert "示例" in (ws["B1"].value or "")
    assert ws["B1"].font.bold is True
    assert ws["B1"].font.size == 12
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.size == 11
    assert ws["A3"].border.left.style == "thin"
    assert ws["A1"].fill.fill_type == "solid"
    assert len(list(ws.merged_cells.ranges)) >= 1


def test_validate_bom_import_batch_detect_cycle_with_db_edges(app_ctx):
    semi_a = SemiMaterial(kind="semi", code="S-A", name="半成品A", base_unit="PCS")
    mat_a = SemiMaterial(kind="material", code="M-A", name="原料A", base_unit="KG")
    db.session.add_all([semi_a, mat_a])
    db.session.flush()

    header = BomHeader(parent_kind="semi", parent_product_id=0, parent_material_id=semi_a.id, version_no=1, is_active=True)
    db.session.add(header)
    db.session.flush()
    db.session.add(
        BomLine(
            bom_header_id=header.id,
            line_no=1,
            child_kind="material",
            child_material_id=mat_a.id,
            quantity=Decimal("1"),
            unit="KG",
        )
    )
    db.session.commit()

    groups = {
        ("material", mat_a.id): [
            {
                "line_no": 1,
                "child_kind": "semi",
                "child_material_id": semi_a.id,
                "quantity": Decimal("1"),
                "unit": "PCS",
                "remark": None,
            }
        ]
    }

    with pytest.raises(ValueError):
        bom_svc.validate_bom_import_batch(groups=groups)
