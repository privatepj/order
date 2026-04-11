from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
import pytest

from app import create_app, db
from app.config import Config
from app.models import (
    Company,
    Customer,
    InventoryMovement,
    PurchaseOrder,
    PurchaseReceipt,
    PurchaseRequisition,
    PurchaseRequisitionLine,
    ProductionMaterialPlanDetail,
    ProductionPreplan,
    Role,
    SemiMaterial,
    Supplier,
    SupplierMaterialMap,
    User,
)
from app.main.routes_procurement import (
    _build_compare_summary,
    _next_purchase_order_no,
    _set_default_supplier_for_material,
    _require_supplier_material_mapping,
    _save_supplier_material_maps,
    _supplier_items_for_material,
    create_purchase_orders_from_requisition,
)
from app.services.production_cost_svc import _material_plan_total_and_unpriced
from app.utils.procurement_order_excel import build_purchase_order_workbook


class TestConfig(Config):
    TESTING = True
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_ENGINE_OPTIONS = {}


@pytest.fixture()
def app():
    app = create_app(TestConfig)
    with app.app_context():
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def app_ctx(app):
    with app.app_context():
        yield


@pytest.fixture()
def admin_client(app):
    with app.app_context():
        role = Role(name="Admin", code="admin")
        db.session.add(role)
        db.session.flush()
        user = User(
            username="admin",
            password_hash="admin",
            role_id=role.id,
            is_active=True,
        )
        db.session.add(user)
        db.session.commit()
        user_id = user.id

    client = app.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = str(user_id)
        session["_fresh"] = True
    return client


def test_purchase_order_numbering_uses_company_code_and_daily_sequence(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    db.session.add(company)
    db.session.flush()
    db.session.add_all(
        [
            PurchaseOrder(
                company_id=company.id,
                po_no="C000120260404001",
                buyer_user_id=1,
                supplier_name="A",
                item_name="M1",
                qty=Decimal("1"),
                unit="pcs",
                unit_price=Decimal("1"),
                amount=Decimal("1"),
            ),
            PurchaseOrder(
                company_id=company.id,
                po_no="C000120260404002",
                buyer_user_id=1,
                supplier_name="A",
                item_name="M2",
                qty=Decimal("1"),
                unit="pcs",
                unit_price=Decimal("1"),
                amount=Decimal("1"),
            ),
        ]
    )
    db.session.commit()

    assert _next_purchase_order_no(company, date(2026, 4, 4)) == "C000120260404003"


def test_purchase_order_excel_export_keeps_template_cells(app_ctx):
    company = Company(name="主体公司", code="C0001", is_default=1)
    po = PurchaseOrder(
        company=company,
        po_no="C000120260404001",
        buyer_user_id=1,
        supplier_name="供应商A",
        supplier_contact_name="张三",
        supplier_phone="13800000000",
        supplier_address="深圳市南山区",
        item_name="物料A",
        item_spec="M8",
        qty=Decimal("12"),
        unit="PCS",
        unit_price=Decimal("3.5"),
        amount=Decimal("42"),
        expected_date=date(2026, 4, 10),
        ordered_at=datetime(2026, 4, 4, 10, 0),
    )

    buf = build_purchase_order_workbook(po)
    wb = load_workbook(buf)
    ws = wb[wb.sheetnames[0]]

    assert ws["A1"].value == "主体公司"
    assert ws["B3"].value == "供应商A"
    assert ws["B4"].value == "张三"
    assert ws["B5"].value == "13800000000"
    assert ws["B6"].value == "深圳市南山区"
    assert ws["J3"].value == "C000120260404001"
    assert ws["J5"].value in ("", None)
    assert ws["B8"].value == "物料A"
    assert ws["D8"].value == "M8"
    assert ws["E8"].value == "PCS"
    assert ws["F8"].value == 12
    assert ws["G8"].value == 3.5
    assert ws["H8"].value == "=G8*F8"


def test_signed_requisition_generates_one_purchase_order_per_line(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(
        kind="material", code="M001", name="物料A", spec="10*10", base_unit="KG"
    )
    db.session.add_all([company, supplier, material])
    db.session.flush()
    supplier.company_id = company.id
    db.session.add(
        SupplierMaterialMap(
            company_id=company.id,
            supplier_id=supplier.id,
            material_id=material.id,
            is_active=True,
            last_unit_price=Decimal("8.8"),
        )
    )
    requisition = PurchaseRequisition(
        company_id=company.id,
        req_no="REQ202604040001",
        requester_user_id=1,
        supplier_name=supplier.name,
        item_name=material.name,
        qty=Decimal("5"),
        unit="KG",
        signed_at=datetime(2026, 4, 4, 9, 0),
        signed_by=1,
    )
    db.session.add(requisition)
    db.session.flush()
    requisition.lines.append(
        PurchaseRequisitionLine(
            company_id=company.id,
            requisition_id=requisition.id,
            line_no=1,
            supplier_id=supplier.id,
            material_id=material.id,
            supplier_name=supplier.name,
            item_name=material.name,
            item_spec=material.spec,
            qty=Decimal("5"),
            unit="KG",
            status="pending_order",
        )
    )
    db.session.flush()

    orders = create_purchase_orders_from_requisition(
        requisition,
        buyer_user_id=2,
        order_date=date(2026, 4, 4),
    )

    assert len(orders) == 1
    assert orders[0].po_no == "C000120260404001"
    assert orders[0].unit_price == Decimal("8.8")
    assert orders[0].supplier_name == "供应商A"
    assert requisition.lines[0].status == "ordered"
    assert requisition.status == "ordered"


def test_compare_summary_uses_order_receipt_and_inventory_quantities(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, material])
    db.session.flush()
    po = PurchaseOrder(
        company_id=company.id,
        po_no="C000120260404001",
        buyer_user_id=1,
        material_id=material.id,
        supplier_name="供应商A",
        item_name="物料A",
        qty=Decimal("10"),
        unit="KG",
        unit_price=Decimal("2"),
        amount=Decimal("20"),
        status="ordered",
    )
    db.session.add(po)
    db.session.flush()
    receipt = PurchaseReceipt(
        company_id=company.id,
        receipt_no="RCV202604040001",
        purchase_order_id=po.id,
        receiver_user_id=1,
        received_qty=Decimal("10"),
        received_at=datetime(2026, 4, 4, 11, 0),
        status="posted",
    )
    db.session.add(receipt)
    db.session.flush()
    db.session.add(
        InventoryMovement(
            category="material",
            direction="in",
            product_id=0,
            material_id=material.id,
            storage_area="A01",
            quantity=Decimal("10"),
            unit="KG",
            biz_date=date(2026, 4, 4),
            source_type="procurement",
            source_purchase_order_id=po.id,
            source_purchase_receipt_id=receipt.id,
            created_by=1,
        )
    )
    db.session.commit()

    summary = _build_compare_summary(receipt)

    assert summary["ordered_qty"] == Decimal("10")
    assert summary["received_qty"] == Decimal("10")
    assert summary["warehouse_qty"] == Decimal("10")
    assert summary["is_matched"] is True


def test_setting_new_default_supplier_clears_old_default(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier_a = Supplier(company_id=1, name="供应商A", is_active=True)
    supplier_b = Supplier(company_id=1, name="供应商B", is_active=True)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier_a, supplier_b, material])
    db.session.flush()
    supplier_a.company_id = company.id
    supplier_b.company_id = company.id
    db.session.add_all(
        [
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_a.id,
                material_id=material.id,
                is_active=True,
                is_preferred=True,
                last_unit_price=Decimal("8.8"),
            ),
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_b.id,
                material_id=material.id,
                is_active=True,
                is_preferred=False,
                last_unit_price=Decimal("9.9"),
            ),
        ]
    )
    db.session.commit()

    _save_supplier_material_maps(
        supplier_b,
        [
            {
                "company_id": company.id,
                "material_id": material.id,
                "is_preferred": True,
                "remark": None,
                "last_unit_price": Decimal("9.9"),
            }
        ],
    )
    db.session.commit()

    maps = (
        SupplierMaterialMap.query.filter_by(
            company_id=company.id, material_id=material.id
        )
        .order_by(SupplierMaterialMap.supplier_id.asc())
        .all()
    )

    assert len(maps) == 2
    assert maps[0].supplier_id == supplier_a.id
    assert maps[0].is_preferred is False
    assert maps[1].supplier_id == supplier_b.id
    assert maps[1].is_preferred is True


def test_inactive_supplier_cannot_be_saved_as_default(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier = Supplier(company_id=1, name="停用供应商", is_active=False)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier, material])
    db.session.flush()
    supplier.company_id = company.id

    with pytest.raises(ValueError, match="停用供应商不能设置为默认供应商"):
        _save_supplier_material_maps(
            supplier,
            [
                {
                    "company_id": company.id,
                    "material_id": material.id,
                    "is_preferred": True,
                    "remark": None,
                    "last_unit_price": Decimal("5.5"),
                }
            ],
        )


def test_material_supplier_search_orders_default_supplier_first(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier_a = Supplier(company_id=1, name="供应商A", is_active=True)
    supplier_b = Supplier(company_id=1, name="供应商B", is_active=True)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier_a, supplier_b, material])
    db.session.flush()
    supplier_a.company_id = company.id
    supplier_b.company_id = company.id
    db.session.add_all(
        [
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_a.id,
                material_id=material.id,
                is_active=True,
                is_preferred=False,
                last_unit_price=Decimal("8.8"),
            ),
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_b.id,
                material_id=material.id,
                is_active=True,
                is_preferred=True,
                last_unit_price=Decimal("9.9"),
            ),
        ]
    )
    db.session.commit()

    items = _supplier_items_for_material(company.id, material.id)

    assert items[0]["supplier_id"] == supplier_b.id
    assert items[0]["is_default_supplier"] is True
    assert items[0]["last_unit_price"] == "9.9"
    assert {item["supplier_id"] for item in items} == {supplier_a.id, supplier_b.id}


def test_unmapped_supplier_material_pair_is_rejected(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier, material])
    db.session.flush()
    supplier.company_id = company.id

    with pytest.raises(ValueError, match="供应商与物料尚未建立关联"):
        _require_supplier_material_mapping(
            company.id, supplier.id, material.id, row_label="采购单"
        )
def test_supplier_save_without_default_flag_preserves_existing_default(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier, material])
    db.session.flush()
    supplier.company_id = company.id
    db.session.add(
        SupplierMaterialMap(
            company_id=company.id,
            supplier_id=supplier.id,
            material_id=material.id,
            is_active=True,
            is_preferred=True,
            last_unit_price=Decimal("8.8"),
        )
    )
    db.session.commit()

    _save_supplier_material_maps(
        supplier,
        [
            {
                "company_id": company.id,
                "material_id": material.id,
                "remark": "updated",
                "last_unit_price": Decimal("9.5"),
            }
        ],
    )
    db.session.commit()

    refreshed = SupplierMaterialMap.query.filter_by(
        company_id=company.id,
        supplier_id=supplier.id,
        material_id=material.id,
    ).one()

    assert refreshed.is_preferred is True
    assert refreshed.last_unit_price == Decimal("9.50")
    assert refreshed.remark == "updated"


def test_setting_default_supplier_from_material_side_clears_old_default(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    supplier_a = Supplier(company_id=1, name="供应商A", is_active=True)
    supplier_b = Supplier(company_id=1, name="供应商B", is_active=True)
    material = SemiMaterial(kind="material", code="M001", name="物料A", base_unit="KG")
    db.session.add_all([company, supplier_a, supplier_b, material])
    db.session.flush()
    supplier_a.company_id = company.id
    supplier_b.company_id = company.id
    db.session.add_all(
        [
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_a.id,
                material_id=material.id,
                is_active=True,
                is_preferred=True,
                last_unit_price=Decimal("8.8"),
            ),
            SupplierMaterialMap(
                company_id=company.id,
                supplier_id=supplier_b.id,
                material_id=material.id,
                is_active=True,
                is_preferred=False,
                last_unit_price=Decimal("9.9"),
            ),
        ]
    )
    db.session.commit()

    _set_default_supplier_for_material(
        company_id=company.id,
        material_id=material.id,
        supplier_id=supplier_b.id,
    )
    db.session.commit()

    maps = (
        SupplierMaterialMap.query.filter_by(
            company_id=company.id,
            material_id=material.id,
        )
        .order_by(SupplierMaterialMap.supplier_id.asc())
        .all()
    )

    assert maps[0].is_preferred is False
    assert maps[1].is_preferred is True


def test_old_material_list_route_redirects_to_procurement_materials(admin_client, app_ctx):
    response = admin_client.get("/semi-materials?kind=material", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/procurement/materials?keyword=&page=1")


def test_material_cost_prefers_default_supplier_price(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    customer = Customer(customer_code="C001", name="客户A", company_id=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(
        kind="material",
        code="M001",
        name="物料A",
        base_unit="KG",
        standard_unit_cost=Decimal("3.5"),
    )
    db.session.add_all([company, customer, supplier, material])
    db.session.flush()
    customer.company_id = company.id
    supplier.company_id = company.id
    preplan = ProductionPreplan(
        source_type="manual",
        plan_date=date(2026, 4, 4),
        customer_id=customer.id,
        created_by=1,
    )
    db.session.add(preplan)
    db.session.flush()
    db.session.add(
        SupplierMaterialMap(
            company_id=company.id,
            supplier_id=supplier.id,
            material_id=material.id,
            is_active=True,
            is_preferred=True,
            last_unit_price=Decimal("9.9"),
        )
    )
    db.session.add(
        ProductionMaterialPlanDetail(
            preplan_id=preplan.id,
            work_order_id=1,
            child_kind="material",
            child_material_id=material.id,
            required_qty=Decimal("2"),
            scrap_qty=Decimal("0"),
            net_required_qty=Decimal("2"),
            stock_covered_qty=Decimal("0"),
            shortage_qty=Decimal("2"),
            unit="KG",
        )
    )
    db.session.commit()

    total, unpriced = _material_plan_total_and_unpriced(preplan_id=preplan.id)

    assert total == Decimal("19.8")
    assert unpriced == 0


def test_material_cost_falls_back_to_standard_cost_when_default_price_missing(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    customer = Customer(customer_code="C001", name="客户A", company_id=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(
        kind="material",
        code="M001",
        name="物料A",
        base_unit="KG",
        standard_unit_cost=Decimal("3.5"),
    )
    db.session.add_all([company, customer, supplier, material])
    db.session.flush()
    customer.company_id = company.id
    supplier.company_id = company.id
    preplan = ProductionPreplan(
        source_type="manual",
        plan_date=date(2026, 4, 4),
        customer_id=customer.id,
        created_by=1,
    )
    db.session.add(preplan)
    db.session.flush()
    db.session.add(
        SupplierMaterialMap(
            company_id=company.id,
            supplier_id=supplier.id,
            material_id=material.id,
            is_active=True,
            is_preferred=True,
            last_unit_price=None,
        )
    )
    db.session.add(
        ProductionMaterialPlanDetail(
            preplan_id=preplan.id,
            work_order_id=1,
            child_kind="material",
            child_material_id=material.id,
            required_qty=Decimal("2"),
            scrap_qty=Decimal("0"),
            net_required_qty=Decimal("2"),
            stock_covered_qty=Decimal("0"),
            shortage_qty=Decimal("2"),
            unit="KG",
        )
    )
    db.session.commit()

    total, unpriced = _material_plan_total_and_unpriced(preplan_id=preplan.id)

    assert total == Decimal("7.0")
    assert unpriced == 0


def test_material_cost_marks_unpriced_when_no_default_price_and_no_standard_cost(app_ctx):
    company = Company(name="Test Co", code="C0001", is_default=1)
    customer = Customer(customer_code="C001", name="客户A", company_id=1)
    supplier = Supplier(company_id=1, name="供应商A", is_active=True)
    material = SemiMaterial(
        kind="material",
        code="M001",
        name="物料A",
        base_unit="KG",
        standard_unit_cost=None,
    )
    db.session.add_all([company, customer, supplier, material])
    db.session.flush()
    customer.company_id = company.id
    supplier.company_id = company.id
    preplan = ProductionPreplan(
        source_type="manual",
        plan_date=date(2026, 4, 4),
        customer_id=customer.id,
        created_by=1,
    )
    db.session.add(preplan)
    db.session.flush()
    db.session.add(
        SupplierMaterialMap(
            company_id=company.id,
            supplier_id=supplier.id,
            material_id=material.id,
            is_active=True,
            is_preferred=True,
            last_unit_price=None,
        )
    )
    db.session.add(
        ProductionMaterialPlanDetail(
            preplan_id=preplan.id,
            work_order_id=1,
            child_kind="material",
            child_material_id=material.id,
            required_qty=Decimal("2"),
            scrap_qty=Decimal("0"),
            net_required_qty=Decimal("2"),
            stock_covered_qty=Decimal("0"),
            shortage_qty=Decimal("2"),
            unit="KG",
        )
    )
    db.session.commit()

    total, unpriced = _material_plan_total_and_unpriced(preplan_id=preplan.id)

    assert total == Decimal("0")
    assert unpriced == 1
