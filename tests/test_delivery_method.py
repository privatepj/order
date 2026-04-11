from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app import create_app, db
from app.config import Config
from app.models import Company, Customer, Delivery, ExpressCompany, ExpressWaybill, OrderItem, Role, SalesOrder, User
from app.services.delivery_svc import create_delivery_from_data, preview_delivery_create
from app.utils.delivery_records_excel import build_delivery_records_workbook


class TestConfig(Config):
    TESTING = True
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_ENGINE_OPTIONS = {}


@pytest.fixture()
def app():
    app = create_app(TestConfig)
    with app.app_context():
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def app_ctx(app):
    with app.app_context():
        yield


@pytest.fixture()
def admin_client(app):
    with app.app_context():
        role = Role(name="Admin", code="admin")
        db.session.add(role)
        db.session.flush()
        user = User(
            username="admin",
            password_hash="admin",
            role_id=role.id,
            is_active=True,
        )
        db.session.add(user)
        db.session.commit()
        user_id = user.id

    client = app.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = str(user_id)
        session["_fresh"] = True
    return client


@pytest.fixture()
def delivery_detail_only_client(app):
    """有送货菜单与详情能力，但不含 delivery.filter.status（列表应固定仅待发）。"""
    with app.app_context():
        role = Role(
            name="送货只详情",
            code="delivery_detail_only",
            allowed_menu_keys=["delivery"],
            allowed_capability_keys=["delivery.action.detail"],
        )
        db.session.add(role)
        db.session.flush()
        user = User(
            username="delivery_reader",
            password_hash="x",
            role_id=role.id,
            is_active=True,
        )
        db.session.add(user)
        db.session.commit()
        user_id = user.id

    client = app.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = str(user_id)
        session["_fresh"] = True
    return client


def _seed_order(with_sf_waybill: bool = False) -> dict:
    company = Company(
        name="测试主体",
        code="TST",
        delivery_no_prefix="DL",
        billing_cycle_day=1,
        is_default=1,
    )
    customer = Customer(
        customer_code="C001",
        short_code="客户A",
        name="客户A",
        company=company,
    )
    order = SalesOrder(
        order_no="SO001",
        customer=customer,
        salesperson="tester",
        order_date=date(2026, 4, 7),
        payment_type="monthly",
    )
    order_item = OrderItem(
        order=order,
        product_name="产品A",
        product_spec="10*10",
        quantity=Decimal("12"),
        unit="PCS",
    )
    db.session.add_all([company, customer, order, order_item])

    express_company = None
    if with_sf_waybill:
        express_company = ExpressCompany(name="顺丰", code="SF", is_active=True)
        db.session.add(express_company)
        db.session.flush()
        db.session.add(
            ExpressWaybill(
                express_company_id=express_company.id,
                waybill_no="SF0001",
                status="available",
            )
        )

    db.session.commit()
    return {
        "company": company,
        "customer": customer,
        "order": order,
        "order_item": order_item,
        "express_company": express_company,
    }


def test_create_delivery_legacy_payload_defaults_to_express(app_ctx):
    seeded = _seed_order(with_sf_waybill=True)

    delivery, err = create_delivery_from_data(
        {
            "customer_id": seeded["customer"].id,
            "lines": [{"order_item_id": seeded["order_item"].id, "quantity": 5}],
            "delivery_date": "2026-04-07",
        }
    )

    assert err is None
    assert delivery is not None
    assert delivery.resolved_delivery_method == "express"
    assert delivery.express_company_id == seeded["express_company"].id
    assert delivery.waybill_no == "SF0001"


def test_create_delivery_pickup_ignores_waybill_and_express_company(app_ctx):
    seeded = _seed_order(with_sf_waybill=True)

    delivery, err = create_delivery_from_data(
        {
            "customer_id": seeded["customer"].id,
            "delivery_method": "pickup",
            "express_company_id": seeded["express_company"].id,
            "waybill_no": "MANUAL001",
            "lines": [{"order_item_id": seeded["order_item"].id, "quantity": 3}],
        }
    )

    assert err is None
    assert delivery is not None
    assert delivery.resolved_delivery_method == "pickup"
    assert delivery.express_company_id is None
    assert delivery.express_waybill_id is None
    assert delivery.waybill_no is None


def test_preview_delivery_returns_pickup_summary(app_ctx):
    seeded = _seed_order(with_sf_waybill=False)

    err, summary = preview_delivery_create(
        {
            "customer_id": seeded["customer"].id,
            "delivery_method": "pickup",
            "waybill_no": "IGNORED",
            "lines": [{"order_item_id": seeded["order_item"].id, "quantity": 2}],
        }
    )

    assert err is None
    assert summary["delivery_method"] == "pickup"
    assert summary["delivery_method_label"] == "自提"
    assert summary["express_company_id"] is None
    assert summary["waybill_no"] is None
    assert "不占运单池" in summary["waybill_note"]


def test_delivery_records_export_shows_pickup_label(app_ctx):
    seeded = _seed_order(with_sf_waybill=False)
    delivery, err = create_delivery_from_data(
        {
            "customer_id": seeded["customer"].id,
            "delivery_method": "pickup",
            "delivery_date": "2026-04-07",
            "lines": [{"order_item_id": seeded["order_item"].id, "quantity": 4}],
        }
    )

    assert err is None
    bio = build_delivery_records_workbook(date(2026, 4, 7), date(2026, 4, 7))
    wb = load_workbook(bio)
    ws = wb[wb.sheetnames[0]]

    assert ws["D1"].value == "配送方式/单号"
    assert ws["D2"].value == "自提"
    assert ws["C2"].value == delivery.delivery_no


def test_delivery_new_web_form_supports_pickup(admin_client, app):
    with app.app_context():
        seeded = _seed_order(with_sf_waybill=False)
        customer_id = seeded["customer"].id
        order_item_id = seeded["order_item"].id

    response = admin_client.post(
        "/deliveries/new",
        data={
            "customer_id": customer_id,
            "delivery_method": "pickup",
            "delivery_date": "2026-04-07",
            "remark": "客户自提",
            "order_item_id": str(order_item_id),
            "delivery_quantity": "6",
        },
        follow_redirects=False,
    )

    assert response.status_code == 302
    with app.app_context():
        delivery = Delivery.query.order_by(Delivery.id.desc()).first()
        assert delivery is not None
        assert delivery.resolved_delivery_method == "pickup"
        assert delivery.waybill_no is None


def test_openclaw_create_delivery_returns_pickup_method(app_ctx, app, monkeypatch):
    seeded = _seed_order(with_sf_waybill=False)
    monkeypatch.setenv("OPENCLAW_API_KEY", "test-openclaw-key")
    client = app.test_client()

    response = client.post(
        "/api/openclaw/deliveries",
        json={
            "customer_id": seeded["customer"].id,
            "delivery_method": "pickup",
            "lines": [{"order_item_id": seeded["order_item"].id, "quantity": 2}],
        },
        headers={"Authorization": "Bearer test-openclaw-key"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["ok"] is True
    assert payload["delivery_method"] == "pickup"


def test_delivery_list_defaults_to_created_without_status_param(app, admin_client):
    with app.app_context():
        seeded = _seed_order(with_sf_waybill=False)
        cid = seeded["customer"].id
        db.session.add_all(
            [
                Delivery(
                    delivery_no="DL-PEND-DEFAULT",
                    delivery_date=date(2026, 4, 7),
                    customer_id=cid,
                    status="created",
                ),
                Delivery(
                    delivery_no="DL-SHIPPED-DEFAULT",
                    delivery_date=date(2026, 4, 7),
                    customer_id=cid,
                    status="shipped",
                ),
            ]
        )
        db.session.commit()

    r = admin_client.get("/deliveries")
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    assert "DL-PEND-DEFAULT" in body
    assert "DL-SHIPPED-DEFAULT" not in body

    r_all = admin_client.get("/deliveries?status=")
    assert r_all.status_code == 200
    body_all = r_all.get_data(as_text=True)
    assert "DL-PEND-DEFAULT" in body_all
    assert "DL-SHIPPED-DEFAULT" in body_all


def test_delivery_list_without_status_filter_cap_forces_created(
    app, delivery_detail_only_client
):
    with app.app_context():
        seeded = _seed_order(with_sf_waybill=False)
        cid = seeded["customer"].id
        db.session.add_all(
            [
                Delivery(
                    delivery_no="DL-PEND-NOFILTER",
                    delivery_date=date(2026, 4, 7),
                    customer_id=cid,
                    status="created",
                ),
                Delivery(
                    delivery_no="DL-SHIP-NOFILTER",
                    delivery_date=date(2026, 4, 7),
                    customer_id=cid,
                    status="shipped",
                ),
            ]
        )
        db.session.commit()

    r = delivery_detail_only_client.get("/deliveries")
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    assert "DL-PEND-NOFILTER" in body
    assert "DL-SHIP-NOFILTER" not in body

    r_bypass = delivery_detail_only_client.get("/deliveries?status=shipped")
    assert r_bypass.status_code == 200
    body_b = r_bypass.get_data(as_text=True)
    assert "DL-PEND-NOFILTER" in body_b
    assert "DL-SHIP-NOFILTER" not in body_b
