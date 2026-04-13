from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Product, SemiMaterial
from app.services import bom_svc

# 第1行：A1「品名」略小；B~工序合并区内父项全称比原四号(14)小一号→12pt
_FONT_SIZE_ROW1_LABEL = 11
_FONT_SIZE_ROW1_PARENT = 12
_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# 第1行浅蓝底（ARGB，无 #）
_FILL_ROW1_BLUE = PatternFill(fill_type="solid", fgColor="FFBDD7EE")


HEADER_ALIASES = {
    "name": ("产品名称", "名称"),
    "spec": ("规格",),
    "qty": ("单个用量", "用量", "用量数量"),
    "unit": ("单位",),
}


@dataclass
class ParseResult:
    root: Tuple[str, int]
    groups: Dict[Tuple[str, int], List[Dict[str, Any]]]
    errors: List[str]


def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _find_header_index(header_row: List[str], aliases: Tuple[str, ...]) -> Optional[int]:
    for idx, title in enumerate(header_row, start=1):
        t = (title or "").strip()
        if not t:
            continue
        if any(alias in t for alias in aliases):
            return idx
    return None


def _parse_qty(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        q = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    if q <= 0:
        return None
    return q


def _format_parent_full_name_product(p: Product) -> str:
    spec = (p.spec or "").strip()
    base = f"{p.product_code} - {p.name}"
    if spec:
        return f"{base}（{spec}）"
    return base


def _format_parent_full_name_semi(sm: SemiMaterial) -> str:
    spec = (sm.spec or "").strip()
    base = f"{sm.code} - {sm.name}"
    if spec:
        return f"{base}（{spec}）"
    return base


def _apply_title_row(ws: Worksheet, *, b1_text: str, merge_end_col: int) -> None:
    """第1行：A1=品名（加粗略小）；B1~merge_end_col 合并为父项全称（加粗，小一号）。"""
    if merge_end_col < 2:
        merge_end_col = 2
    for c in range(1, merge_end_col + 1):
        ws.cell(1, c).fill = _FILL_ROW1_BLUE
    c_a = ws.cell(1, 1, "品名")
    c_a.font = Font(bold=True, size=_FONT_SIZE_ROW1_LABEL)
    c_a.alignment = Alignment(vertical="center", horizontal="center")
    ws.cell(1, 2, b1_text)
    if merge_end_col > 2:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=merge_end_col)
    top_left = ws.cell(1, 2)
    top_left.font = Font(bold=True, size=_FONT_SIZE_ROW1_PARENT)
    top_left.alignment = Alignment(vertical="center", wrap_text=True)
    c_a.fill = _FILL_ROW1_BLUE
    top_left.fill = _FILL_ROW1_BLUE


def _apply_used_range_borders(
    ws: Worksheet, *, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    """已用区域内全部单元格细边框。"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = _BORDER_ALL


def resolve_root_product_code_from_b1(b1_raw: Any) -> Optional[str]:
    """B1 可为纯 product_code，或「编码 - 名称（规格）」全称；返回可解析的 product_code。"""
    s = _cell_text(b1_raw)
    if not s:
        return None
    if Product.query.filter_by(product_code=s).first():
        return s
    if " - " in s:
        candidate = s.split(" - ", 1)[0].strip()
        if candidate and Product.query.filter_by(product_code=candidate).first():
            return candidate
    return None


def build_multilevel_template_workbook(*, level_columns: int = 3) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "多级BOM模板"
    headers = ["分段"] + [f"层级{i}" for i in range(1, level_columns + 1)] + [
        "产品名称",
        "规格",
        "单个用量",
        "数量",
        "总用量",
        "单位",
        "工序",
    ]
    last_col = len(headers)
    _apply_title_row(
        ws,
        b1_text="P001 - 示例成品（填写时请改为实际品名全称或仅填编码）",
        merge_end_col=last_col,
    )
    for col, h in enumerate(headers, start=1):
        ws.cell(3, col, h)
    # 可选示例行（第4行）
    name_col_tpl = 2 + level_columns
    spec_col_tpl = name_col_tpl + 1
    qty_col_tpl = name_col_tpl + 2
    unit_col_tpl = name_col_tpl + 5
    ws.cell(4, 2, "示例子件编码")
    ws.cell(4, name_col_tpl, "示例名称")
    ws.cell(4, spec_col_tpl, "示例规格")
    ws.cell(4, qty_col_tpl, 1)
    ws.cell(4, unit_col_tpl, "PCS")

    last_row = 4
    _apply_used_range_borders(ws, min_row=1, max_row=last_row, min_col=1, max_col=last_col)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_multilevel_sheet(ws: Worksheet, *, header_row: int = 3, root_cell: str = "B1") -> ParseResult:
    errors: List[str] = []
    groups: Dict[Tuple[str, int], List[Dict[str, Any]]] = {}

    b1_raw = ws[root_cell].value
    root_code = resolve_root_product_code_from_b1(b1_raw)
    if not root_code:
        hint = _cell_text(b1_raw)
        if not hint:
            return ParseResult(root=(bom_svc.PARENT_FINISHED, 0), groups={}, errors=["第 1 行：B1 成品编码/品名全称不能为空。"])
        return ParseResult(
            root=(bom_svc.PARENT_FINISHED, 0),
            groups={},
            errors=[f"第 1 行：无法从 B1 解析成品编码（{hint}）。请填写 product_code 或「编码 - 名称（规格）」格式。"],
        )
    product = Product.query.filter_by(product_code=root_code).first()
    if not product:
        return ParseResult(
            root=(bom_svc.PARENT_FINISHED, 0),
            groups={},
            errors=[f"第 1 行：未找到成品编码（{root_code}）。"],
        )
    root = (bom_svc.PARENT_FINISHED, int(product.id))

    max_col = ws.max_column or 1
    header_vals = [_cell_text(ws.cell(header_row, c).value) for c in range(1, max_col + 1)]
    idx_name = _find_header_index(header_vals, HEADER_ALIASES["name"])
    idx_qty = _find_header_index(header_vals, HEADER_ALIASES["qty"])
    idx_unit = _find_header_index(header_vals, HEADER_ALIASES["unit"])
    if not idx_name:
        errors.append(f"第 {header_row} 行：缺少“产品名称”列。")
    if not idx_qty:
        errors.append(f"第 {header_row} 行：缺少“单个用量”列。")
    if errors:
        return ParseResult(root=root, groups={}, errors=errors)
    if idx_name <= 2:
        return ParseResult(root=root, groups={}, errors=[f"第 {header_row} 行：层级列不足，需位于 B 到产品名称列之前。"])

    level_cols = list(range(2, idx_name))
    stack: List[Tuple[str, int]] = [root]
    sm_cache: Dict[str, Optional[SemiMaterial]] = {}
    any_data = False

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        level_values = [_cell_text(ws.cell(row_idx, c).value) for c in level_cols]
        depth = 0
        code = ""
        for i, v in enumerate(level_values, start=1):
            if v:
                depth = i
                code = v
                break
        qty_raw = ws.cell(row_idx, idx_qty).value if idx_qty else None
        unit_raw = ws.cell(row_idx, idx_unit).value if idx_unit else None
        if depth == 0 and not any([qty_raw, unit_raw]):
            continue
        if depth == 0:
            errors.append(f"第 {row_idx} 行：未识别到层级编码。")
            continue
        any_data = True

        if depth > len(stack):
            errors.append(f"第 {row_idx} 行：层级跳变过大，缺少上级节点。")
            continue
        while len(stack) > depth:
            stack.pop()

        qty = _parse_qty(qty_raw)
        if qty is None:
            errors.append(f"第 {row_idx} 行：单个用量必须是大于 0 的数字。")
            continue
        unit = _cell_text(unit_raw)[:16] or None

        if code not in sm_cache:
            sm_cache[code] = SemiMaterial.query.filter_by(code=code).first()
        sm = sm_cache[code]
        if not sm:
            errors.append(f"第 {row_idx} 行：未找到物料编码（{code}）。")
            continue
        if sm.kind not in (bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
            errors.append(f"第 {row_idx} 行：子项类别无效（{sm.kind}）。")
            continue

        parent_key = stack[-1]
        line_no = len(groups.get(parent_key, [])) + 1
        groups.setdefault(parent_key, []).append(
            {
                "line_no": line_no,
                "child_kind": sm.kind,
                "child_material_id": int(sm.id),
                "quantity": qty,
                "unit": unit,
                "remark": None,
            }
        )
        stack.append((sm.kind, int(sm.id)))

    if not any_data:
        errors.append(f"第 {header_row + 1} 行起：未识别到可导入数据。")
    return ParseResult(root=root, groups=groups, errors=errors)


def topological_parent_order(groups: Dict[Tuple[str, int], List[Dict[str, Any]]]) -> List[Tuple[str, int]]:
    keys = set(groups.keys())
    deps: Dict[Tuple[str, int], List[Tuple[str, int]]] = {}
    for k, lines in groups.items():
        arr: List[Tuple[str, int]] = []
        for ln in lines:
            ck = (ln.get("child_kind"), int(ln.get("child_material_id") or 0))
            if ck in keys:
                arr.append(ck)
        deps[k] = arr

    visiting: set[Tuple[str, int]] = set()
    done: set[Tuple[str, int]] = set()
    out: List[Tuple[str, int]] = []

    def dfs(node: Tuple[str, int]) -> None:
        if node in done:
            return
        if node in visiting:
            raise ValueError("导入数据存在循环依赖，无法确定写入顺序。")
        visiting.add(node)
        for c in deps.get(node, []):
            dfs(c)
        visiting.remove(node)
        done.add(node)
        out.append(node)

    for node in groups.keys():
        dfs(node)
    return out


def build_multilevel_workbook(*, parent_kind: str, parent_id: int, max_depth: int = 20) -> Tuple[BytesIO, str]:
    parent_full_name: str
    if parent_kind == bom_svc.PARENT_FINISHED:
        p = Product.query.get(parent_id)
        if not p:
            raise ValueError("未找到成品。")
        parent_full_name = _format_parent_full_name_product(p)
        filename = f"BOM_{p.product_code}.xlsx"
    else:
        sm = SemiMaterial.query.get(parent_id)
        if not sm:
            raise ValueError("未找到半成品/物料。")
        parent_full_name = _format_parent_full_name_semi(sm)
        filename = f"BOM_{sm.code}.xlsx"

    rows: List[Dict[str, Any]] = []
    path: set[Tuple[str, int]] = set()

    def walk(kind: str, node_id: int, depth: int, cumulative_qty: Decimal) -> None:
        if depth > max_depth:
            raise ValueError("导出失败：BOM 层级超过上限。")
        node = (kind, node_id)
        if node in path:
            raise ValueError("导出失败：检测到 BOM 循环引用。")
        header = bom_svc.get_active_bom_header(parent_kind=kind, parent_id=node_id)
        if not header or not header.lines:
            return
        path.add(node)
        lines = sorted(list(header.lines), key=lambda x: x.line_no)
        child_ids = [x.child_material_id for x in lines]
        sm_map = {s.id: s for s in SemiMaterial.query.filter(SemiMaterial.id.in_(child_ids)).all()} if child_ids else {}
        for ln in lines:
            child = sm_map.get(ln.child_material_id)
            if not child:
                continue
            qty = Decimal(str(ln.quantity))
            current_total = cumulative_qty * qty
            rows.append(
                {
                    "depth": depth + 1,
                    "code": child.code,
                    "name": child.name,
                    "spec": child.spec,
                    "quantity": qty,
                    "total_quantity": current_total,
                    "unit": ln.unit or child.base_unit or "",
                }
            )
            walk(ln.child_kind, ln.child_material_id, depth + 1, current_total)
        path.remove(node)

    walk(parent_kind, parent_id, 0, Decimal("1"))
    depth_count = max([r["depth"] for r in rows], default=1)

    headers = ["分段"] + [f"层级{i}" for i in range(1, depth_count + 1)] + [
        "产品名称",
        "规格",
        "单个用量",
        "数量",
        "总用量",
        "单位",
        "工序",
    ]
    last_col = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "多级BOM"
    _apply_title_row(ws, b1_text=parent_full_name, merge_end_col=last_col)
    for col, h in enumerate(headers, start=1):
        ws.cell(3, col, h)

    name_col = 2 + depth_count
    spec_col = name_col + 1
    qty_col = name_col + 2
    count_col = name_col + 3
    total_col = name_col + 4
    unit_col = name_col + 5
    process_col = name_col + 6

    for row_idx, item in enumerate(rows, start=4):
        ws.cell(row_idx, 1, None)
        ws.cell(row_idx, 1 + int(item["depth"]), item["code"])
        ws.cell(row_idx, name_col, item["name"])
        ws.cell(row_idx, spec_col, item["spec"])
        ws.cell(row_idx, qty_col, item["quantity"])
        ws.cell(row_idx, count_col, 1)
        ws.cell(row_idx, total_col, item["total_quantity"])
        ws.cell(row_idx, unit_col, item["unit"])
        ws.cell(row_idx, process_col, "")

    last_row = 3 + len(rows) if rows else 3
    _apply_used_range_borders(ws, min_row=1, max_row=last_row, min_col=1, max_col=last_col)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename
