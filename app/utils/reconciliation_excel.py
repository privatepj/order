"""对账单 Excel（送货明细、抬头/账户与样张一致）。"""
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy.orm import joinedload

from app import db
from app.models import Customer, Delivery, DeliveryItem, OrderItem, SalesOrder
from app.models import Company
from app.models.product import CustomerProduct

if TYPE_CHECKING:
    from datetime import date

COL_LAST = 10
FONT_SONG = "宋体"
thin = Side(style="thin")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 模板列宽 A–J
COL_WIDTHS = (10.625, 14.125, 10, 28.75, 10.25, 9.0, 10, 11.5, 12.25, 6.125)

FOOTER_NOTES = (
    "注：1. 我司收款账户只有下列账户，若转入别的账户，我司不接受也不承认，发生纠纷一切责任由贵司承担。",
    "2. 请贵司财务及时核对此单，并请在3日内回签，逾时（视同确认无误）我司将入账，请予以准时付款，谢谢支持！",
)


def _material_no(oi: OrderItem) -> str:
    if oi.customer_product:
        mn = oi.customer_product.material_no or ""
        if oi.customer_product.product:
            pc = oi.customer_product.product.product_code or ""
            return pc or mn
        return mn or ""
    return (oi.customer_material_no or "") or ""


def _qty_display(q) -> str:
    if q is None:
        return ""
    d = Decimal(str(q))
    if d == d.to_integral_value():
        return str(int(d))
    return format(float(d), ".4f").rstrip("0").rstrip(".")


def build_reconciliation_workbook(
    customer: Customer,
    company: Company,
    start: "date",
    end: "date",
    period_caption: str,
    show_amounts: bool,
) -> BytesIO:
    co = company
    cust = customer

    rows = (
        db.session.query(DeliveryItem, Delivery, OrderItem, SalesOrder)
        .join(Delivery, DeliveryItem.delivery_id == Delivery.id)
        .join(OrderItem, DeliveryItem.order_item_id == OrderItem.id)
        .join(SalesOrder, DeliveryItem.order_id == SalesOrder.id)
        .options(
            joinedload(OrderItem.customer_product).joinedload(CustomerProduct.product)
        )
        .filter(Delivery.customer_id == cust.id)
        .filter(Delivery.delivery_date >= start)
        .filter(Delivery.delivery_date <= end)
        .order_by(Delivery.delivery_date, Delivery.id, DeliveryItem.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"对账_{cust.customer_code}"

    def merge_row(r, text, font_size=14, bold=False, row_height=None):
        ws.merge_cells(
            start_row=r, start_column=1, end_row=r, end_column=COL_LAST
        )
        c = ws.cell(r, 1, text)
        c.font = Font(name=FONT_SONG, size=font_size, bold=bold)
        c.alignment = center
        if row_height is not None:
            ws.row_dimensions[r].height = row_height

    merge_row(1, co.name or "", 18, False, 22.5)
    merge_row(2, period_caption, 16, False, 20.25)
    for c in range(1, COL_LAST + 1):
        ws.cell(1, c).border = all_border
        ws.cell(2, c).border = all_border

    font_info = Font(name=FONT_SONG, size=11, bold=False)

    def pair_row(r, lk, lv, rk, rv):
        ws.cell(r, 1, lk).alignment = left
        ws.cell(r, 1).font = font_info
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(r, 2, lv or "").alignment = left
        ws.cell(r, 2).font = font_info
        ws.cell(r, 6, rk).alignment = left
        ws.cell(r, 6).font = font_info
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=COL_LAST)
        ws.cell(r, 7, rv or "").alignment = left
        ws.cell(r, 7).font = font_info

    r = 3
    pair_row(r, "客户：", cust.name, "供应商：", co.name)
    for c in range(1, COL_LAST + 1):
        ws.cell(r, c).border = all_border
    r += 1
    pair_row(r, "联系人：", cust.contact or "", "联系人：", co.contact_person or "")
    for c in range(1, COL_LAST + 1):
        ws.cell(r, c).border = all_border
    r += 1
    pair_row(r, "电话：", cust.phone or "", "电话：", co.phone or "")
    for c in range(1, COL_LAST + 1):
        ws.cell(r, c).border = all_border
    r += 1
    pair_row(r, "传真：", cust.fax or "", "传真：", co.fax or "")
    for c in range(1, COL_LAST + 1):
        ws.cell(r, c).border = all_border
    r += 1
    pair_row(r, "地址：", cust.address or "", "地址：", co.address or "")
    for c in range(1, COL_LAST + 1):
        ws.cell(r, c).border = all_border
    r += 1

    font_table = Font(name=FONT_SONG, size=10, bold=False)
    headers = [
        "日期",
        "单据编号",
        "料号",
        "品名",
        "数量",
        "单位",
        "单价",
        "销售金额",
        "订单单号",
        "备注",
    ]
    hdr_row = r
    ws.row_dimensions[hdr_row].height = 24
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hdr_row, col, h)
        cell.font = font_table
        cell.border = all_border
        cell.alignment = center
    r += 1

    total_sales = Decimal("0")
    for di, dlv, oi, so in rows:
        price = Decimal(str(oi.price)) if oi.price is not None else Decimal("0")
        qty = Decimal(str(di.quantity)) if di.quantity is not None else Decimal("0")
        line_amt = (price * qty).quantize(Decimal("0.01"))
        if show_amounts:
            total_sales += line_amt

        mat = _material_no(oi)
        pname = (di.product_name or oi.product_name or "") or ""
        remark = (dlv.remark or "") or ""

        ws.cell(r, 1, dlv.delivery_date.isoformat() if dlv.delivery_date else "")
        ws.cell(r, 2, dlv.delivery_no or "")
        ws.cell(r, 3, mat)
        ws.cell(r, 4, pname)
        ws.cell(r, 5, _qty_display(di.quantity))
        ws.cell(r, 6, di.unit or oi.unit or "")
        if show_amounts:
            ws.cell(r, 7, float(price))
            ws.cell(r, 8, float(line_amt))
        else:
            ws.cell(r, 7, "")
            ws.cell(r, 8, "")
        ws.cell(r, 9, (so.customer_order_no or "").strip() or (so.order_no or ""))
        ws.cell(r, 10, remark)
        ws.row_dimensions[r].height = 24
        for c in range(1, COL_LAST + 1):
            ws.cell(r, c).border = all_border
            ws.cell(r, c).font = font_table
            if c in (1, 2, 3, 4, 6, 9, 10):
                ws.cell(r, c).alignment = left
            else:
                ws.cell(r, c).alignment = center
        r += 1

    if not rows:
        ws.row_dimensions[r].height = 24
        for c in range(1, COL_LAST + 1):
            cell = ws.cell(r, c, "")
            cell.border = all_border
            cell.font = font_table
        r += 1

    sum_row = r
    ws.row_dimensions[sum_row].height = 23
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
    for c in range(1, 7):
        ws.cell(sum_row, c).border = all_border
        ws.cell(sum_row, c).font = font_table
    ws.cell(sum_row, 7, "未税合计").font = font_table
    ws.cell(sum_row, 7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(sum_row, 7).border = all_border
    if show_amounts:
        ws.cell(sum_row, 8, float(total_sales))
    else:
        ws.cell(sum_row, 8, "")
    ws.cell(sum_row, 8).font = font_table
    ws.cell(sum_row, 8).border = all_border
    ws.merge_cells(start_row=sum_row, start_column=9, end_row=sum_row, end_column=COL_LAST)
    ws.cell(sum_row, 9).border = all_border
    ws.cell(sum_row, 9).font = font_table
    r += 1

    acc_lines = []
    if co.private_account:
        acc_lines.append(f"对私账户：{co.private_account}")
    if co.public_account:
        acc_lines.append(f"对公账户：{co.public_account}")
    if co.account_name:
        acc_lines.append(f"户名：{co.account_name}")
    if co.bank_name:
        acc_lines.append(f"开户行：{co.bank_name}")
    font_account = Font(name=FONT_SONG, size=11, bold=True)
    font_note = Font(name=FONT_SONG, size=10, bold=True)
    font_footer = Font(name=FONT_SONG, size=10, bold=False)
    if acc_lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COL_LAST)
        c = ws.cell(r, 1, "      ".join(acc_lines))
        c.alignment = left
        c.font = font_account
        ws.row_dimensions[r].height = 24
        for col in range(1, COL_LAST + 1):
            ws.cell(r, col).border = all_border
        r += 1

    for note in FOOTER_NOTES:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COL_LAST)
        c = ws.cell(r, 1, note)
        c.alignment = center
        c.font = font_note
        ws.row_dimensions[r].height = 27
        for col in range(1, COL_LAST + 1):
            ws.cell(r, col).border = all_border
        r += 1

    r += 1
    ws.row_dimensions[r].height = 16
    ws.cell(r, 1, "客户回签：").font = font_footer
    ws.cell(r, 4, "核准：").font = font_footer
    ws.cell(r, 7, "审核：").font = font_footer
    ws.cell(r, 9, "制表：" + (co.preparer_name or "")).font = font_footer

    for col in range(1, COL_LAST + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = (
            COL_WIDTHS[col - 1]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
