"""送货单 Excel：按「主体-付款方式」分 sheet，同 sheet 内多送货单段落用空行分隔；未交数量=按时间顺序累计至该行出货后的剩余。"""
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy.orm import joinedload
from openpyxl.worksheet.worksheet import Worksheet

from app import db
from app.models import Delivery, DeliveryItem, OrderItem, SalesOrder
from app.models.product import CustomerProduct
from app.utils.payment_type import normalize_payment_type, payment_type_label
from app.services.delivery_svc import effective_customer_material_no

COLS = 8
ADDRESS_LINE = "深圳市光明新区公明马山头旭发科技园A3栋1楼、2楼"
TEL_LINE = "TEL: 0755-29543726  15989505220"
FOOT_NOTES = (
    "1. 如以上货品有异常，请7天内于我司业务人员联系，换货或退货，过期恕不受理。",
    "2. 若因贵司付款没有准时，我司可以停止供货，本单货品所有权仍属我司。",
)

thin = Side(style="thin")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
font_song_20 = Font(name="宋体", size=20, bold=False)
font_song_20_bold = Font(name="宋体", size=20, bold=True)
font_song_12 = Font(name="宋体", size=12, bold=False)
font_song_10 = Font(name="宋体", size=10, bold=False)

_PAYMENT_TYPE_SORT = {"monthly": 0, "cash": 1}


def _safe_sheet_title(name: str, used: set) -> str:
    s = re.sub(r"[\[\]\*\/\\\?\:]", "_", (name or "sheet")[:31])
    if not s:
        s = "sheet"
    base = s
    n = 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = (
            (base[: 31 - len(suffix)] + suffix)
            if len(base) + len(suffix) > 31
            else base + suffix
        )
    used.add(s)
    return s


def _qty_str(q) -> str:
    if q is None:
        return ""
    d = Decimal(str(q))
    if d == d.to_integral_value():
        return str(int(d))
    return format(float(d), ".4f").rstrip("0").rstrip(".")


def _apply_column_widths(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 13.5
    ws.column_dimensions["B"].width = 14.625
    ws.column_dimensions["C"].width = 40.5
    ws.column_dimensions["D"].width = 9.25
    ws.column_dimensions["E"].width = 13.0
    ws.column_dimensions["F"].width = 13.0
    ws.column_dimensions["G"].width = 13.0
    ws.column_dimensions["H"].width = 10.625


def _payment_type_sort_key(pt: str) -> Tuple[int, str]:
    return (_PAYMENT_TYPE_SORT.get(pt, 50), pt or "")


def _write_delivery_section(
    ws: Worksheet,
    start_row: int,
    delivery: Delivery,
    di_list: List[DeliveryItem],
    oi_map: Dict[int, OrderItem],
    order_cache: Dict[int, Optional[SalesOrder]],
    remaining_after: Dict[int, float],
) -> int:
    """
    从 start_row 起写入一个送货单段落（表头+明细+脚注+签栏）。
    返回本段最后一行（签栏所在行）的行号。
    """
    company = (
        delivery.customer.company
        if delivery.customer and delivery.customer.company
        else None
    )
    company_name = company.name if company else ""
    cust = delivery.customer.name if delivery.customer else ""
    waybill = delivery.waybill_no or "送货"
    dno = delivery.delivery_no or ""
    dstr = (
        delivery.delivery_date.strftime("%Y%m%d") if delivery.delivery_date else ""
    )

    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    c = ws.cell(r, 1, company_name)
    c.font = font_song_20
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 29.1
    r += 1

    for text in (ADDRESS_LINE, TEL_LINE):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
        cell = ws.cell(r, 1, text)
        cell.font = font_song_12
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 21.0
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    t = ws.cell(r, 1, "送  货  单")
    t.font = font_song_20_bold
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 32.1
    r += 1

    ws.cell(r, 1, "客户")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2, cust)
    ws.cell(r, 5, "单号")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    ws.cell(r, 6, dno)
    for col in (1, 2, 5, 6):
        ws.cell(r, col).font = font_song_10
    ws.row_dimensions[r].height = 24.0
    r += 1

    ws.cell(r, 1, "快递单号")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2, waybill)
    ws.cell(r, 5, "日期")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
    ws.cell(r, 6, dstr)
    for col in (1, 2, 5, 6):
        ws.cell(r, col).font = font_song_10
    ws.row_dimensions[r].height = 24.0
    r += 1

    headers = (
        "料号",
        "物料编号",
        "品名规格",
        "订购数量",
        "出货数量",
        "未交数量",
        "单位",
        "客户订单号",
    )
    hr = r
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hr, col, h)
        cell.font = font_song_10
        cell.border = all_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[hr].height = 27.0
    r = hr + 1

    for di in di_list:
        oi = oi_map.get(di.order_item_id)
        if not oi:
            continue
        oid = oi.order_id
        if oid not in order_cache:
            order_cache[oid] = db.session.get(SalesOrder, oid)
        so = order_cache[oid]

        product_code = None
        material_no = ""
        if oi.customer_product and oi.customer_product.product:
            product_code = oi.customer_product.product.product_code
            material_no = product_code or ""

        liao = effective_customer_material_no(oi)
        name_spec = (
            " ".join(
                x for x in (oi.product_name or "", oi.product_spec or "") if x
            ).strip()
            or (oi.product_name or "")
        )

        remaining = remaining_after.get(di.id, 0.0)
        order_cell = (
            (so.customer_order_no or "").strip()
            if so
            else ""
        ) or ((so.order_no or "").strip() if so else "")

        vals = (
            liao,
            material_no,
            name_spec,
            _qty_str(oi.quantity),
            _qty_str(di.quantity),
            _qty_str(remaining),
            di.unit or oi.unit or "",
            order_cell,
        )
        for col, val in enumerate(vals, 1):
            cell = ws.cell(r, col, val)
            cell.font = font_song_12
            cell.border = all_border
            if 4 <= col <= 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        ws.row_dimensions[r].height = 27.0
        r += 1

    empty_rows = max(0, 6 - len(di_list))
    for _ in range(empty_rows):
        for col in range(1, COLS + 1):
            cell = ws.cell(r, col, "")
            cell.font = font_song_12
            cell.border = all_border
        ws.row_dimensions[r].height = 27.0
        r += 1

    foot_row = r
    for note in FOOT_NOTES:
        ws.merge_cells(
            start_row=foot_row, start_column=1, end_row=foot_row, end_column=COLS
        )
        cell = ws.cell(foot_row, 1, note)
        cell.font = font_song_12
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.row_dimensions[foot_row].height = 21.0
        foot_row += 1

    foot_row += 1
    ws.merge_cells(
        start_row=foot_row, start_column=1, end_row=foot_row, end_column=COLS
    )
    sign_cell = ws.cell(
        foot_row,
        1,
        "制单人：________________    仓库：________________    客户签收：________________",
    )
    sign_cell.font = font_song_12
    ws.row_dimensions[foot_row].height = 21.0
    return foot_row


def build_delivery_notes_workbook(deliveries: List[Delivery]) -> Optional[BytesIO]:
    if not deliveries:
        return None

    deliveries_sorted = sorted(
        deliveries,
        key=lambda d: (d.delivery_date or date.min, d.id),
    )

    all_pairs: List[Tuple[Delivery, DeliveryItem]] = []
    for d in deliveries_sorted:
        items = (
            DeliveryItem.query.filter_by(delivery_id=d.id)
            .order_by(DeliveryItem.id)
            .all()
        )
        for di in items:
            all_pairs.append((d, di))

    if not all_pairs:
        return None

    oi_ids = {di.order_item_id for _, di in all_pairs}
    order_items = (
        OrderItem.query.filter(OrderItem.id.in_(oi_ids))
        .options(
            joinedload(OrderItem.customer_product).joinedload(CustomerProduct.product)
        )
        .all()
    )
    oi_map = {oi.id: oi for oi in order_items}

    delivered_running: Dict[int, float] = defaultdict(float)
    remaining_after: Dict[int, float] = {}
    for _, di in all_pairs:
        oi = oi_map.get(di.order_item_id)
        if not oi:
            continue
        need = float(oi.quantity or 0)
        q = float(di.quantity or 0)
        delivered_running[oi.id] += q
        remaining_after[di.id] = max(0.0, need - delivered_running[oi.id])

    order_cache: Dict[int, Optional[SalesOrder]] = {}

    wb = openpyxl.Workbook()
    used_titles: set = set()
    sheet_ws: Dict[Tuple[int, str], Worksheet] = {}
    sheet_next_row: Dict[Tuple[int, str], int] = {}
    first_sheet = True

    for d in deliveries_sorted:
        items = (
            DeliveryItem.query.filter_by(delivery_id=d.id)
            .order_by(DeliveryItem.id)
            .all()
        )
        if not items:
            continue

        buckets: Dict[str, List[DeliveryItem]] = defaultdict(list)
        for di in items:
            oi = oi_map.get(di.order_item_id)
            if not oi:
                continue
            oid = oi.order_id
            if oid not in order_cache:
                order_cache[oid] = db.session.get(SalesOrder, oid)
            so = order_cache[oid]
            pt = normalize_payment_type(so.payment_type if so else None)
            buckets[pt].append(di)

        company = (
            d.customer.company if d.customer and d.customer.company else None
        )
        cid = company.id if company else 0

        for pt in sorted(buckets.keys(), key=_payment_type_sort_key):
            dis = buckets[pt]
            if not dis:
                continue
            key = (cid, pt)
            if key not in sheet_ws:
                if first_sheet:
                    wb.remove(wb.active)
                    first_sheet = False
                cname = company.name if company else "未知主体"
                title = _safe_sheet_title(f"{cname}-{payment_type_label(pt)}", used_titles)
                ws = wb.create_sheet(title=title)
                sheet_ws[key] = ws
                sheet_next_row[key] = 1
                _apply_column_widths(ws)

            ws = sheet_ws[key]
            start_row = sheet_next_row[key]
            # 单个表格最多 6 条明细，超出则分页分段
            chunk_size = 6
            last_row = start_row - 1
            for i in range(0, len(dis), chunk_size):
                chunk = dis[i : i + chunk_size]
                last_row = _write_delivery_section(
                    ws, start_row, d, chunk, oi_map, order_cache, remaining_after
                )
                start_row = last_row + 2
            sheet_next_row[key] = start_row

    if first_sheet:
        return None

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
