from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.utils.decimal_scale import quantize_decimal


TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "static"
    / "templates"
    / "procurement"
    / "purchase_order_template.xlsx"
)


def _decimal_to_number(v):
    if v is None:
        return None
    d = quantize_decimal(v)
    if d == d.to_integral():
        return int(d)
    return float(d)


def _date_text(v) -> str:
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def build_purchase_order_workbook(purchase_order) -> BytesIO:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    if getattr(wb, "calculation", None) is not None:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    company_name = ""
    if purchase_order.company and purchase_order.company.name:
        company_name = purchase_order.company.name
    supplier_name = (purchase_order.supplier_name or "").strip()
    supplier_contact = (purchase_order.supplier_contact_name or "").strip()
    supplier_phone = (purchase_order.supplier_phone or "").strip()
    supplier_address = (purchase_order.supplier_address or "").strip()
    order_date = purchase_order.ordered_at or purchase_order.created_at

    ws["A1"] = company_name
    ws["B3"] = supplier_name
    ws["B4"] = supplier_contact
    ws["B5"] = supplier_phone
    ws["B6"] = supplier_address
    ws["J3"] = purchase_order.po_no or ""
    ws["J4"] = _date_text(order_date)
    ws["J5"] = ""

    ws["B8"] = purchase_order.item_name or ""
    ws["D8"] = purchase_order.item_spec or ""
    ws["E8"] = purchase_order.unit or ""
    ws["F8"] = _decimal_to_number(purchase_order.qty)
    ws["G8"] = _decimal_to_number(purchase_order.unit_price)
    ws["H8"] = "=G8*F8"
    ws["I8"] = _date_text(purchase_order.expected_date)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
