"""送货记录 Excel：按「主体-付款方式」分 sheet，每行一条送货明细；备注列=客户订单号。"""
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import joinedload

from app import db
from app.models import Customer, Delivery, DeliveryItem, OrderItem, SalesOrder
from app.models.product import CustomerProduct
from app.utils.payment_type import normalize_payment_type, payment_type_label
from app.services.delivery_svc import effective_customer_material_no

thin = Side(style="thin")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(bold=True)
_PAYMENT_TYPE_SORT = {"monthly": 0, "cash": 1}


def _safe_sheet_title(name: str, used: set) -> str:
    s = re.sub(r"[\[\]\*\/\\\?\:]", "_", (name or "sheet")[:31])
    if not s:
        s = "sheet"
    base = s
    n = 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = (
            (base[: 31 - len(suffix)] + suffix)
            if len(base) + len(suffix) > 31
            else base + suffix
        )
    used.add(s)
    return s


def _qty_str(q) -> str:
    if q is None:
        return ""
    d = Decimal(str(q))
    if d == d.to_integral_value():
        return str(int(d))
    return format(float(d), ".4f").rstrip("0").rstrip(".")


def _payment_type_sort_key(pt: str) -> Tuple[int, str]:
    return (_PAYMENT_TYPE_SORT.get(pt, 50), pt or "")


def _name_spec(oi: OrderItem, di: DeliveryItem) -> str:
    n = (di.product_name or oi.product_name or "") or ""
    sp = oi.product_spec or ""
    return " ".join(x for x in (n, sp) if x).strip() or n


def build_delivery_records_workbook(
    d_from: date,
    d_to: date,
    company_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    payment_type: Optional[str] = None,
) -> Optional[BytesIO]:
    q = (
        db.session.query(DeliveryItem, Delivery, OrderItem, SalesOrder, Customer)
        .join(Delivery, DeliveryItem.delivery_id == Delivery.id)
        .join(OrderItem, DeliveryItem.order_item_id == OrderItem.id)
        .join(SalesOrder, DeliveryItem.order_id == SalesOrder.id)
        .join(Customer, Delivery.customer_id == Customer.id)
        .options(
            joinedload(OrderItem.customer_product).joinedload(CustomerProduct.product),
            joinedload(Customer.company),
        )
        .filter(Delivery.delivery_date >= d_from)
        .filter(Delivery.delivery_date <= d_to)
    )
    if company_id:
        q = q.filter(Customer.company_id == company_id)
    if customer_id:
        q = q.filter(Delivery.customer_id == customer_id)
    rows = q.order_by(
        Delivery.delivery_date, Delivery.id, DeliveryItem.id
    ).all()

    if payment_type:
        pt_norm = normalize_payment_type(payment_type)
        rows = [r for r in rows if normalize_payment_type(r[3].payment_type) == pt_norm]

    if not rows:
        return None

    # 计算“未交数量”：与送货单导出一致，按 (送货日期 -> 送货单id -> 明细id)
    # 的顺序对同一 order_item_id 累计已出货量，remaining_after[delivery_item_id]
    # 表示“处理到当前这条出货明细之后”的剩余未交。
    delivered_running = defaultdict(lambda: Decimal("0"))
    remaining_after: Dict[int, Decimal] = {}
    for di, dlv, oi, so, cust in rows:
        need = Decimal(str(oi.quantity or 0))
        q = Decimal(str(di.quantity or 0))
        delivered_running[oi.id] += q
        left = need - delivered_running[oi.id]
        remaining_after[di.id] = left if left > 0 else Decimal("0")

    buckets: Dict[Tuple[int, str], List] = defaultdict(list)
    for di, dlv, oi, so, cust in rows:
        pt = normalize_payment_type(so.payment_type)
        cid = cust.company_id or 0
        buckets[(cid, pt)].append((di, dlv, oi, so, cust))

    wb = openpyxl.Workbook()
    used_titles: set = set()
    first = True

    for key in sorted(buckets.keys(), key=lambda k: (k[0], _payment_type_sort_key(k[1]))):
        chunk = buckets[key]
        cid, pt = key
        cust0 = chunk[0][4]
        company = cust0.company if cust0 and cust0.company else None
        cname = company.name if company else "未知主体"
        title = _safe_sheet_title(f"{cname}-{payment_type_label(pt)}", used_titles)
        if first:
            wb.remove(wb.active)
            first = False
        ws = wb.create_sheet(title=title)
        _write_records_sheet(ws, chunk, remaining_after)

    if first:
        return None

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _write_records_sheet(
    ws: Worksheet, chunk: List, remaining_after: Dict[int, Decimal]
) -> None:
    headers = (
        "客户",
        "送货日期",
        "送货单号",
        "料号",
        "公司料号",
        "物料编号",
        "品名规格",
        "订单数量",
        "此次出货数量",
        "未交数量",
        "单位",
        "单价",
        "金额",
        "备注",
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = header_font
        c.border = all_border
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    r = 2
    for di, dlv, oi, so, cust in chunk:
        delivery_date_s = (
            dlv.delivery_date.strftime("%Y%m%d") if dlv.delivery_date else ""
        )
        remark = (so.customer_order_no or "").strip() or ""

        customer_liao = effective_customer_material_no(oi)

        company_liao = ""
        if oi.customer_product and oi.customer_product.product:
            company_liao = oi.customer_product.product.product_code or ""

        material_no = ""
        if oi.customer_product and oi.customer_product.product:
            material_no = oi.customer_product.product.product_code or ""

        product_name_spec = _name_spec(oi, di)

        order_qty = Decimal(str(oi.quantity or 0))
        delivery_qty = Decimal(str(di.quantity or 0))
        remaining_qty = remaining_after.get(di.id, Decimal("0"))

        is_sample = bool(getattr(oi, "is_sample", False))
        price = Decimal(str(oi.price or 0)) if not is_sample else Decimal("0")
        amount = (price * delivery_qty).quantize(Decimal("0.01"))

        vals = (
            cust.name or "",
            delivery_date_s,
            dlv.delivery_no or "",
            customer_liao,
            company_liao,
            material_no,
            product_name_spec,
            _qty_str(order_qty),
            _qty_str(delivery_qty),
            _qty_str(remaining_qty),
            di.unit or oi.unit or "",
            _qty_str(price),
            float(amount),
            remark,
        )

        for col, val in enumerate(vals, 1):
            cell = ws.cell(r, col, val)
            cell.border = all_border
            # 文字字段靠左，其它居中；金额居右（更像报表）
            if col == 13:
                cell.alignment = Alignment(
                    horizontal="right", vertical="center", wrap_text=True
                )
            elif col in (4, 5, 6, 7, 14):
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        r += 1

    # 列宽（按模板常见比例设置）
    widths = {
        1: 20,  # 客户
        2: 10,  # 送货日期
        3: 14,  # 送货单号
        4: 14,  # 料号
        5: 14,  # 公司料号
        6: 14,  # 物料编号
        7: 24,  # 品名规格
        8: 12,  # 订单数量
        9: 12,  # 此次出货数量
        10: 12,  # 未交数量
        11: 10,  # 单位
        12: 10,  # 单价
        13: 16,  # 金额
        14: 18,  # 备注
    }
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 12)
