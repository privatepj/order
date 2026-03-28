from decimal import Decimal, InvalidOperation

from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required

from app.auth.capabilities import customer_list_read_filters
from app.auth.decorators import capability_required, menu_required

from app import db
from app.models import Customer, Company
from app.utils.query import keyword_like_or, cast_str
from app.utils.visibility import is_admin, customer_view
from sqlalchemy.exc import IntegrityError
from io import BytesIO


def register_customer_routes(bp):
    @bp.route("/customers")
    @login_required
    @menu_required("customer")
    def customer_list():
        page = request.args.get("page", 1, type=int)
        keyword = customer_list_read_filters()
        q = Customer.query
        if keyword:
            q = q.outerjoin(Company, Customer.company_id == Company.id)
            cond = keyword_like_or(
                keyword,
                Customer.customer_code,
                Customer.short_code,
                Customer.name,
                Customer.contact,
                Customer.phone,
                Customer.fax,
                Customer.address,
                Customer.payment_terms,
                Customer.remark,
                cast_str(Customer.tax_point),
                Company.name,
                Company.code,
            )
            if cond is not None:
                q = q.filter(cond)
            q = q.distinct()
        q = q.order_by(Customer.customer_code)
        pagination = q.paginate(page=page, per_page=20)
        rows = [customer_view(c) for c in pagination.items]
        return render_template(
            "customer/list.html",
            pagination=pagination,
            rows=rows,
            keyword=keyword,
        )

    @bp.route("/customers/new", methods=["GET", "POST"])
    @login_required
    @menu_required("customer")
    @capability_required("customer.action.create")
    def customer_new():
        companies = Company.query.order_by(Company.id).all()
        if request.method == "POST":
            return _customer_save(None, companies)
        return render_template(
            "customer/form.html",
            customer=None,
            customer_view=customer_view(None),
            companies=companies,
        )

    @bp.route("/customers/<int:customer_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("customer")
    @capability_required("customer.action.edit")
    def customer_edit(customer_id):
        customer = Customer.query.get_or_404(customer_id)
        companies = Company.query.order_by(Company.id).all()
        if request.method == "POST":
            return _customer_save(customer, companies)
        return render_template(
            "customer/form.html",
            customer=customer,
            customer_view=customer_view(customer),
            companies=companies,
        )

    @bp.route("/customers/<int:customer_id>/delete", methods=["POST"])
    @login_required
    @menu_required("customer")
    @capability_required("customer.action.delete")
    def customer_delete(customer_id):
        customer = Customer.query.get_or_404(customer_id)
        db.session.delete(customer)
        db.session.commit()
        flash("客户已删除。", "success")
        return redirect(url_for("main.customer_list"))

    @bp.route("/customers/import", methods=["GET", "POST"])
    @login_required
    @menu_required("customer")
    @capability_required("customer.action.import")
    def customer_import():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("customer/import.html", result=None)

            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("customer/import.html", result=None)
            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("customer/import.html", result=None)

            default_company_id = db.session.query(db.func.min(Company.id)).scalar()
            if not default_company_id:
                flash("系统中无经营主体，请先初始化公司数据。", "danger")
                return render_template("customer/import.html", result=None)

            # 读取表头（第 1 行），用于按列名取值；若缺少关键表头则回退到旧的固定前 5 列读取。
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True), None
            ) or ()
            header_map = {}
            for i, h in enumerate(header_row):
                if isinstance(h, str):
                    k = h.strip()
                    if k:
                        header_map[k] = i

            def _get(row, idx):
                return row[idx] if idx is not None and idx < len(row) else None

            required_headers = {"客户名称", "主体", "税点", "客户短码"}
            has_header_mode = required_headers.issubset(set(header_map.keys()))

            max_tries = 3
            last_result = {"success": 0, "errors": []}
            for attempt in range(max_tries):
                try:
                    success = 0
                    errors = []

                    for idx, row in enumerate(
                        ws.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        # row 可能比我们索引少很多列，所以统一做一下扩展再读。
                        row = tuple(row or ())

                        if has_header_mode:
                            name = _get(row, header_map.get("客户名称"))
                            contact = _get(row, header_map.get("联系人"))
                            phone = _get(row, header_map.get("电话"))
                            address = _get(row, header_map.get("地址"))
                            remark = _get(row, header_map.get("备注"))
                            tax_point_raw = _get(row, header_map.get("税点"))
                            subject_name = _get(row, header_map.get("主体"))
                            short_code = _get(row, header_map.get("客户短码"))
                        else:
                            # 回退旧模板：客户名称/联系人/电话/地址/备注（前 5 列）
                            name, contact, phone, address, remark = (row + (None,) * 5)[:5]
                            tax_point_raw = None
                            subject_name = None
                            short_code = None

                        name = (name or "").strip() if isinstance(name, str) else (name or "")
                        if not name:
                            if any(row):
                                errors.append(f"第 {idx} 行：客户名称为空")
                            continue

                        # 主体：主体为空用默认；否则按 Company.name 查找
                        company_id = default_company_id
                        if has_header_mode:
                            if subject_name is None or (
                                isinstance(subject_name, str) and not subject_name.strip()
                            ):
                                company_id = default_company_id
                            else:
                                subject_name_s = (
                                    subject_name.strip()
                                    if isinstance(subject_name, str)
                                    else str(subject_name)
                                )
                                company = Company.query.filter_by(name=subject_name_s).first()
                                if not company:
                                    errors.append(f"第 {idx} 行：主体不存在：{subject_name_s}")
                                    continue
                                company_id = company.id

                        # 税点
                        tax_point = None
                        if has_header_mode:
                            if tax_point_raw is not None and not (
                                isinstance(tax_point_raw, str) and not tax_point_raw.strip()
                            ):
                                try:
                                    if isinstance(tax_point_raw, str):
                                        t = tax_point_raw.strip()
                                        # 兼容少量可能的 "13%" 输入
                                        if t.endswith("%"):
                                            tax_point = Decimal(t[:-1].strip()) / Decimal("100")
                                        else:
                                            tax_point = Decimal(t)
                                    else:
                                        tax_point = Decimal(str(tax_point_raw))
                                except Exception:
                                    errors.append(f"第 {idx} 行：税点格式不正确：{tax_point_raw}")
                                    continue

                        # 客户短码
                        short_code = (
                            (short_code or "").strip()
                            if isinstance(short_code, str)
                            else (str(short_code).strip() if short_code is not None else "")
                        ) or None

                        # 已存在判定：客户名称 + 主体(company_id)
                        existing = Customer.query.filter_by(name=name, company_id=company_id).first()
                        if existing:
                            continue

                        customer = Customer(
                            customer_code=_next_customer_code(),
                            name=name,
                            contact=(contact or "").strip() if isinstance(contact, str) else contact,
                            phone=(phone or "").strip() if isinstance(phone, str) else phone,
                            address=(address or "").strip() if isinstance(address, str) else address,
                            remark=(remark or "").strip() if isinstance(remark, str) else remark,
                            tax_point=tax_point,
                            short_code=short_code,
                            company_id=company_id,
                        )
                        db.session.add(customer)
                        success += 1

                    db.session.commit()
                    last_result = {"success": success, "errors": errors}
                    break
                except IntegrityError:
                    db.session.rollback()
                    db.session.expunge_all()
                    if attempt == max_tries - 1:
                        flash("导入失败：客户编码冲突，请稍后重试。", "danger")
                        return render_template("customer/import.html", result=None)

            result = last_result
            if result.get("success"):
                flash(f"成功导入 {result['success']} 条客户。", "success")
            if result.get("errors"):
                flash(f"有 {len(result['errors'])} 条记录导入失败，请查看错误列表。", "danger")
            return render_template("customer/import.html", result=result)

        return render_template("customer/import.html", result=None)

    @bp.route("/customers/export-import-template", methods=["GET"])
    @login_required
    @menu_required("customer")
    def export_customer_import_template():
        """
        客户导入模板（xlsx）：表头行+1行空白。
        表头列名需要与 `customer_import()` 的精确匹配保持一致。
        """
        from openpyxl import Workbook

        headers = [
            "客户名称",
            "联系人",
            "电话",
            "地址",
            "备注",
            "税点",
            "主体",
            "客户短码",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "客户导入模板"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="客户导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _next_customer_code():
        # 兼容历史：可能存在小写 `c0001` 等编码。
        # 同时做“候选存在则递增”兜底，避免并发/历史数据导致的冲突。
        q = (
            db.session.query(Customer.customer_code)
            .filter(
                db.or_(
                    Customer.customer_code.like("C%"),
                    Customer.customer_code.like("c%"),
                )
            )
        )
        max_num = 0
        for (code,) in q.all():
            if not code or len(code) < 2:
                continue
            if code[0] not in ("C", "c"):
                continue
            try:
                max_num = max(max_num, int(code[1:]))
            except ValueError:
                continue

        next_num = max_num + 1
        while True:
            candidate = f"C{next_num:04d}"
            exists = Customer.query.filter_by(customer_code=candidate).first()
            if not exists:
                return candidate
            next_num += 1

    def _customer_save(customer, companies):
        name = (request.form.get("name") or "").strip()
        raw_ids = request.form.getlist("company_ids")
        company_ids = []
        for x in raw_ids:
            try:
                v = int(x)
            except (TypeError, ValueError):
                continue
            if v and v not in company_ids:
                company_ids.append(v)
        if not name:
            flash("客户名称为必填。", "danger")
            return render_template(
                "customer/form.html",
                customer=customer or Customer(name=name),
                customer_view=customer_view(customer),
                companies=companies,
            )
        if not company_ids:
            flash("请选择经营主体。", "danger")
            return render_template(
                "customer/form.html",
                customer=customer or Customer(name=name),
                customer_view=customer_view(customer),
                companies=companies,
            )
        payment_terms = (request.form.get("payment_terms") or "").strip() or None
        if payment_terms:
            low = payment_terms.strip().lower()
            if low == "monthly":
                payment_terms = "月结"
            elif low == "cash":
                payment_terms = "现金"

        def _apply_fields(cust: Customer, co_id: int) -> None:
            cust.name = name
            cust.company_id = co_id
            sc = (request.form.get("short_code") or "").strip() or None
            cust.short_code = sc
            cust.address = (request.form.get("address") or "").strip() or None
            cust.payment_terms = payment_terms
            cust.remark = (request.form.get("remark") or "").strip() or None
            if is_admin():
                cust.contact = (request.form.get("contact") or "").strip() or None
                cust.phone = (request.form.get("phone") or "").strip() or None
                cust.fax = (request.form.get("fax") or "").strip() or None
                tp = (request.form.get("tax_point") or "").strip()
                if tp:
                    try:
                        cust.tax_point = Decimal(tp)
                    except (InvalidOperation, ValueError):
                        cust.tax_point = None
                else:
                    cust.tax_point = None

        created = 0
        if customer is None:
            new_specs: list[tuple[Customer, int]] = []
            for co_id in company_ids:
                c = Customer()
                c.customer_code = _next_customer_code()
                _apply_fields(c, co_id)
                db.session.add(c)
                new_specs.append((c, co_id))
                created += 1

            max_tries = 3
            for attempt in range(max_tries):
                try:
                    db.session.commit()
                    break
                except IntegrityError:
                    db.session.rollback()
                    db.session.expunge_all()
                    if attempt == max_tries - 1:
                        flash("保存失败：客户编码冲突，请稍后重试。", "danger")
                        return redirect(url_for("main.customer_list"))
                    for c, co_id in new_specs:
                        c.customer_code = _next_customer_code()
                        _apply_fields(c, co_id)
                        db.session.add(c)
            flash(f"客户已新增（共 {created} 条）。", "success")
            return redirect(url_for("main.customer_list"))

        # 编辑：更新当前记录为第一个主体，其余主体复制新增
        _apply_fields(customer, company_ids[0])
        db.session.add(customer)
        new_specs: list[tuple[Customer, int]] = []
        for co_id in company_ids[1:]:
            c = Customer()
            c.customer_code = _next_customer_code()
            _apply_fields(c, co_id)
            db.session.add(c)
            new_specs.append((c, co_id))
            created += 1
        max_tries = 3
        for attempt in range(max_tries):
            try:
                db.session.commit()
                break
            except IntegrityError:
                db.session.rollback()
                db.session.expunge_all()
                if attempt == max_tries - 1:
                    flash("保存失败：客户编码冲突，请稍后重试。", "danger")
                    return redirect(url_for("main.customer_list"))
                _apply_fields(customer, company_ids[0])
                db.session.add(customer)
                for c, co_id in new_specs:
                    c.customer_code = _next_customer_code()
                    _apply_fields(c, co_id)
                    db.session.add(c)
        if created:
            flash(f"客户已保存，并新增 {created} 条（按多主体复制）。", "success")
        else:
            flash("客户已保存。", "success")
        return redirect(url_for("main.customer_list"))
