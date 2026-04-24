from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required

from app.auth.capabilities import product_list_read_filters
from app.auth.decorators import capability_required, menu_required

from app import db
from app.models import Product
from app.utils.query import keyword_like_or
from app.utils.form_display import clean_optional_text
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func
from io import BytesIO


def register_product_routes(bp):
    def _normalize_name_spec(name, spec):
        name_n = (name or "").strip()
        if spec is None:
            spec_n = ""
        elif isinstance(spec, str):
            spec_n = spec.strip()
        else:
            spec_n = str(spec).strip()
        return name_n, spec_n

    def _find_product_by_name_spec(name, spec):
        name_n, spec_n = _normalize_name_spec(name, spec)
        if not name_n:
            return None, None
        matches = (
            Product.query.filter(
                Product.name == name_n,
                func.coalesce(Product.spec, "") == spec_n,
            )
            .order_by(Product.id.asc())
            .all()
        )
        if len(matches) > 1:
            return None, "同品名+规格匹配到多条产品，请先清理主数据"
        return (matches[0] if matches else None), None

    @bp.route("/products")
    @login_required
    @menu_required("product")
    def product_list():
        page = request.args.get("page", 1, type=int)
        keyword = product_list_read_filters()
        q = Product.query
        cond = keyword_like_or(
            keyword,
            Product.product_code,
            Product.name,
            Product.spec,
            Product.series,
            Product.base_unit,
            Product.remark,
        )
        if cond is not None:
            q = q.filter(cond)
        q = q.order_by(Product.product_code)
        pagination = q.paginate(page=page, per_page=20)
        return render_template(
            "product/list.html",
            pagination=pagination,
            keyword=keyword,
        )

    @bp.route("/products/new", methods=["GET", "POST"])
    @login_required
    @menu_required("product")
    @capability_required("product.action.create")
    def product_new():
        if request.method == "POST":
            return _product_save(None)
        return render_template("product/form.html", product=None)

    @bp.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("product")
    @capability_required("product.action.edit")
    def product_edit(product_id):
        product = Product.query.get_or_404(product_id)
        if request.method == "POST":
            return _product_save(product)
        return render_template("product/form.html", product=product)

    @bp.route("/products/<int:product_id>/delete", methods=["POST"])
    @login_required
    @menu_required("product")
    @capability_required("product.action.delete")
    def product_delete(product_id):
        product = Product.query.get_or_404(product_id)
        db.session.delete(product)
        db.session.commit()
        flash("产品已删除。", "success")
        return redirect(url_for("main.product_list"))

    @bp.route("/products/export-import-template", methods=["GET"])
    @login_required
    @menu_required("product")
    @capability_required("product.action.import")
    def export_product_import_template():
        """产品导入模板（xlsx）：表头行+1行空白。"""
        from openpyxl import Workbook

        headers = ["产品编号", "产品名称", "规格", "基础单位", "备注", "系列"]

        wb = Workbook()
        ws = wb.active
        ws.title = "产品导入模板"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="产品导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/products/import", methods=["GET", "POST"])
    @login_required
    @menu_required("product")
    @capability_required("product.action.import")
    def product_import():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("product/import.html", result=None)
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("product/import.html", result=None)
            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("product/import.html", result=None)
            max_tries = 3
            for attempt in range(max_tries):
                try:
                    success = 0
                    errors = []
                    seen_name_spec = set()
                    for idx, row in enumerate(
                        ws.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        code, name, spec, base_unit, remark, series = (
                            row + (None,) * 6
                        )[:6]
                        code = (code or "").strip() if isinstance(code, str) else ""
                        name = (name or "").strip() if isinstance(name, str) else (name or "")
                        if not name:
                            if any(row):
                                errors.append(f"第 {idx} 行：产品名称为空")
                            continue
                        spec_clean = clean_optional_text(spec, max_len=128)
                        name_n, spec_n = _normalize_name_spec(name, spec_clean)
                        row_key = (name_n, spec_n)
                        if row_key in seen_name_spec:
                            errors.append(
                                f"第 {idx} 行：同一文件中品名+规格重复（{name_n}/{spec_n or '空规格'}）"
                            )
                            continue
                        seen_name_spec.add(row_key)

                        product_by_name, name_match_err = _find_product_by_name_spec(
                            name_n, spec_clean
                        )
                        if name_match_err:
                            errors.append(f"第 {idx} 行：{name_match_err}")
                            continue

                        product_by_code = (
                            Product.query.filter_by(product_code=code).first() if code else None
                        )
                        if product_by_code and product_by_name and product_by_code.id != product_by_name.id:
                            errors.append(
                                f"第 {idx} 行：产品编号与品名+规格命中不同记录，无法自动合并"
                            )
                            continue

                        product = product_by_code or product_by_name
                        if not product:
                            if not code:
                                code = _next_product_code()
                                while Product.query.filter_by(product_code=code).first():
                                    code = _bump_product_code(code)
                            product = Product(product_code=code)
                            db.session.add(product)

                        product.name = name_n
                        product.spec = spec_clean
                        product.base_unit = clean_optional_text(base_unit, max_len=16)
                        product.remark = clean_optional_text(remark, max_len=255)
                        product.series = clean_optional_text(series, max_len=64)
                        success += 1
                        db.session.flush()

                    db.session.commit()
                    result = {"success": success, "errors": errors}
                    if success:
                        flash(f"成功导入或更新 {success} 条产品。", "success")
                    if errors:
                        flash(f"有 {len(errors)} 条记录导入失败，请查看错误列表。", "danger")
                    return render_template("product/import.html", result=result)
                except IntegrityError:
                    db.session.rollback()
                    db.session.expunge_all()
                    if attempt == max_tries - 1:
                        flash("导入失败：产品编码冲突，请稍后重试。", "danger")
                        return render_template("product/import.html", result=None)
        return render_template("product/import.html", result=None)

    def _next_product_code():
        m = 0
        for (c,) in db.session.query(Product.product_code).filter(
            db.or_(
                Product.product_code.like("P%"),
                Product.product_code.like("p%"),
            )
        ).all():
            try:
                if not c or len(c) < 2:
                    continue
                m = max(m, int(c[1:]))
            except ValueError:
                pass
        return f"P{m + 1:04d}"

    def _bump_product_code(code):
        try:
            return f"P{int(code[1:]) + 1:04d}"
        except (ValueError, IndexError):
            return code + "N"

    def _product_save(product):
        name = (request.form.get("name") or "").strip()
        if not name:
            flash("产品名称为必填。", "danger")
            return render_template(
                "product/form.html",
                product=product or Product(name=name),
            )
        if product is None:
            product = Product()
            product.product_code = _next_product_code()
            while Product.query.filter_by(product_code=product.product_code).first():
                product.product_code = _bump_product_code(product.product_code)
        product.name = name
        product.spec = clean_optional_text(request.form.get("spec"), max_len=128)
        product.series = clean_optional_text(request.form.get("series"), max_len=64)
        product.base_unit = clean_optional_text(request.form.get("base_unit"), max_len=16)
        product.remark = clean_optional_text(request.form.get("remark"), max_len=255)
        max_tries = 3
        db.session.add(product)
        for attempt in range(max_tries):
            try:
                db.session.commit()
                flash("产品已保存。", "success")
                return redirect(url_for("main.product_list"))
            except IntegrityError:
                db.session.rollback()
                db.session.expunge_all()
                if product.id is None:
                    product.product_code = _next_product_code()
                    while Product.query.filter_by(product_code=product.product_code).first():
                        product.product_code = _bump_product_code(product.product_code)
                db.session.add(product)
                if attempt == max_tries - 1:
                    flash("保存失败：产品编码冲突，请稍后重试。", "danger")
                    return redirect(url_for("main.product_list"))
