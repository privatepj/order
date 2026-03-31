from io import BytesIO

from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from sqlalchemy import func

from app import db
from app.auth.decorators import capability_required, menu_required
from app.models import Delivery, ExpressCompany, ExpressWaybill
from app.utils.waybill_pool import apply_waybill_to_pool
from app.utils.waybill_range import expand_waybill_range
from app.utils.query import keyword_like_or

_INVALID_SAMPLES_MAX = 20


def _excel_cell_to_waybill_str(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, bool):
        return ""
    if isinstance(cell, str):
        return cell.strip()
    if isinstance(cell, int):
        return str(cell)
    if isinstance(cell, float):
        if cell == int(cell):
            return str(int(cell))
        return str(cell).strip()
    return str(cell).strip()


def register_express_routes(bp):
    @bp.route("/express-companies")
    @login_required
    @menu_required("express")
    def express_company_list():
        companies = ExpressCompany.query.order_by(ExpressCompany.id).all()
        stats = {}
        for c in companies:
            av = (
                db.session.query(func.count(ExpressWaybill.id))
                .filter(
                    ExpressWaybill.express_company_id == c.id,
                    ExpressWaybill.status == "available",
                )
                .scalar()
                or 0
            )
            us = (
                db.session.query(func.count(ExpressWaybill.id))
                .filter(
                    ExpressWaybill.express_company_id == c.id,
                    ExpressWaybill.status == "used",
                )
                .scalar()
                or 0
            )
            stats[c.id] = {"available": av, "used": us}
        return render_template(
            "express/company_list.html", companies=companies, stats=stats
        )

    @bp.route("/express-companies/new", methods=["GET", "POST"])
    @login_required
    @menu_required("express")
    @capability_required("express.action.company_create")
    def express_company_new():
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            code = (request.form.get("code") or "").strip().upper()
            if not name or not code:
                flash("请填写名称与短码。", "danger")
                return render_template("express/company_form.html", company=None)
            if ExpressCompany.query.filter_by(code=code).first():
                flash("短码已存在。", "danger")
                return render_template("express/company_form.html", company=None)
            c = ExpressCompany(name=name, code=code, is_active=True)
            db.session.add(c)
            db.session.commit()
            flash("已添加快递公司。", "success")
            return redirect(url_for("main.express_company_list"))
        return render_template("express/company_form.html", company=None)

    @bp.route("/express-companies/<int:cid>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("express")
    @capability_required("express.action.company_edit")
    def express_company_edit(cid):
        c = ExpressCompany.query.get_or_404(cid)
        if request.method == "POST":
            c.name = (request.form.get("name") or "").strip() or c.name
            code = (request.form.get("code") or "").strip().upper()
            if code:
                other = ExpressCompany.query.filter(
                    ExpressCompany.code == code, ExpressCompany.id != c.id
                ).first()
                if other:
                    flash("短码已存在。", "danger")
                    return render_template("express/company_form.html", company=c)
                c.code = code
            c.is_active = request.form.get("is_active") == "1"
            db.session.commit()
            flash("已保存。", "success")
            return redirect(url_for("main.express_company_list"))
        return render_template("express/company_form.html", company=c)

    @bp.route("/express-waybills/import-template", methods=["GET"])
    @login_required
    @menu_required("express")
    @capability_required("express.action.waybill_import")
    def express_waybill_import_template():
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "单号导入"
        ws.cell(1, 1, "快递单号")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="快递单号导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/express-waybills/import", methods=["GET", "POST"])
    @login_required
    @menu_required("express")
    @capability_required("express.action.waybill_import")
    def express_waybill_import():
        companies = (
            ExpressCompany.query.filter(ExpressCompany.is_active.is_(True))
            .filter(ExpressCompany.code != "LEGACY")
            .order_by(ExpressCompany.name)
            .all()
        )
        if not companies:
            flash("请先添加快递公司。", "warning")
            return redirect(url_for("main.express_company_new"))

        if request.method == "POST":
            import_mode = (request.form.get("import_mode") or "range").strip()
            mode_hint = import_mode if import_mode in ("range", "excel") else "range"
            ec_id = request.form.get("express_company_id", type=int)
            if not ec_id:
                flash("请选择快递公司。", "danger")
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint=mode_hint,
                    result=None,
                )
            ec = ExpressCompany.query.get(ec_id)
            if not ec or ec.code == "LEGACY":
                flash("无效的快递公司。", "danger")
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint=mode_hint,
                    result=None,
                )

            if import_mode == "excel":
                file = request.files.get("file")
                if not file or not (file.filename or "").strip():
                    flash("请先选择要上传的 Excel 文件。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
                try:
                    wb = load_workbook(file, data_only=True)
                    ws = wb.active
                except Exception:
                    flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
                inserted = 0
                skipped = 0
                errors: list[str] = []
                for idx, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    cell = row[0] if row else None
                    raw = _excel_cell_to_waybill_str(cell)
                    if not raw:
                        continue
                    res, msg = apply_waybill_to_pool(ec_id, ec.code, raw)
                    if res == "inserted":
                        inserted += 1
                    elif res == "skipped":
                        skipped += 1
                    else:
                        errors.append(f"第 {idx} 行（{raw}）：{msg}")
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                    flash("保存失败，请重试。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
                if inserted or skipped:
                    flash(
                        f"Excel 导入：新增 {inserted} 条，跳过（已存在）{skipped} 条。",
                        "success",
                    )
                elif not errors:
                    flash("未导入任何单号：请确认第一列从第 2 行起有数据。", "warning")
                if errors:
                    flash(f"有 {len(errors)} 条未写入，请查看下方列表。", "warning")
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint="excel",
                    result={
                        "inserted": inserted,
                        "skipped": skipped,
                        "errors": errors,
                        "invalid_prefix": 0,
                        "invalid_samples": [],
                    },
                )

            step_raw = (request.form.get("waybill_step") or "").strip()
            if not step_raw:
                step = 1
            else:
                try:
                    step = int(step_raw)
                except ValueError:
                    flash("单号间隔须为正整数。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
                if step < 1:
                    flash("单号间隔须为正整数。", "danger")
                    return render_template(
                        "express/waybill_import.html",
                        companies=companies,
                        import_mode_hint=mode_hint,
                        result=None,
                    )
            start = (request.form.get("waybill_start") or "").strip()
            end = (request.form.get("waybill_end") or "").strip()
            try:
                numbers = expand_waybill_range(start, end, step)
            except ValueError as e:
                flash(str(e), "danger")
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint=mode_hint,
                    result=None,
                )
            inserted = 0
            skipped = 0
            invalid_prefix = 0
            invalid_samples: list[str] = []
            for no in numbers:
                res, _msg = apply_waybill_to_pool(ec_id, ec.code, no)
                if res == "inserted":
                    inserted += 1
                elif res == "skipped":
                    skipped += 1
                else:
                    invalid_prefix += 1
                    if len(invalid_samples) < _INVALID_SAMPLES_MAX:
                        invalid_samples.append(no)
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                flash("保存失败，请重试。", "danger")
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint=mode_hint,
                    result=None,
                )
            if invalid_prefix:
                flash(
                    f"录入结束：新增 {inserted} 条，跳过（已存在）{skipped} 条，"
                    f"短码不符跳过 {invalid_prefix} 条。",
                    "warning",
                )
                return render_template(
                    "express/waybill_import.html",
                    companies=companies,
                    import_mode_hint="range",
                    result={
                        "inserted": inserted,
                        "skipped": skipped,
                        "errors": [],
                        "invalid_prefix": invalid_prefix,
                        "invalid_samples": invalid_samples,
                    },
                )
            flash(
                f"录入完成：新增 {inserted} 条，跳过（已存在）{skipped} 条。",
                "success",
            )
            return redirect(url_for("main.express_waybill_import"))

        return render_template(
            "express/waybill_import.html",
            companies=companies,
            import_mode_hint=None,
            result=None,
        )

    @bp.route("/express-waybills", methods=["GET"])
    @login_required
    @menu_required("express")
    def express_waybill_list():
        companies = (
            ExpressCompany.query.filter(ExpressCompany.is_active.is_(True))
            .filter(ExpressCompany.code != "LEGACY")
            .order_by(ExpressCompany.name)
            .all()
        )

        express_company_id = request.args.get("express_company_id", type=int)
        keyword = (request.args.get("q") or "").strip()
        page = request.args.get("page", 1, type=int)

        selected_company = None
        pagination = None
        if express_company_id:
            selected_company = ExpressCompany.query.get(express_company_id)
            if selected_company and selected_company.code != "LEGACY":
                q = ExpressWaybill.query.filter(
                    ExpressWaybill.status == "available",
                    ExpressWaybill.express_company_id == express_company_id,
                )
                kw_cond = keyword_like_or(keyword, ExpressWaybill.waybill_no)
                if kw_cond is not None:
                    q = q.filter(kw_cond)
                q = q.order_by(ExpressWaybill.id.desc())
                pagination = q.paginate(page=page, per_page=20)

        return render_template(
            "express/waybill_list.html",
            companies=companies,
            express_company_id=express_company_id,
            selected_company=selected_company,
            keyword=keyword,
            pagination=pagination,
        )

    @bp.route("/express-waybills/batch-delete", methods=["POST"])
    @login_required
    @menu_required("express")
    @capability_required("express.action.waybill_batch_delete")
    def express_waybill_batch_delete():
        express_company_id = request.form.get("express_company_id", type=int)
        keyword = (request.form.get("q") or "").strip()
        page = request.form.get("page", type=int) or 1

        if not express_company_id:
            flash("请选择快递公司。", "danger")
            return redirect(url_for("main.express_waybill_list"))

        raw_ids = request.form.getlist("ids")
        ids: list[int] = []
        for rid in raw_ids:
            try:
                ids.append(int(rid))
            except (TypeError, ValueError):
                continue
        requested_ids = list(dict.fromkeys(ids))  # 保序去重
        if not requested_ids:
            flash("请至少勾选一条可删除的单号。", "warning")
            return redirect(
                url_for(
                    "main.express_waybill_list",
                    express_company_id=express_company_id,
                    q=keyword,
                    page=page,
                )
            )

        valid_ids = [
            rid
            for (rid,) in db.session.query(ExpressWaybill.id)
            .filter(
                ExpressWaybill.id.in_(requested_ids),
                ExpressWaybill.status == "available",
                ExpressWaybill.express_company_id == express_company_id,
            )
            .all()
        ]

        if not valid_ids:
            flash("未找到可删除的单号（可能已被占用或不存在）。", "warning")
            return redirect(
                url_for(
                    "main.express_waybill_list",
                    express_company_id=express_company_id,
                    q=keyword,
                    page=page,
                )
            )

        used_ids = set(
            rid
            for (rid,) in db.session.query(Delivery.express_waybill_id)
            .filter(Delivery.express_waybill_id.isnot(None))
            .filter(Delivery.express_waybill_id.in_(valid_ids))
            .all()
        )
        delete_ids = [rid for rid in valid_ids if rid not in used_ids]

        if not delete_ids:
            flash("所选单号均已被占用，无法删除。", "warning")
            return redirect(
                url_for(
                    "main.express_waybill_list",
                    express_company_id=express_company_id,
                    q=keyword,
                    page=page,
                )
            )

        try:
            db.session.query(ExpressWaybill).filter(
                ExpressWaybill.id.in_(delete_ids)
            ).delete(synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("删除失败，请重试。", "danger")
            return redirect(
                url_for(
                    "main.express_waybill_list",
                    express_company_id=express_company_id,
                    q=keyword,
                    page=page,
                )
            )

        deleted = len(delete_ids)
        skipped = len(requested_ids) - deleted
        if skipped > 0:
            flash(
                f"已删除 {deleted} 条；其余 {skipped} 条未删除（可能已被占用或不存在）。",
                "warning",
            )
        else:
            flash(f"已删除 {deleted} 条快递单号。", "success")

        return redirect(
            url_for(
                "main.express_waybill_list",
                express_company_id=express_company_id,
                q=keyword,
                page=1,
            )
        )
