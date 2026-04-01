from decimal import Decimal, InvalidOperation

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, abort
from flask_login import login_required

from app.auth.capabilities import current_user_can_cap, customer_product_list_read_filters
from app.auth.decorators import capability_required, menu_required
from sqlalchemy.orm import joinedload

from app import db
from app.models import Customer, Company, Product, CustomerProduct
from app.utils.query import is_valid_customer_search_keyword, keyword_like_or
from app.utils.visibility import is_admin, customer_product_view
from io import BytesIO


def register_customer_product_routes(bp):
    @bp.route("/customer-products")
    @login_required
    @menu_required("customer_product")
    def customer_product_list():
        page, customer_id, keyword = customer_product_list_read_filters()
        q = CustomerProduct.query.options(
            joinedload(CustomerProduct.customer).joinedload(Customer.company),
            joinedload(CustomerProduct.product),
        )
        if customer_id:
            q = q.filter(CustomerProduct.customer_id == customer_id)
        if keyword:
            q = (
                q.outerjoin(Customer, CustomerProduct.customer_id == Customer.id)
                .outerjoin(Product, CustomerProduct.product_id == Product.id)
            )
            cond = keyword_like_or(
                keyword,
                Customer.name,
                Customer.customer_code,
                Product.product_code,
                Product.name,
                Product.spec,
                CustomerProduct.customer_material_no,
            )
            if cond is not None:
                q = q.filter(cond)
            q = q.distinct()
        q = q.order_by(CustomerProduct.id.desc())
        pagination = q.paginate(page=page, per_page=20)
        rows = [customer_product_view(cp) for cp in pagination.items]
        customers = Customer.query.order_by(Customer.name).all()
        return render_template(
            "customer_product/list.html",
            pagination=pagination,
            rows=rows,
            customers=customers,
            customer_id=customer_id,
            keyword=keyword,
            is_admin=is_admin(),
        )

    @bp.route("/customer-products/new", methods=["GET", "POST"])
    @login_required
    @menu_required("customer_product")
    @capability_required("customer_product.action.create")
    def customer_product_new():
        if request.method == "POST":
            return _customer_product_save(None)
        return render_template(
            "customer_product/form.html",
            cp=None,
            is_admin=is_admin(),
            customer_id_val=None,
            product_id_val=None,
            customer_label="",
            product_label="",
        )

    @bp.route("/customer-products/<int:cp_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("customer_product")
    @capability_required("customer_product.action.edit")
    def customer_product_edit(cp_id):
        cp = (
            CustomerProduct.query.options(
                joinedload(CustomerProduct.customer).joinedload(Customer.company),
                joinedload(CustomerProduct.product),
            ).get_or_404(cp_id)
        )
        if request.method == "POST":
            return _customer_product_save(cp)
        cust = cp.customer
        co_txt = "-"
        if cust and cust.company:
            co_txt = f"{cust.company.name}({cust.company.code})"
        customer_label = (
            f"{cust.customer_code} - {cust.name} · 主体 {co_txt}" if cust else ""
        )
        p = cp.product
        product_label = (
            f"{p.product_code} - {p.name}" + (f"（{p.spec}）" if p and p.spec else "")
            if p
            else ""
        )
        return render_template(
            "customer_product/form.html",
            cp=cp,
            is_admin=is_admin(),
            customer_id_val=cp.customer_id,
            product_id_val=cp.product_id,
            customer_label=customer_label,
            product_label=product_label,
        )

    @bp.route("/customer-products/<int:cp_id>/delete", methods=["POST"])
    @login_required
    @menu_required("customer_product")
    @capability_required("customer_product.action.delete")
    def customer_product_delete(cp_id):
        cp = CustomerProduct.query.get_or_404(cp_id)
        db.session.delete(cp)
        db.session.commit()
        flash("客户产品已删除。", "success")
        return redirect(url_for("main.customer_product_list"))

    @bp.route("/customer-products/export-import-template", methods=["GET"])
    @login_required
    @menu_required("customer_product")
    @capability_required("customer_product.action.export_template")
    def export_customer_product_import_template():
        """客户产品导入模板（xlsx）：表头行+1行空白。"""
        from openpyxl import Workbook

        headers = [
            "客户名称",
            "产品编号",
            "客户料号",
            "物料编号（自动，勿填）",
            "单位",
            "单价",
            "币种",
            "备注",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "客户产品导入模板"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="客户产品导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/api/customers-search-cp")
    @login_required
    @menu_required("customer_product")
    def customers_search_cp():
        if not (
            current_user_can_cap("customer_product.action.create")
            or current_user_can_cap("customer_product.action.edit")
        ):
            abort(403)
        qstr = (request.args.get("q") or "").strip()
        limit = request.args.get("limit", 20, type=int)
        limit = max(1, min(limit, 20))
        if not is_valid_customer_search_keyword(qstr):
            return jsonify({"items": []})

        like = f"%{qstr}%"
        q = (
            Customer.query.options(joinedload(Customer.company))
            .outerjoin(Company, Customer.company_id == Company.id)
            .filter(
                db.or_(
                    Customer.name.like(like),
                    Customer.customer_code.like(like),
                    Customer.short_code.like(like),
                    Company.name.like(like),
                    Company.code.like(like),
                )
            )
            .order_by(Customer.customer_code)
        )
        items = []
        for c in q.limit(limit).all():
            co = c.company
            co_txt = f"{co.name}({co.code})" if co else "-"
            items.append(
                {
                    "id": c.id,
                    "label": f"{c.customer_code} - {c.name} · 主体 {co_txt}",
                }
            )
        return jsonify({"items": items})

    @bp.route("/api/products-search-cp")
    @login_required
    @menu_required("customer_product")
    def products_search_cp():
        if not (
            current_user_can_cap("customer_product.action.create")
            or current_user_can_cap("customer_product.action.edit")
        ):
            abort(403)
        qstr = (request.args.get("q") or "").strip()
        limit = request.args.get("limit", 20, type=int)
        limit = max(1, min(limit, 20))
        q = Product.query.order_by(Product.product_code)
        if qstr:
            like = f"%{qstr}%"
            q = q.filter(
                db.or_(
                    Product.product_code.like(like),
                    Product.name.like(like),
                    Product.spec.like(like),
                )
            )
        items = [
            {
                "id": p.id,
                "label": f"{p.product_code} - {p.name}"
                + (f"（{p.spec}）" if p.spec else ""),
            }
            for p in q.limit(limit).all()
        ]
        return jsonify({"items": items})

    @bp.route("/customer-products/import", methods=["GET", "POST"])
    @login_required
    @menu_required("customer_product")
    @capability_required("customer_product.action.import")
    def customer_product_import():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("customer_product/import.html", result=None)
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("customer_product/import.html", result=None)
            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("customer_product/import.html", result=None)
            admin = is_admin()
            success = 0
            errors = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                cols = (row + (None,) * 8)[:8]
                cust_name, prod_code, cust_mat_no, _mat_no_ignored, unit, price, currency, remark = cols
                cust_name = (cust_name or "").strip() if isinstance(cust_name, str) else (cust_name or "")
                prod_code = (prod_code or "").strip() if isinstance(prod_code, str) else (prod_code or "")
                if not cust_name or not prod_code:
                    if any(row):
                        errors.append(f"第 {idx} 行：客户名称或产品编号为空")
                    continue
                customer = Customer.query.filter_by(name=cust_name).first()
                if not customer:
                    errors.append(f"第 {idx} 行：客户“{cust_name}”不存在，请先在客户表维护")
                    continue
                product = Product.query.filter_by(product_code=prod_code).first()
                if not product:
                    errors.append(f"第 {idx} 行：产品编号“{prod_code}”不存在，请先在产品表维护")
                    continue
                cp = CustomerProduct.query.filter_by(
                    customer_id=customer.id, product_id=product.id
                ).first()
                if not cp:
                    cp = CustomerProduct(customer_id=customer.id, product_id=product.id)
                    db.session.add(cp)
                cp.customer_material_no = (
                    (cust_mat_no or "").strip() if isinstance(cust_mat_no, str) else cust_mat_no
                )
                cp.material_no = product.product_code
                cp.unit = (unit or "").strip() if isinstance(unit, str) else unit
                if admin:
                    price_raw = (price or "").strip() if isinstance(price, str) else price
                    try:
                        cp.price = Decimal(str(price_raw)) if price_raw not in (None, "") else None
                    except (InvalidOperation, ValueError):
                        cp.price = None
                    cp.currency = (currency or "").strip() if isinstance(currency, str) else currency
                cp.remark = (remark or "").strip() if isinstance(remark, str) else remark
                success += 1
            db.session.commit()
            result = {"success": success, "errors": errors}
            if success:
                flash(f"成功导入或更新 {success} 条客户产品。", "success")
            if errors:
                flash(f"有 {len(errors)} 条记录导入失败，请查看错误列表。", "danger")
            return render_template("customer_product/import.html", result=result)
        return render_template("customer_product/import.html", result=None)

    def _customer_product_save(cp):
        customer_id = request.form.get("customer_id", type=int)
        product_id = request.form.get("product_id", type=int)
        if not customer_id or not product_id:
            flash("客户和产品为必选（请搜索并点选）。", "danger")
            return render_template(
                "customer_product/form.html",
                cp=cp,
                is_admin=is_admin(),
                customer_id_val=customer_id,
                product_id_val=product_id,
                customer_label=(request.form.get("customer_label") or "").strip(),
                product_label=(request.form.get("product_label") or "").strip(),
            )
        if cp is None:
            cp = CustomerProduct()
        cp.customer_id = customer_id
        cp.product_id = product_id
        cp.customer_material_no = (request.form.get("customer_material_no") or "").strip() or None
        prod_row = db.session.get(Product, product_id)
        cp.material_no = (prod_row.product_code or "") if prod_row else None
        cp.unit = (request.form.get("unit") or "").strip() or None
        if is_admin():
            price_raw = (request.form.get("price") or "").strip()
            try:
                cp.price = Decimal(price_raw) if price_raw else None
            except (InvalidOperation, ValueError):
                cp.price = None
            cp.currency = (request.form.get("currency") or "").strip() or None
        cp.remark = (request.form.get("remark") or "").strip() or None
        db.session.add(cp)
        db.session.commit()
        flash("客户产品已保存。", "success")
        return redirect(url_for("main.customer_product_list"))
