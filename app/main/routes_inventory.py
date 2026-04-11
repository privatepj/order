import json
from datetime import date
from decimal import Decimal, InvalidOperation

from io import BytesIO

from flask import abort, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from app import db
from app.auth.capabilities import current_user_can_cap, inventory_stock_query_read_filters
from app.auth.decorators import capability_required, menu_required
from app.models import (
    InventoryDailyLine,
    InventoryDailyRecord,
    InventoryMovement,
    InventoryMovementBatch,
    InventoryOpeningBalance,
    Product,
    PurchaseOrder,
    PurchaseReceipt,
    SemiMaterial,
    User,
)
from app.services import inventory_svc
from app.utils.query import keyword_like_or

INVENTORY_OPS_MENU_CODES = (
    "inventory_ops_finished",
    "inventory_ops_semi",
    "inventory_ops_material",
)
_INVENTORY_CAP_SUFFIXES = (
    "api.products_search",
    "api.suggest_storage_area",
    "movement.list",
    "movement.create",
    "movement.delete",
    "movement_batch.void",
    "opening.list",
    "opening.create",
    "opening.edit",
    "opening.delete",
    "daily.list",
    "daily.create",
    "daily.detail",
    "daily.edit",
    "daily.delete",
)
INVENTORY_CAP_KEYS = {
    suffix: tuple(f"{menu_code}.{suffix}" for menu_code in INVENTORY_OPS_MENU_CODES)
    for suffix in _INVENTORY_CAP_SUFFIXES
}


def _menu_code_for_category(category: str) -> str:
    if category == inventory_svc.INV_SEMI:
        return "inventory_ops_semi"
    if category == inventory_svc.INV_MATERIAL:
        return "inventory_ops_material"
    return "inventory_ops_finished"


def _parse_line_rows():
    """从表单解析 (product_id, quantity, unit, note) 列表；非法时 raise ValueError。"""
    pids = request.form.getlist("line_product_id")
    qtys = request.form.getlist("line_quantity")
    units = request.form.getlist("line_unit")
    notes = request.form.getlist("line_note")
    rows = []
    seen_pid = set()
    for i, pid in enumerate(pids):
        pid = (pid or "").strip()
        if not pid:
            continue
        try:
            product_id = int(pid)
        except ValueError:
            continue
        if product_id in seen_pid:
            raise ValueError("同一单据中不能重复选择同一产品。")
        seen_pid.add(product_id)
        qraw = (qtys[i] if i < len(qtys) else "") or "0"
        qraw = str(qraw).strip()
        try:
            qty = Decimal(qraw)
        except InvalidOperation:
            raise ValueError("数量格式不正确。")
        if qty < 0:
            raise ValueError("数量不能为负数。")
        unit = (units[i] if i < len(units) else None) or None
        if unit:
            unit = unit.strip()[:16] or None
        note = (notes[i] if i < len(notes) else None) or None
        if note:
            note = note.strip()[:255] or None
        rows.append((product_id, qty, unit, note))
    return rows


def _validate_products(product_ids):
    if not product_ids:
        return
    cnt = Product.query.filter(Product.id.in_(product_ids)).count()
    if cnt != len(set(product_ids)):
        raise ValueError("存在无效的产品。")


def _validate_semi_materials(kind: str, item_ids):
    if not item_ids:
        return
    cnt = (
        SemiMaterial.query.filter(SemiMaterial.kind == kind, SemiMaterial.id.in_(item_ids))
        .count()
    )
    if cnt != len(set(item_ids)):
        raise ValueError(f"存在无效的{kind}主数据。")


def _parse_movement_line_rows():
    """解析批量手工出入库行：(product_id, storage_area, quantity, unit, remark)；非法时 raise ValueError。"""
    pids = request.form.getlist("line_product_id")
    areas = request.form.getlist("line_storage_area")
    qtys = request.form.getlist("line_quantity")
    units = request.form.getlist("line_unit")
    remarks = request.form.getlist("line_remark")
    rows = []
    for i, pid in enumerate(pids):
        pid = (pid or "").strip()
        if not pid:
            continue
        try:
            product_id = int(pid)
        except ValueError:
            continue
        area = (areas[i] if i < len(areas) else "") or ""
        area = area.strip()
        if not area:
            raise ValueError("每一行有产品的记录都必须填写仓储区。")
        qraw = (qtys[i] if i < len(qtys) else "") or "0"
        qraw = str(qraw).strip()
        try:
            qty = Decimal(qraw)
        except InvalidOperation:
            raise ValueError("数量格式不正确。")
        if qty <= 0:
            raise ValueError("每一行数量须大于 0。")
        unit = (units[i] if i < len(units) else None) or None
        if unit:
            unit = unit.strip()[:16] or None
        remark = (remarks[i] if i < len(remarks) else None) or None
        if remark:
            remark = remark.strip()[:255] or None
        rows.append((product_id, area, qty, unit, remark))
    return rows


def _parse_movement_material_line_rows():
    """解析批量手工半成品/物料出入库行：(material_id, storage_area, quantity, unit, remark)。"""
    pids = request.form.getlist("line_material_id")
    areas = request.form.getlist("line_storage_area")
    qtys = request.form.getlist("line_quantity")
    units = request.form.getlist("line_unit")
    remarks = request.form.getlist("line_remark")
    rows = []
    for i, pid in enumerate(pids):
        pid = (pid or "").strip()
        if not pid:
            continue
        try:
            material_id = int(pid)
        except ValueError:
            continue
        area = (areas[i] if i < len(areas) else "") or ""
        area = area.strip()
        if not area:
            raise ValueError("每一行有物料的记录都必须填写仓储区。")
        qraw = (qtys[i] if i < len(qtys) else "") or "0"
        qraw = str(qraw).strip()
        try:
            qty = Decimal(qraw)
        except InvalidOperation:
            raise ValueError("数量格式不正确。")
        if qty <= 0:
            raise ValueError("每一行数量须大于 0。")
        unit = (units[i] if i < len(units) else None) or None
        if unit:
            unit = unit.strip()[:16] or None
        remark = (remarks[i] if i < len(remarks) else None) or None
        if remark:
            remark = remark.strip()[:255] or None
        rows.append((material_id, area, qty, unit, remark))
    return rows


def _cell_str_movement_import(val):
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()


def _movement_import_qty_display(qty: Decimal) -> str:
    s = format(qty, "f").rstrip("0").rstrip(".")
    return s if s else "0"


def _parse_movement_import_excel(ws):
    """解析库存录入 Excel：列顺序 品名、规格、仓储区、数量、单位、备注；第 2 行起为数据。"""
    parsed = []
    errors = []
    failed_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (tuple(row) + (None,) * 6)[:6]
        name, spec, area, qraw, unit, remark = vals
        name_s = _cell_str_movement_import(name)
        spec_s = _cell_str_movement_import(spec)
        area_s = _cell_str_movement_import(area)
        unit_s = _cell_str_movement_import(unit)
        remark_s = _cell_str_movement_import(remark)
        q_empty = qraw is None or (isinstance(qraw, str) and not str(qraw).strip())
        if not any([name_s, spec_s, area_s, unit_s, remark_s]) and q_empty:
            continue
        if q_empty:
            reason = "数量不能为空"
            errors.append(f"{inventory_svc.movement_import_label(name_s, spec_s)}：{reason}")
            failed_rows.append(
                inventory_svc.movement_import_failed_row(
                    name=name_s,
                    spec=spec_s,
                    area=area_s,
                    quantity="",
                    unit=unit_s or None,
                    remark=remark_s or None,
                    reason=reason,
                )
            )
            continue
        try:
            qty = Decimal(str(qraw))
        except InvalidOperation:
            reason = "数量格式不正确"
            q_disp = "" if qraw is None else str(qraw).strip()
            errors.append(f"{inventory_svc.movement_import_label(name_s, spec_s)}：{reason}")
            failed_rows.append(
                inventory_svc.movement_import_failed_row(
                    name=name_s,
                    spec=spec_s,
                    area=area_s,
                    quantity=q_disp,
                    unit=unit_s or None,
                    remark=remark_s or None,
                    reason=reason,
                )
            )
            continue
        parsed.append((0, name_s, spec_s, area_s, qty, unit_s or None, remark_s or None))
    return parsed, errors, failed_rows


def _dedupe_movement_import_parsed(parsed):
    """
    同一文件内六列（品名、规格、仓储区、数量、单位、备注）规范化后完全相同的行，
    仅保留首次出现，其余记为重复失败。
    """
    seen = set()
    out = []
    errors = []
    failed_rows = []

    def _unit_key(u):
        if u is None:
            return ""
        return u.strip()[:16] if isinstance(u, str) else str(u).strip()[:16]

    def _remark_key(r):
        if r is None:
            return ""
        return r.strip()[:255] if isinstance(r, str) else str(r).strip()[:255]

    for row in parsed:
        _, name_s, spec_s, area_s, qty, unit, remark = row
        spec_n = inventory_svc.normalize_spec_for_match(spec_s)
        key = (
            name_s.strip(),
            spec_n,
            area_s.strip(),
            str(qty),
            _unit_key(unit),
            _remark_key(remark),
        )
        if key in seen:
            reason = "与文件中前面的行完全重复（品名、规格、仓储区、数量、单位、备注均相同）"
            errors.append(f"{inventory_svc.movement_import_label(name_s, spec_s)}：{reason}")
            failed_rows.append(
                inventory_svc.movement_import_failed_row(
                    name=name_s,
                    spec=spec_s,
                    area=area_s,
                    quantity=_movement_import_qty_display(qty),
                    unit=_unit_key(unit) or None,
                    remark=_remark_key(remark) or None,
                    reason=reason,
                )
            )
            continue
        seen.add(key)
        out.append(row)
    return out, errors, failed_rows


def register_inventory_routes(bp):
    @bp.route("/inventory/finished")
    @login_required
    @menu_required("inventory_ops_finished")
    def inventory_finished_entry():
        return redirect(url_for("main.inventory_movement_new", category=inventory_svc.INV_FINISHED))

    @bp.route("/inventory/semi")
    @login_required
    @menu_required("inventory_ops_semi")
    def inventory_semi_entry():
        return redirect(url_for("main.inventory_movement_new", category=inventory_svc.INV_SEMI))

    @bp.route("/inventory/material")
    @login_required
    @menu_required("inventory_ops_material")
    def inventory_material_entry():
        return redirect(url_for("main.inventory_movement_new", category=inventory_svc.INV_MATERIAL))

    # ----- API：产品搜索（库存录入用） -----
    @bp.route("/api/inventory/products-search", methods=["GET"])
    @login_required
    @menu_required("inventory_ops_finished")
    def inventory_products_search():
        if not current_user_can_cap("inventory_ops_finished.api.products_search"):
            abort(403)
        qstr = (request.args.get("q") or "").strip()
        limit = request.args.get("limit", 20, type=int)
        limit = max(1, min(limit, 20))
        q = Product.query.order_by(Product.product_code)
        cond = keyword_like_or(
            qstr,
            Product.product_code,
            Product.name,
            Product.spec,
            Product.base_unit,
            Product.remark,
        )
        if cond is not None:
            q = q.filter(cond)
        items = []
        for p in q.limit(limit).all():
            label = f"{p.product_code} — {p.name}"
            if p.spec:
                label += f"（{p.spec}）"
            items.append(
                {
                    "id": p.id,
                    "label": label,
                    "spec": p.spec or "",
                    "base_unit": p.base_unit or "",
                    "category": inventory_svc.INV_FINISHED,
                }
            )
        return jsonify({"items": items})

    @bp.route("/api/inventory/suggest-storage-area", methods=["GET"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    def inventory_suggest_storage_area():
        category = (request.args.get("category") or "").strip() or inventory_svc.INV_FINISHED
        item_id = request.args.get("item_id", type=int)
        if not item_id:
            item_id = request.args.get("product_id", type=int)
        if not item_id:
            item_id = request.args.get("material_id", type=int)

        if not item_id:
            return jsonify({"storage_area": ""})

        if category not in (
            inventory_svc.INV_FINISHED,
            inventory_svc.INV_SEMI,
            inventory_svc.INV_MATERIAL,
        ):
            category = inventory_svc.INV_FINISHED
        menu_code = _menu_code_for_category(category)
        if not current_user_can_cap(f"{menu_code}.api.suggest_storage_area"):
            abort(403)

        area = inventory_svc.suggest_storage_area_for_category_item(category, item_id)
        return jsonify({"storage_area": area})

    # ----- API：半成品/物料搜索（库存录入用） -----
    @bp.route("/api/inventory/semi-materials-search", methods=["GET"])
    @login_required
    @menu_required("inventory_ops_semi", "inventory_ops_material")
    def inventory_semi_materials_search():
        qstr = (request.args.get("q") or "").strip()
        kind = (request.args.get("kind") or request.args.get("category") or "").strip()
        if kind not in (inventory_svc.INV_SEMI, inventory_svc.INV_MATERIAL):
            kind = inventory_svc.INV_SEMI
        menu_code = _menu_code_for_category(kind)
        if not current_user_can_cap(f"{menu_code}.api.products_search"):
            abort(403)
        limit = request.args.get("limit", 20, type=int)
        limit = max(1, min(limit, 20))

        q = SemiMaterial.query.filter(SemiMaterial.kind == kind).order_by(SemiMaterial.code)
        cond = keyword_like_or(
            qstr,
            SemiMaterial.code,
            SemiMaterial.name,
            SemiMaterial.spec,
            SemiMaterial.base_unit,
            SemiMaterial.remark,
        )
        if cond is not None:
            q = q.filter(cond)

        items = []
        for it in q.limit(limit).all():
            label = f"{it.code} — {it.name}"
            if it.spec:
                label += f"（{it.spec}）"
            items.append(
                {
                    "id": it.id,
                    "label": label,
                    "spec": it.spec or "",
                    "base_unit": it.base_unit or "",
                    "category": it.kind,
                }
            )
        return jsonify({"items": items})

    # ----- 库存查询（结存聚合） -----
    @bp.route("/inventory/query")
    @login_required
    @menu_required("inventory_query")
    def inventory_stock_query():
        page, category, storage_area, spec_kw, name_spec_kw = inventory_stock_query_read_filters()
        rows, total = inventory_svc.query_stock_aggregate(
            category=category,
            storage_area_kw=storage_area,
            spec_kw=spec_kw,
            name_spec_kw=name_spec_kw,
            page=page,
            per_page=30,
        )
        total_pages = (total + 30 - 1) // 30 if total else 1
        return render_template(
            "inventory/stock_query.html",
            rows=rows,
            page=page,
            total=total,
            total_pages=total_pages,
            category=category,
            storage_area=storage_area,
            spec_kw=spec_kw,
            name_spec_kw=name_spec_kw,
        )

    # ----- 库存录入（批量手工出入库） -----
    @bp.route("/inventory/movement/import-template", methods=["GET"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["movement.create"])
    def inventory_movement_import_template():
        from openpyxl import Workbook

        headers = ["品名", "规格", "仓储区", "数量", "单位", "备注"]
        wb = Workbook()
        ws = wb.active
        ws.title = "库存录入导入"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="库存录入导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/inventory/movement/export-failed", methods=["POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["movement.create"])
    def inventory_movement_export_failed():
        raw = (request.form.get("failed_rows_json") or "").strip()
        if not raw:
            flash("没有可导出的失败明细。", "warning")
            return redirect(url_for("main.inventory_movement_new"))
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            flash("导出失败：数据格式无效。", "danger")
            return redirect(url_for("main.inventory_movement_new"))
        if not isinstance(data, list):
            flash("导出失败：数据格式无效。", "danger")
            return redirect(url_for("main.inventory_movement_new"))
        max_rows = 500
        if len(data) > max_rows:
            flash(f"导出失败：失败行超过 {max_rows} 条上限。", "danger")
            return redirect(url_for("main.inventory_movement_new"))
        from openpyxl import Workbook

        headers = ["品名", "规格", "仓储区", "数量", "单位", "备注", "失败原因"]
        wb = Workbook()
        ws = wb.active
        ws.title = "导入失败"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)
        r = 2
        for item in data:
            if not isinstance(item, dict):
                continue
            ws.cell(r, 1, item.get("name") or "")
            ws.cell(r, 2, item.get("spec") or "")
            ws.cell(r, 3, item.get("area") or "")
            ws.cell(r, 4, item.get("quantity") or "")
            ws.cell(r, 5, item.get("unit") or "")
            ws.cell(r, 6, item.get("remark") or "")
            ws.cell(r, 7, item.get("reason") or "")
            r += 1
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="库存导入失败明细.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/inventory/movement/new", methods=["GET", "POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["movement.create"])
    def inventory_movement_new():
        import_result = None
        category_from_query = (request.args.get("category") or "").strip()
        purchase_order_id = (
            request.form.get("purchase_order_id", type=int)
            if request.method == "POST"
            else request.args.get("purchase_order_id", type=int)
        )
        purchase_receipt_id = (
            request.form.get("purchase_receipt_id", type=int)
            if request.method == "POST"
            else request.args.get("purchase_receipt_id", type=int)
        )
        purchase_receipt = (
            PurchaseReceipt.query.options(selectinload(PurchaseReceipt.purchase_order)).get(purchase_receipt_id)
            if purchase_receipt_id
            else None
        )
        purchase_order = None
        if purchase_order_id:
            purchase_order = PurchaseOrder.query.options(selectinload(PurchaseOrder.material)).get(purchase_order_id)
        elif purchase_receipt and purchase_receipt.purchase_order_id:
            purchase_order = PurchaseOrder.query.options(selectinload(PurchaseOrder.material)).get(
                purchase_receipt.purchase_order_id
            )
            purchase_order_id = purchase_order.id if purchase_order else None
        if purchase_receipt and purchase_order and purchase_receipt.purchase_order_id != purchase_order.id:
            purchase_receipt = None
            purchase_receipt_id = None
        purchase_category = None
        purchase_context = None
        if purchase_order:
            purchase_category = (
                purchase_order.material.kind
                if purchase_order.material
                and purchase_order.material.kind in (inventory_svc.INV_SEMI, inventory_svc.INV_MATERIAL)
                else inventory_svc.INV_MATERIAL
            )
            purchase_context = {
                "purchase_order_id": purchase_order.id,
                "purchase_receipt_id": purchase_receipt.id if purchase_receipt else None,
                "po_no": purchase_order.po_no,
                "receipt_no": purchase_receipt.receipt_no if purchase_receipt else None,
                "supplier_name": purchase_order.supplier_name,
                "item_name": purchase_order.item_name,
                "item_spec": purchase_order.item_spec,
                "material_id": purchase_order.material_id,
                "category": purchase_category,
            }
        if request.method == "POST":
            do_excel = request.form.get("do_excel_import") == "1"
            try:
                cat = (request.form.get("category") or "").strip()
                if cat not in (
                    inventory_svc.INV_FINISHED,
                    inventory_svc.INV_SEMI,
                    inventory_svc.INV_MATERIAL,
                ):
                    raise ValueError("请选择类别。")
                if not current_user_can_cap(f"{_menu_code_for_category(cat)}.movement.create"):
                    raise ValueError("您没有该类别库存录入权限。")
                direction = (request.form.get("direction") or "").strip()
                if direction not in ("in", "out"):
                    raise ValueError("请选择入库或出库。")
                rd = request.form.get("biz_date") or ""
                biz_date = date.fromisoformat(rd) if rd else None
                if not biz_date:
                    raise ValueError("请选择业务日期。")

                if purchase_context:
                    if direction != "in":
                        raise ValueError("采购来源关联入库只允许录入入库明细。")
                    if purchase_category and cat != purchase_category:
                        raise ValueError("请使用采购单对应的物料类别进行入库。")

                if do_excel:
                    if purchase_context:
                        raise ValueError("关联采购来源时暂不支持 Excel 导入，请手工录入。")
                    file = request.files.get("excel_file")
                    if not file or not (file.filename or "").strip():
                        raise ValueError("请先选择要上传的 Excel 文件（.xlsx）。")
                    try:
                        from openpyxl import load_workbook
                    except ImportError:
                        raise ValueError("服务器缺少 openpyxl 依赖，无法导入。")
                    try:
                        wb = load_workbook(file, data_only=True)
                        ws = wb.active
                    except Exception:
                        raise ValueError("Excel 文件无法读取，请确认格式为 .xlsx。")
                    parsed, parse_errors, parse_failed_rows = _parse_movement_import_excel(ws)
                    deduped, dup_errors, dup_failed_rows = _dedupe_movement_import_parsed(parsed)
                    all_errors = list(parse_errors) + list(dup_errors)
                    all_failed_rows = list(parse_failed_rows) + list(dup_failed_rows)
                    success_count = 0
                    if deduped:
                        xname = (file.filename or "").strip()
                        success_count, svc_errors, svc_failed_rows = (
                            inventory_svc.import_movements_from_parsed_lines_by_category(
                                deduped,
                                category=cat,
                                direction=direction,
                                biz_date=biz_date,
                                created_by=current_user.id,
                                original_filename=xname or None,
                            )
                        )
                        all_errors.extend(svc_errors)
                        all_failed_rows.extend(svc_failed_rows)
                    elif not parse_errors and not dup_errors:
                        flash("Excel 中无有效数据行。", "warning")
                    import_result = {
                        "success": success_count,
                        "errors": all_errors,
                        "failed_rows": all_failed_rows,
                    }
                    if success_count:
                        flash(f"已从 Excel 保存批次（{success_count} 条明细）。", "success")
                    if all_errors:
                        flash(
                            "部分行未导入或校验失败，请查看下列明细，可导出失败行修正后重试。",
                            "warning",
                        )
                    today = date.today().isoformat()
                    return render_template(
                        "inventory/movement_form.html",
                        default_biz_date=rd or today,
                        form_category=request.form.get("category"),
                        form_direction=request.form.get("direction"),
                        import_result=import_result,
                        purchase_context=purchase_context,
                    )

                if cat == inventory_svc.INV_FINISHED:
                    line_rows = _parse_movement_line_rows()
                    if not line_rows:
                        raise ValueError("请至少录入一行有效的产品、仓储区与数量。")
                    _validate_products([r[0] for r in line_rows])
                    batch = inventory_svc.create_movement_batch(
                        category=inventory_svc.INV_FINISHED,
                        biz_date=biz_date,
                        direction=direction,
                        source=inventory_svc.BATCH_SOURCE_FORM,
                        line_count=len(line_rows),
                        created_by=current_user.id,
                    )
                    for product_id, area, qty, unit, remark in line_rows:
                        p = Product.query.get(product_id)
                        if not p:
                            raise ValueError("存在无效的产品。")
                        inventory_svc.create_manual_movement(
                            category=inventory_svc.INV_FINISHED,
                            direction=direction,
                            product_id=product_id,
                            material_id=0,
                            storage_area=area,
                            quantity=qty,
                            unit=unit or (p.base_unit or None),
                            biz_date=biz_date,
                            remark=remark,
                            created_by=current_user.id,
                            movement_batch_id=batch.id,
                            source_purchase_order_id=purchase_order_id,
                            source_purchase_receipt_id=purchase_receipt_id,
                        )
                else:
                    line_rows = _parse_movement_material_line_rows()
                    if not line_rows:
                        raise ValueError("请至少录入一行有效的半成品/物料、仓储区与数量。")
                    _validate_semi_materials(cat, [r[0] for r in line_rows])
                    batch = inventory_svc.create_movement_batch(
                        category=cat,
                        biz_date=biz_date,
                        direction=direction,
                        source=inventory_svc.BATCH_SOURCE_FORM,
                        line_count=len(line_rows),
                        created_by=current_user.id,
                    )
                    for material_id, area, qty, unit, remark in line_rows:
                        item = SemiMaterial.query.get(material_id)
                        if not item or item.kind != cat:
                            raise ValueError("存在无效的半成品/物料主数据。")
                        inventory_svc.create_manual_movement(
                            category=cat,
                            direction=direction,
                            product_id=0,
                            material_id=material_id,
                            storage_area=area,
                            quantity=qty,
                            unit=unit or (item.base_unit or None),
                            biz_date=biz_date,
                            remark=remark,
                            created_by=current_user.id,
                            movement_batch_id=batch.id,
                            source_purchase_order_id=purchase_order_id,
                            source_purchase_receipt_id=purchase_receipt_id,
                        )
                db.session.commit()
                flash(f"已保存批次（{len(line_rows)} 条明细）。", "success")
                if purchase_receipt_id:
                    return redirect(url_for("main.procurement_receipt_compare", receipt_id=purchase_receipt_id))
                if purchase_order_id:
                    return redirect(url_for("main.procurement_order_detail", po_id=purchase_order_id))
                return redirect(url_for("main.inventory_list"))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except InvalidOperation:
                db.session.rollback()
                flash("数量格式不正确。", "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：数据冲突。", "danger")
        today = date.today().isoformat()
        bd = (request.form.get("biz_date") or today) if request.method == "POST" else today
        return render_template(
            "inventory/movement_form.html",
            default_biz_date=bd,
            form_category=(
                request.form.get("category")
                if request.method == "POST"
                else category_from_query or purchase_category or None
            ),
            form_direction=request.form.get("direction") if request.method == "POST" else ("in" if purchase_context else None),
            import_result=import_result,
            purchase_context=purchase_context,
        )

    # ----- 库存批次列表（主「库存」入口） -----
    @bp.route("/inventory")
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    def inventory_list():
        page = request.args.get("page", 1, type=int)
        q = InventoryMovementBatch.query.options(
            selectinload(InventoryMovementBatch.delivery),
        ).order_by(
            InventoryMovementBatch.id.desc(),
        )
        pagination = q.paginate(page=page, per_page=30)
        uid_set = {b.created_by for b in pagination.items}
        users = (
            {
                u.id: (u.name or u.username)
                for u in User.query.filter(User.id.in_(uid_set)).all()
            }
            if uid_set
            else {}
        )
        return render_template(
            "inventory/movement_list.html",
            pagination=pagination,
            creators=users,
        )

    @bp.route("/inventory/batch/<int:batch_id>")
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    def inventory_batch_detail(batch_id):
        if not any(current_user_can_cap(k) for k in INVENTORY_CAP_KEYS["movement.list"]):
            abort(403)
        batch = InventoryMovementBatch.query.options(
            selectinload(InventoryMovementBatch.delivery),
        ).get_or_404(batch_id)
        movements = (
            InventoryMovement.query.options(
                selectinload(InventoryMovement.product),
                selectinload(InventoryMovement.material),
            )
            .filter_by(movement_batch_id=batch_id)
            .order_by(InventoryMovement.id.asc())
            .all()
        )
        creator = User.query.get(batch.created_by)
        creator_name = (creator.name or creator.username) if creator else batch.created_by
        return render_template(
            "inventory/batch_detail.html",
            batch=batch,
            movements=movements,
            creator_name=creator_name,
        )

    @bp.route("/inventory/batch/<int:batch_id>/void", methods=["POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["movement_batch.void"])
    def inventory_batch_void(batch_id):
        try:
            inventory_svc.void_movement_batch(batch_id)
            db.session.commit()
            flash("已撤销该批次及其全部明细。", "success")
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "warning")
        return redirect(url_for("main.inventory_list"))

    @bp.route("/inventory/movement/<int:movement_id>/delete", methods=["POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["movement.delete"])
    def inventory_movement_delete(movement_id):
        m = InventoryMovement.query.get_or_404(movement_id)
        if m.source_type != inventory_svc.SOURCE_MANUAL:
            flash("仅可删除手工录入的流水。", "warning")
            return redirect(url_for("main.inventory_list"))
        if m.movement_batch_id is not None:
            flash("已归入批次的明细请使用「撤销整批」，勿单条删除。", "warning")
            return redirect(url_for("main.inventory_list"))
        db.session.delete(m)
        db.session.commit()
        flash("已删除该条流水。", "success")
        return redirect(url_for("main.inventory_list"))

    # ----- 期初结存维护 -----
    @bp.route("/inventory/opening")
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["opening.list"])
    def inventory_opening_list():
        page = request.args.get("page", 1, type=int)
        q = InventoryOpeningBalance.query.options(
            selectinload(InventoryOpeningBalance.product),
            selectinload(InventoryOpeningBalance.material),
        ).order_by(
            InventoryOpeningBalance.storage_area,
            InventoryOpeningBalance.id,
        )
        pagination = q.paginate(page=page, per_page=30)
        return render_template(
            "inventory/opening_list.html",
            pagination=pagination,
        )

    @bp.route("/inventory/opening/new", methods=["GET", "POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["opening.create"])
    def inventory_opening_new():
        if request.method == "POST":
            try:
                cat = (request.form.get("category") or "").strip()
                if cat not in (
                    inventory_svc.INV_FINISHED,
                    inventory_svc.INV_SEMI,
                    inventory_svc.INV_MATERIAL,
                ):
                    raise ValueError("请选择类别。")

                area = (request.form.get("storage_area") or "").strip()
                if not area:
                    raise ValueError("请填写仓储区。")

                qraw = (request.form.get("opening_qty") or "").strip()
                oq = Decimal(qraw)

                unit = (request.form.get("unit") or "").strip() or None
                remark = (request.form.get("remark") or "").strip() or None

                if cat == inventory_svc.INV_FINISHED:
                    pid = int((request.form.get("product_id") or "").strip())
                    if not Product.query.get(pid):
                        raise ValueError("产品无效。")
                    row = InventoryOpeningBalance(
                        category=inventory_svc.INV_FINISHED,
                        product_id=pid,
                        material_id=0,
                        storage_area=area[:32],
                        opening_qty=oq,
                        unit=unit[:16] if unit else None,
                        remark=remark[:255] if remark else None,
                    )
                else:
                    mid_raw = (request.form.get("material_id") or request.form.get("product_id") or "").strip()
                    mid = int(mid_raw) if mid_raw else 0
                    if not mid:
                        raise ValueError("请选择半成品/物料。")
                    item = SemiMaterial.query.get(mid)
                    if not item or item.kind != cat:
                        raise ValueError("半成品/物料无效。")
                    row = InventoryOpeningBalance(
                        category=cat,
                        product_id=0,
                        material_id=mid,
                        storage_area=area[:32],
                        opening_qty=oq,
                        unit=unit[:16] if unit else None,
                        remark=remark[:255] if remark else None,
                    )
                db.session.add(row)
                db.session.commit()
                flash("期初已保存。", "success")
                return redirect(url_for("main.inventory_opening_list"))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except InvalidOperation:
                db.session.rollback()
                flash("数量格式不正确。", "danger")
            except IntegrityError:
                db.session.rollback()
                flash("该维度已存在期初，请编辑原记录。", "danger")
        return render_template("inventory/opening_form.html", row=None)

    @bp.route("/inventory/opening/<int:opening_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["opening.edit"])
    def inventory_opening_edit(opening_id):
        row = InventoryOpeningBalance.query.options(
            selectinload(InventoryOpeningBalance.product),
            selectinload(InventoryOpeningBalance.material),
        ).get_or_404(opening_id)
        if request.method == "POST":
            try:
                qraw = (request.form.get("opening_qty") or "").strip()
                row.opening_qty = Decimal(qraw)
                row.unit = (request.form.get("unit") or "").strip()[:16] or None
                row.remark = (request.form.get("remark") or "").strip()[:255] or None
                db.session.commit()
                flash("已更新。", "success")
                return redirect(url_for("main.inventory_opening_list"))
            except InvalidOperation:
                db.session.rollback()
                flash("数量格式不正确。", "danger")
        return render_template("inventory/opening_form.html", row=row)

    @bp.route("/inventory/opening/<int:opening_id>/delete", methods=["POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["opening.delete"])
    def inventory_opening_delete(opening_id):
        row = InventoryOpeningBalance.query.get_or_404(opening_id)
        db.session.delete(row)
        db.session.commit()
        flash("已删除该条期初。", "success")
        return redirect(url_for("main.inventory_opening_list"))

    # ----- 历史：每日库存快照（旧功能） -----
    @bp.route("/inventory/daily")
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    def inventory_daily_list():
        page = request.args.get("page", 1, type=int)
        q = InventoryDailyRecord.query.options(
            selectinload(InventoryDailyRecord.lines)
        ).order_by(
            InventoryDailyRecord.record_date.desc(),
            InventoryDailyRecord.id.desc(),
        )
        pagination = q.paginate(page=page, per_page=20)
        cid_set = {r.created_by for r in pagination.items}
        creators = (
            {
                u.id: (u.name or u.username)
                for u in User.query.filter(User.id.in_(cid_set)).all()
            }
            if cid_set
            else {}
        )
        return render_template(
            "inventory/daily_list.html",
            pagination=pagination,
            creators=creators,
        )

    @bp.route("/inventory/daily/new", methods=["GET", "POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["daily.create"])
    def inventory_daily_new():
        if request.method == "POST":
            try:
                rd = request.form.get("record_date") or ""
                record_date = date.fromisoformat(rd) if rd else None
                if not record_date:
                    raise ValueError("请选择业务日期。")
                status = (request.form.get("status") or "confirmed").strip()
                if status not in ("draft", "confirmed"):
                    status = "confirmed"
                remark = (request.form.get("remark") or "").strip()[:500] or None
                rows = _parse_line_rows()
                if not rows:
                    raise ValueError("请至少录入一行产品数量。")
                _validate_products([r[0] for r in rows])
                rec = InventoryDailyRecord(
                    record_date=record_date,
                    status=status,
                    remark=remark,
                    created_by=current_user.id,
                )
                db.session.add(rec)
                db.session.flush()
                for product_id, qty, unit, note in rows:
                    db.session.add(
                        InventoryDailyLine(
                            header_id=rec.id,
                            product_id=product_id,
                            quantity=qty,
                            unit=unit,
                            note=note,
                        )
                    )
                db.session.commit()
                flash("每日库存录入已保存。", "success")
                return redirect(
                    url_for("main.inventory_daily_detail", record_id=rec.id)
                )
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：产品重复或数据冲突。", "danger")
        today = date.today().isoformat()
        return render_template(
            "inventory/daily_form.html",
            record=None,
            lines=None,
            default_record_date=today,
        )

    @bp.route("/inventory/daily/<int:record_id>")
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    def inventory_daily_detail(record_id):
        record = InventoryDailyRecord.query.options(
            selectinload(InventoryDailyRecord.lines).selectinload(
                InventoryDailyLine.product
            )
        ).get_or_404(record_id)
        creator = User.query.get(record.created_by)
        creator_label = (
            (creator.name or creator.username) if creator else str(record.created_by)
        )
        return render_template(
            "inventory/daily_detail.html",
            record=record,
            creator_label=creator_label,
        )

    @bp.route("/inventory/daily/<int:record_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["daily.edit"])
    def inventory_daily_edit(record_id):
        record = InventoryDailyRecord.query.options(
            selectinload(InventoryDailyRecord.lines).selectinload(
                InventoryDailyLine.product
            )
        ).get_or_404(record_id)
        if request.method == "POST":
            try:
                rd = request.form.get("record_date") or ""
                record_date = date.fromisoformat(rd) if rd else None
                if not record_date:
                    raise ValueError("请选择业务日期。")
                status = (request.form.get("status") or "confirmed").strip()
                if status not in ("draft", "confirmed"):
                    status = "confirmed"
                remark = (request.form.get("remark") or "").strip()[:500] or None
                rows = _parse_line_rows()
                if not rows:
                    raise ValueError("请至少录入一行产品数量。")
                _validate_products([r[0] for r in rows])
                record.record_date = record_date
                record.status = status
                record.remark = remark
                InventoryDailyLine.query.filter_by(header_id=record.id).delete(
                    synchronize_session=False
                )
                for product_id, qty, unit, note in rows:
                    db.session.add(
                        InventoryDailyLine(
                            header_id=record.id,
                            product_id=product_id,
                            quantity=qty,
                            unit=unit,
                            note=note,
                        )
                    )
                db.session.commit()
                flash("已更新。", "success")
                return redirect(
                    url_for("main.inventory_daily_detail", record_id=record.id)
                )
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：产品重复或数据冲突。", "danger")
        record = InventoryDailyRecord.query.options(
            selectinload(InventoryDailyRecord.lines).selectinload(
                InventoryDailyLine.product
            )
        ).get_or_404(record_id)
        lines = list(record.lines)
        return render_template(
            "inventory/daily_form.html",
            record=record,
            lines=lines,
            default_record_date=record.record_date.isoformat(),
        )

    @bp.route("/inventory/daily/<int:record_id>/delete", methods=["POST"])
    @login_required
    @menu_required(*INVENTORY_OPS_MENU_CODES)
    @capability_required(*INVENTORY_CAP_KEYS["daily.delete"])
    def inventory_daily_delete(record_id):
        record = InventoryDailyRecord.query.get_or_404(record_id)
        db.session.delete(record)
        db.session.commit()
        flash("已删除该条每日录入。", "success")
        return redirect(url_for("main.inventory_daily_list"))
