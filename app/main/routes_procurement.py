from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re

from flask import (
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from app import db
from app.auth.capabilities import current_user_can_cap
from app.auth.decorators import capability_required, menu_required
from app.models import (
    Company,
    InventoryMovement,
    PurchaseOrder,
    PurchaseReceipt,
    PurchaseRequisition,
    PurchaseRequisitionLine,
    PurchaseStockIn,
    SemiMaterial,
    Supplier,
    SupplierMaterialMap,
)
from app.utils.decimal_scale import json_decimal
from app.utils.procurement_order_excel import build_purchase_order_workbook

REQUISITION_STATUS = ("draft", "signed", "partial_ordered", "ordered", "cancelled")
PO_STATUS = ("draft", "ordered", "partially_received", "received", "cancelled")
RECEIPT_STATUS = ("draft", "posted")
RECONCILE_STATUS = ("pending", "matched", "exception")


def _as_decimal(value, default: str = "0") -> Decimal:
    if value is None:
        return Decimal(default)
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _companies():
    return Company.query.order_by(Company.id.asc()).all()


def _resolve_company_id():
    companies = _companies()
    return Company.get_default_id(), companies


def _next_no(
    prefix: str, model, field_name: str, width: int = 4, on_date: date | None = None
) -> str:
    current_day = on_date or date.today()
    start = f"{prefix}{current_day:%Y%m%d}"
    column = getattr(model, field_name)
    latest = (
        model.query.filter(column.like(f"{start}%"))
        .order_by(column.desc(), model.id.desc())
        .first()
    )
    seq = 1
    if latest:
        raw = getattr(latest, field_name) or ""
        tail = str(raw)[len(start) :]
        if tail.isdigit():
            seq = int(tail) + 1
    return f"{start}{seq:0{width}d}"


def _next_purchase_order_no(company: Company, on_date: date | None = None) -> str:
    if not company or not (company.code or "").strip():
        raise ValueError("默认主体缺少主体编号，无法生成采购单号。")
    current_day = on_date or date.today()
    prefix = f"{company.code.strip()}{current_day:%Y%m%d}"
    latest = (
        PurchaseOrder.query.filter(PurchaseOrder.po_no.like(f"{prefix}%"))
        .order_by(PurchaseOrder.po_no.desc(), PurchaseOrder.id.desc())
        .first()
    )
    seq = 1
    if latest and latest.po_no:
        tail = str(latest.po_no)[len(prefix) :]
        if tail.isdigit():
            seq = int(tail) + 1
    return f"{prefix}{seq:03d}"


def _parse_decimal(
    raw: str | None, field_name: str, *, allow_zero: bool = True
) -> Decimal:
    try:
        value = Decimal((raw or "").strip())
    except InvalidOperation as exc:
        raise ValueError(f"{field_name}格式不正确。") from exc
    if value < 0 or (not allow_zero and value <= 0):
        raise ValueError(f"{field_name}必须大于0。")
    return value


def _parse_optional_date(raw: str | None, field_name: str) -> date | None:
    text = (raw or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name}格式错误。") from exc


def _parse_datetime_local(raw: str | None, field_name: str) -> datetime:
    text = (raw or "").strip()
    if not text:
        raise ValueError(f"{field_name}不能为空。")
    try:
        return datetime.strptime(text, "%Y-%m-%dT%H:%M")
    except ValueError as exc:
        raise ValueError(f"{field_name}格式错误。") from exc


def _touch_supplier_material_map(
    *,
    company_id: int,
    supplier_id: int | None,
    material_id: int | None,
    unit_price: Decimal | None = None,
    create_missing: bool = False,
) -> SupplierMaterialMap | None:
    if not supplier_id or not material_id:
        return None
    mapping = SupplierMaterialMap.query.filter_by(
        company_id=company_id,
        supplier_id=supplier_id,
        material_id=material_id,
    ).first()
    if not mapping and create_missing:
        mapping = SupplierMaterialMap(
            company_id=company_id,
            supplier_id=supplier_id,
            material_id=material_id,
            is_active=True,
        )
        db.session.add(mapping)
    if not mapping:
        return None
    if unit_price is not None:
        mapping.last_unit_price = unit_price
    mapping.is_active = True
    return mapping


def _clear_other_default_supplier_flags(
    *, company_id: int, material_id: int, keep_mapping_id: int
) -> None:
    (
        SupplierMaterialMap.query.filter(
            SupplierMaterialMap.company_id == company_id,
            SupplierMaterialMap.material_id == material_id,
            SupplierMaterialMap.id != keep_mapping_id,
            SupplierMaterialMap.is_preferred.is_(True),
        ).update({SupplierMaterialMap.is_preferred: False}, synchronize_session=False)
    )


def _best_material_mapping(
    company_id: int | None, material_id: int
) -> SupplierMaterialMap | None:
    if not company_id:
        return None
    return (
        SupplierMaterialMap.query.options(selectinload(SupplierMaterialMap.supplier))
        .filter_by(company_id=company_id, material_id=material_id, is_active=True)
        .join(Supplier, Supplier.id == SupplierMaterialMap.supplier_id)
        .filter(Supplier.is_active.is_(True))
        .order_by(
            SupplierMaterialMap.is_preferred.desc(),
            SupplierMaterialMap.updated_at.desc(),
            SupplierMaterialMap.id.desc(),
        )
        .first()
    )


def _default_material_mapping(
    company_id: int | None, material_id: int
) -> SupplierMaterialMap | None:
    if not company_id:
        return None
    return (
        SupplierMaterialMap.query.options(selectinload(SupplierMaterialMap.supplier))
        .filter_by(
            company_id=company_id,
            material_id=material_id,
            is_active=True,
            is_preferred=True,
        )
        .join(Supplier, Supplier.id == SupplierMaterialMap.supplier_id)
        .filter(Supplier.is_active.is_(True))
        .order_by(
            SupplierMaterialMap.updated_at.desc(),
            SupplierMaterialMap.id.desc(),
        )
        .first()
    )


def _supplier_payload(
    row: Supplier,
    *,
    last_unit_price: Decimal | None = None,
    is_default_supplier: bool = False,
) -> dict:
    return {
        "id": row.id,
        "supplier_id": row.id,
        "name": row.name,
        "label": row.name,
        "contact_name": row.contact_name or "",
        "phone": row.phone or "",
        "address": row.address or "",
        "last_unit_price": json_decimal(last_unit_price) if last_unit_price is not None else "",
        "is_default_supplier": bool(is_default_supplier),
    }


def _material_payload(
    row: SemiMaterial,
    *,
    last_unit_price: Decimal | None = None,
    is_default_supplier: bool = False,
    supplier_id: int | None = None,
) -> dict:
    label = row.name
    if row.spec:
        label = f"{row.name} / {row.spec}"
    return {
        "id": row.id,
        "material_id": row.id,
        "name": row.name,
        "label": label,
        "spec": row.spec or "",
        "base_unit": row.base_unit or "",
        "last_unit_price": json_decimal(last_unit_price) if last_unit_price is not None else "",
        "is_default_supplier": bool(is_default_supplier),
        "supplier_id": supplier_id,
    }


def _supplier_search_query(
    company_id: int, keyword: str = "", *, active_only: bool = True
):
    q = Supplier.query.filter(Supplier.company_id == company_id)
    if active_only:
        q = q.filter(Supplier.is_active.is_(True))
    if keyword:
        q = q.filter(
            or_(
                Supplier.name.contains(keyword),
                Supplier.contact_name.contains(keyword),
                Supplier.phone.contains(keyword),
                Supplier.address.contains(keyword),
            )
        )
    return q.order_by(Supplier.name.asc(), Supplier.id.asc())


def _material_search_query(keyword: str = ""):
    q = SemiMaterial.query.filter(SemiMaterial.kind == "material")
    if keyword:
        q = q.filter(
            or_(
                SemiMaterial.code.contains(keyword),
                SemiMaterial.name.contains(keyword),
                SemiMaterial.spec.contains(keyword),
            )
        )
    return q.order_by(SemiMaterial.name.asc(), SemiMaterial.id.asc())


def _next_material_code() -> str:
    prefix = "MT"
    m = 0
    for (code,) in (
        db.session.query(SemiMaterial.code)
        .filter(SemiMaterial.code.like(f"{prefix}%"))
        .all()
    ):
        if not code:
            continue
        raw = str(code)
        if not raw.startswith(prefix):
            continue
        tail = raw[len(prefix) :]
        if tail.isdigit():
            try:
                m = max(m, int(tail))
            except ValueError:
                pass
    return f"{prefix}{m + 1:04d}"


def _bump_material_code(code: str) -> str:
    raw = (code or "").strip()
    matched = re.match(r"^(MT)(\d+)$", raw)
    if not matched:
        return raw + "N"
    prefix = matched.group(1)
    tail = matched.group(2)
    try:
        return f"{prefix}{int(tail) + 1:04d}"
    except ValueError:
        return raw + "N"


def _require_supplier_material_mapping(
    company_id: int,
    supplier_id: int,
    material_id: int,
    *,
    row_label: str,
) -> tuple[Supplier, SemiMaterial, SupplierMaterialMap]:
    supplier = db.session.get(Supplier, supplier_id)
    material = db.session.get(SemiMaterial, material_id)
    if not supplier or supplier.company_id != company_id or not supplier.is_active:
        raise ValueError(f"{row_label}供应商无效或已停用。")
    if not material or material.kind != "material":
        raise ValueError(f"{row_label}物料无效。")
    mapping = SupplierMaterialMap.query.filter_by(
        company_id=company_id,
        supplier_id=supplier_id,
        material_id=material_id,
        is_active=True,
    ).first()
    if not mapping:
        raise ValueError(f"{row_label}供应商与物料尚未建立关联。")
    return supplier, material, mapping


def _save_supplier_material_maps(
    supplier: Supplier, rows: list[dict]
) -> list[SupplierMaterialMap]:
    if not supplier.is_active and any(item["is_preferred"] for item in rows):
        raise ValueError("停用供应商不能设置为默认供应商。")
    SupplierMaterialMap.query.filter_by(supplier_id=supplier.id).delete(
        synchronize_session=False
    )
    saved_rows: list[SupplierMaterialMap] = []
    for item in rows:
        mapping = SupplierMaterialMap(
            company_id=supplier.company_id,
            supplier_id=supplier.id,
            material_id=item["material_id"],
            is_preferred=item["is_preferred"],
            is_active=True,
            last_unit_price=item["last_unit_price"],
            remark=item["remark"],
        )
        db.session.add(mapping)
        saved_rows.append(mapping)
    db.session.flush()
    for mapping in saved_rows:
        if mapping.is_preferred:
            _clear_other_default_supplier_flags(
                company_id=supplier.company_id,
                material_id=mapping.material_id,
                keep_mapping_id=mapping.id,
            )
    return saved_rows


def _material_items_for_supplier(
    company_id: int,
    supplier_id: int,
    *,
    keyword: str = "",
    limit: int = 20,
) -> list[dict]:
    rows = (
        SupplierMaterialMap.query.options(selectinload(SupplierMaterialMap.material))
        .filter_by(company_id=company_id, supplier_id=supplier_id, is_active=True)
        .join(SemiMaterial, SemiMaterial.id == SupplierMaterialMap.material_id)
        .filter(SemiMaterial.kind == "material")
    )
    if keyword:
        rows = rows.filter(
            or_(
                SemiMaterial.code.contains(keyword),
                SemiMaterial.name.contains(keyword),
                SemiMaterial.spec.contains(keyword),
            )
        )
    items = (
        rows.order_by(SupplierMaterialMap.is_preferred.desc(), SemiMaterial.name.asc())
        .limit(limit)
        .all()
    )
    return [
        _material_payload(
            item.material,
            last_unit_price=_as_decimal(item.last_unit_price)
            if item.last_unit_price is not None
            else None,
            is_default_supplier=bool(item.is_preferred),
            supplier_id=item.supplier_id,
        )
        for item in items
        if item.material
    ]


def _supplier_items_for_material(
    company_id: int,
    material_id: int,
    *,
    keyword: str = "",
    limit: int = 20,
) -> list[dict]:
    rows = (
        SupplierMaterialMap.query.options(selectinload(SupplierMaterialMap.supplier))
        .filter_by(company_id=company_id, material_id=material_id, is_active=True)
        .join(Supplier, Supplier.id == SupplierMaterialMap.supplier_id)
        .filter(Supplier.is_active.is_(True))
    )
    if keyword:
        rows = rows.filter(
            or_(
                Supplier.name.contains(keyword),
                Supplier.contact_name.contains(keyword),
                Supplier.phone.contains(keyword),
                Supplier.address.contains(keyword),
            )
        )
    items = (
        rows.order_by(SupplierMaterialMap.is_preferred.desc(), Supplier.name.asc())
        .limit(limit)
        .all()
    )
    return [
        _supplier_payload(
            item.supplier,
            last_unit_price=_as_decimal(item.last_unit_price)
            if item.last_unit_price is not None
            else None,
            is_default_supplier=bool(item.is_preferred),
        )
        for item in items
        if item.supplier
    ]


def _material_default_supplier_state(
    company_id: int | None, material_id: int | None
) -> tuple[list[dict], int | None, str]:
    if not company_id or not material_id:
        return [], None, ""
    options = _supplier_items_for_material(company_id, material_id, limit=200)
    default_mapping = _default_material_mapping(company_id, material_id)
    if not default_mapping:
        return options, None, ""
    default_price = (
        json_decimal(_as_decimal(default_mapping.last_unit_price))
        if default_mapping.last_unit_price is not None
        else ""
    )
    return options, default_mapping.supplier_id, default_price


def _set_default_supplier_for_material(
    *, company_id: int, material_id: int, supplier_id: int | None
) -> None:
    if not supplier_id:
        (
            SupplierMaterialMap.query.filter_by(
                company_id=company_id,
                material_id=material_id,
            ).update(
                {SupplierMaterialMap.is_preferred: False},
                synchronize_session=False,
            )
        )
        return
    mapping = (
        SupplierMaterialMap.query.join(
            Supplier, Supplier.id == SupplierMaterialMap.supplier_id
        )
        .filter(
            SupplierMaterialMap.company_id == company_id,
            SupplierMaterialMap.material_id == material_id,
            SupplierMaterialMap.supplier_id == supplier_id,
            SupplierMaterialMap.is_active.is_(True),
            Supplier.is_active.is_(True),
        )
        .first()
    )
    if not mapping:
        raise ValueError("默认供应商必须从当前主体已关联且启用的供应商中选择。")
    mapping.is_preferred = True
    db.session.flush()
    _clear_other_default_supplier_flags(
        company_id=company_id,
        material_id=material_id,
        keep_mapping_id=mapping.id,
    )


def _sync_requisition_summary(row: PurchaseRequisition) -> None:
    lines = list(row.lines or [])
    if not lines:
        return
    row.qty = sum((_as_decimal(line.qty) for line in lines), Decimal("0"))
    row.expected_date = next(
        (line.expected_date for line in lines if line.expected_date), None
    )
    supplier_names = [line.supplier_name for line in lines if line.supplier_name]
    first_line = lines[0]
    unique_suppliers = list(dict.fromkeys(supplier_names))
    row.supplier_name = (
        unique_suppliers[0]
        if len(unique_suppliers) == 1
        else f"{unique_suppliers[0]}等{len(unique_suppliers)}家"
    )
    row.item_name = (
        first_line.item_name
        if len(lines) == 1
        else f"{first_line.item_name}等{len(lines)}项"
    )
    row.item_spec = first_line.item_spec if len(lines) == 1 else None
    row.unit = (
        first_line.unit
        if len({line.unit for line in lines if line.unit}) == 1
        else "项"
    )


def _sync_requisition_line_status(line: PurchaseRequisitionLine) -> None:
    req = line.requisition
    if req and req.status == "cancelled":
        line.status = "cancelled"
        return
    active_order = line.purchase_orders.filter(
        PurchaseOrder.status != "cancelled"
    ).first()
    line.status = "ordered" if active_order else "pending_order"


def _sync_requisition_status(row: PurchaseRequisition) -> None:
    lines = list(row.lines or [])
    if row.status == "cancelled":
        return
    if not lines:
        row.status = "draft"
        return
    ordered_count = sum(1 for line in lines if line.status == "ordered")
    if ordered_count == len(lines):
        row.status = "ordered"
    elif ordered_count > 0:
        row.status = "partial_ordered"
    elif row.signed_at:
        row.status = "signed"
    else:
        row.status = "draft"


def _sum_posted_receipt_qty(po: PurchaseOrder) -> Decimal:
    qty = (
        db.session.query(func.coalesce(func.sum(PurchaseReceipt.received_qty), 0))
        .filter(
            PurchaseReceipt.purchase_order_id == po.id,
            PurchaseReceipt.status == "posted",
        )
        .scalar()
    )
    return _as_decimal(qty)


def _warehouse_query_for_order(po: PurchaseOrder):
    receipt_ids = [
        rid
        for (rid,) in db.session.query(PurchaseReceipt.id)
        .filter(PurchaseReceipt.purchase_order_id == po.id)
        .all()
    ]
    filters = [InventoryMovement.source_purchase_order_id == po.id]
    if receipt_ids:
        filters.append(InventoryMovement.source_purchase_receipt_id.in_(receipt_ids))
    return InventoryMovement.query.filter(
        InventoryMovement.direction == "in", or_(*filters)
    )


def _sum_warehouse_qty(po: PurchaseOrder) -> Decimal:
    qty = (
        _warehouse_query_for_order(po)
        .with_entities(func.coalesce(func.sum(InventoryMovement.quantity), 0))
        .scalar()
    )
    return _as_decimal(qty)


def _latest_storage_area(
    po: PurchaseOrder, receipt: PurchaseReceipt | None = None
) -> str | None:
    q = _warehouse_query_for_order(po)
    if receipt:
        q = q.filter(
            or_(
                InventoryMovement.source_purchase_receipt_id == receipt.id,
                InventoryMovement.source_purchase_order_id == po.id,
            )
        )
    movement = q.order_by(InventoryMovement.id.desc()).first()
    return movement.storage_area if movement and movement.storage_area else None


def _build_compare_summary(receipt: PurchaseReceipt) -> dict:
    po = receipt.purchase_order
    ordered_qty = _as_decimal(po.qty)
    received_qty = _sum_posted_receipt_qty(po)
    warehouse_qty = _sum_warehouse_qty(po)
    receipt_warehouse_qty = _as_decimal(
        InventoryMovement.query.filter(
            InventoryMovement.direction == "in",
            InventoryMovement.source_purchase_receipt_id == receipt.id,
        )
        .with_entities(func.coalesce(func.sum(InventoryMovement.quantity), 0))
        .scalar()
    )
    return {
        "ordered_qty": ordered_qty,
        "received_qty": received_qty,
        "warehouse_qty": warehouse_qty,
        "receipt_warehouse_qty": receipt_warehouse_qty,
        "variance_qty": warehouse_qty - received_qty,
        "is_matched": ordered_qty == received_qty == warehouse_qty,
        "category": (
            po.material.kind
            if po.material and po.material.kind in ("semi", "material")
            else "material"
        ),
    }


def _sync_purchase_order_status(po: PurchaseOrder) -> None:
    if po.status == "cancelled":
        return
    posted_qty = _sum_posted_receipt_qty(po)
    if posted_qty >= _as_decimal(po.qty) and _as_decimal(po.qty) > 0:
        po.status = "received"
    elif posted_qty > 0:
        po.status = "partially_received"
    elif po.ordered_at:
        po.status = "ordered"
    else:
        po.status = "draft"


def _sync_purchase_order_reconcile_status(po: PurchaseOrder) -> None:
    ordered_qty = _as_decimal(po.qty)
    received_qty = _sum_posted_receipt_qty(po)
    warehouse_qty = _sum_warehouse_qty(po)
    if ordered_qty == 0 and received_qty == 0 and warehouse_qty == 0:
        po.reconcile_status = "pending"
    elif ordered_qty == received_qty == warehouse_qty:
        po.reconcile_status = "matched"
    else:
        po.reconcile_status = (
            "exception" if (received_qty > 0 or warehouse_qty > 0) else "pending"
        )


def _update_requisition_from_order(po: PurchaseOrder | None) -> None:
    if not po or not po.requisition_line:
        return
    _sync_requisition_line_status(po.requisition_line)
    if po.requisition:
        _sync_requisition_summary(po.requisition)
        _sync_requisition_status(po.requisition)


def _parse_supplier_map_rows(company_id: int) -> list[dict]:
    material_ids = request.form.getlist("map_material_id")
    prices = request.form.getlist("map_last_unit_price")
    remarks = request.form.getlist("map_remark")
    preferred_values = request.form.getlist("map_is_preferred")
    rows: list[dict] = []
    seen_material_ids: set[int] = set()
    for idx, raw_material_id in enumerate(material_ids):
        raw_material_id = (raw_material_id or "").strip()
        raw_price = (prices[idx] if idx < len(prices) else "").strip()
        raw_remark = ((remarks[idx] if idx < len(remarks) else "") or "").strip()
        raw_preferred = (
            preferred_values[idx] if idx < len(preferred_values) else "0"
        ) == "1"
        if not raw_material_id:
            if raw_price or raw_remark or raw_preferred:
                raise ValueError(f"供应物料第 {idx + 1} 行必须从物料表选择物料。")
            continue
        material_id = int(raw_material_id)
        if material_id in seen_material_ids:
            raise ValueError("同一个供应商不能重复维护同一物料。")
        material = db.session.get(SemiMaterial, material_id)
        if not material or material.kind != "material":
            raise ValueError("供应商关联物料存在无效物料。")
        price = None
        if raw_price:
            price = _parse_decimal(raw_price, "最近单价")
        rows.append(
            {
                "company_id": company_id,
                "material_id": material_id,
                "is_preferred": raw_preferred,
                "remark": raw_remark[:500] or None,
                "last_unit_price": price,
            }
        )
        seen_material_ids.add(material_id)
    return rows


def _save_supplier_material_maps(
    supplier: Supplier, rows: list[dict]
) -> list[SupplierMaterialMap]:
    if not supplier.is_active and any(item.get("is_preferred") is True for item in rows):
        raise ValueError("停用供应商不能设置为默认供应商。")
    existing_rows = SupplierMaterialMap.query.filter_by(supplier_id=supplier.id).all()
    existing_by_material_id = {item.material_id: item for item in existing_rows}
    keep_material_ids: set[int] = set()
    saved_rows: list[SupplierMaterialMap] = []
    for item in rows:
        keep_material_ids.add(item["material_id"])
        mapping = existing_by_material_id.get(item["material_id"])
        if not mapping:
            mapping = SupplierMaterialMap(
                company_id=supplier.company_id,
                supplier_id=supplier.id,
                material_id=item["material_id"],
            )
            db.session.add(mapping)
        mapping.company_id = supplier.company_id
        mapping.is_active = True
        mapping.last_unit_price = item["last_unit_price"]
        mapping.remark = item["remark"]
        preferred_flag = item.get("is_preferred")
        if not supplier.is_active:
            mapping.is_preferred = False
        elif preferred_flag is not None:
            mapping.is_preferred = bool(preferred_flag)
        elif mapping.id is None:
            mapping.is_preferred = False
        saved_rows.append(mapping)
    for material_id, mapping in existing_by_material_id.items():
        if material_id not in keep_material_ids:
            db.session.delete(mapping)
    db.session.flush()
    for mapping in saved_rows:
        if mapping.is_preferred:
            _clear_other_default_supplier_flags(
                company_id=supplier.company_id,
                material_id=mapping.material_id,
                keep_mapping_id=mapping.id,
            )
    return saved_rows


def _upsert_supplier_material_maps_no_delete(
    supplier: Supplier, rows: list[dict]
) -> list[SupplierMaterialMap]:
    """
    upsert-only：只插入/更新 rows 中出现的 supplier-material 映射，不删除数据库中其他映射。

    默认供应商语义：
    - `rows[i]["is_preferred"] is True`：设置为默认，并清理同一物料下其它默认供应商。
    - `rows[i]["is_preferred"] is False`：显式取消默认标记，但不影响其它默认供应商。
    - `rows[i]["is_preferred"] is None`：不修改 is_preferred（仅对“新建映射”保持非默认）。
    """

    if not supplier.is_active and any(item.get("is_preferred") is True for item in rows):
        raise ValueError("停用供应商不能设置为默认供应商。")

    existing_rows = SupplierMaterialMap.query.filter_by(supplier_id=supplier.id).all()
    existing_by_material_id = {item.material_id: item for item in existing_rows}

    saved_rows: list[SupplierMaterialMap] = []
    for item in rows:
        material_id = item["material_id"]
        mapping = existing_by_material_id.get(material_id)
        if not mapping:
            mapping = SupplierMaterialMap(
                company_id=supplier.company_id,
                supplier_id=supplier.id,
                material_id=material_id,
            )
            db.session.add(mapping)
            existing_by_material_id[material_id] = mapping

        mapping.company_id = supplier.company_id
        mapping.is_active = True
        mapping.last_unit_price = item["last_unit_price"]
        mapping.remark = item["remark"]

        preferred_flag = item.get("is_preferred")
        if not supplier.is_active:
            mapping.is_preferred = False
        elif preferred_flag is not None:
            mapping.is_preferred = bool(preferred_flag)
        elif mapping.id is None:
            mapping.is_preferred = False

        saved_rows.append(mapping)

    db.session.flush()
    for mapping in saved_rows:
        if mapping.is_preferred:
            _clear_other_default_supplier_flags(
                company_id=supplier.company_id,
                material_id=mapping.material_id,
                keep_mapping_id=mapping.id,
            )
    return saved_rows


def _parse_supplier_map_rows(company_id: int) -> list[dict]:
    material_ids = request.form.getlist("map_material_id")
    prices = request.form.getlist("map_last_unit_price")
    remarks = request.form.getlist("map_remark")
    has_preferred_values = "map_is_preferred" in request.form
    preferred_values = request.form.getlist("map_is_preferred")
    rows: list[dict] = []
    seen_material_ids: set[int] = set()
    for idx, raw_material_id in enumerate(material_ids):
        raw_material_id = (raw_material_id or "").strip()
        raw_price = (prices[idx] if idx < len(prices) else "").strip()
        raw_remark = ((remarks[idx] if idx < len(remarks) else "") or "").strip()
        raw_preferred = (
            (preferred_values[idx] if idx < len(preferred_values) else "0") == "1"
            if has_preferred_values
            else None
        )
        if not raw_material_id:
            if raw_price or raw_remark or raw_preferred:
                raise ValueError(f"供应物料第 {idx + 1} 行必须从物料表选择物料。")
            continue
        material_id = int(raw_material_id)
        if material_id in seen_material_ids:
            raise ValueError("同一个供应商不能重复维护同一物料。")
        material = db.session.get(SemiMaterial, material_id)
        if not material or material.kind != "material":
            raise ValueError("供应商关联物料存在无效物料。")
        price = None
        if raw_price:
            price = _parse_decimal(raw_price, "最近单价")
        rows.append(
            {
                "company_id": company_id,
                "material_id": material_id,
                "is_preferred": raw_preferred,
                "remark": raw_remark[:500] or None,
                "last_unit_price": price,
            }
        )
        seen_material_ids.add(material_id)
    return rows


def _parse_supplier_import_excel_ws(ws, company_id: int) -> tuple[dict, list[str]]:
    """
    解析供应商 Excel 导入（含供应商-物料映射）。

    模板列顺序（从第 1 行表头开始）：
    1 供应商名称（必填）
    2 联系人
    3 电话
    4 地址
    5 供应商状态（1启用/0停用，可空默认启用）
    6 供应商备注
    7 物料编号（SemiMaterial.code，必填）
    8 最近单价（可空）
    9 物料-供应商备注
    10 是否默认供应商（1/0，可空=不修改默认）
    """

    def _cell_to_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            return str(v)
        return str(v).strip()

    def _parse_int01(raw, *, field_name: str) -> bool:
        s = _cell_to_str(raw)
        if not s:
            raise ValueError(f"{field_name}不能为空。")
        if s in ("1", "true", "True", "启用"):
            return True
        if s in ("0", "false", "False", "停用"):
            return False
        if s.isdigit() and s in ("1", "0"):
            return s == "1"
        raise ValueError(f"{field_name}格式错误，请填 1/0。")

    def _parse_optional_int01(raw, *, field_name: str) -> bool | None:
        s = _cell_to_str(raw)
        if not s:
            return None
        if s in ("1", "true", "True", "是", "默认", "启用"):
            return True
        if s in ("0", "false", "False", "否", "非默认", "停用"):
            return False
        if s.isdigit() and s in ("1", "0"):
            return s == "1"
        raise ValueError(f"{field_name}格式错误，请填 1/0 或留空。")

    def _parse_optional_decimal(raw, *, field_name: str) -> Decimal | None:
        s = _cell_to_str(raw)
        if not s:
            return None
        return _parse_decimal(s, field_name)

    suppliers: dict = {}
    errors: list[str] = []
    seen_supplier_material: set[tuple[str, str]] = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = (tuple(row) + (None,) * 10)[:10]
        (
            supplier_name,
            contact_name,
            phone,
            address,
            supplier_status,
            supplier_remark,
            material_code,
            last_unit_price,
            map_remark,
            is_preferred,
        ) = vals

        supplier_name_s = _cell_to_str(supplier_name)[:128]
        if not supplier_name_s:
            errors.append(f"第 {idx} 行：供应商名称不能为空。")
            continue

        try:
            # 供应商状态：可空默认启用
            supplier_status_s = _cell_to_str(supplier_status)
            is_active = (
                True if not supplier_status_s else _parse_int01(supplier_status, field_name="供应商状态")
            )
            contact_name_s = _cell_to_str(contact_name)[:64] or None
            phone_s = _cell_to_str(phone)[:32] or None
            address_s = _cell_to_str(address)[:255] or None
            supplier_remark_s = _cell_to_str(supplier_remark)[:500] or None
        except ValueError as exc:
            errors.append(f"第 {idx} 行：{str(exc)}")
            continue

        material_code_s = _cell_to_str(material_code)
        if not material_code_s:
            # 允许该行只更新供应商基础信息，不更新映射
            errors.append(f"第 {idx} 行：物料编号不能为空。")
            suppliers.setdefault(
                supplier_name_s,
                {"base": {"is_active": is_active, "contact_name": contact_name_s, "phone": phone_s, "address": address_s, "remark": supplier_remark_s}, "maps": []},
            )
            # 若已存在则覆盖基础字段（后出现的行优先）
            if supplier_name_s in suppliers:
                suppliers[supplier_name_s]["base"] = {
                    "is_active": is_active,
                    "contact_name": contact_name_s,
                    "phone": phone_s,
                    "address": address_s,
                    "remark": supplier_remark_s,
                }
            continue

        key = (supplier_name_s, material_code_s)
        if key in seen_supplier_material:
            errors.append(f"第 {idx} 行：同一供应商的同一物料在文件中重复。")
            continue
        seen_supplier_material.add(key)

        material = SemiMaterial.query.filter_by(code=material_code_s).first()
        if not material or material.kind != "material":
            errors.append(f"第 {idx} 行：物料编号 {material_code_s} 不存在或不是物料。")
            continue

        try:
            price = _parse_optional_decimal(last_unit_price, field_name="最近单价")
            mapping_remark_s = _cell_to_str(map_remark)[:500] or None
            preferred_flag = _parse_optional_int01(
                is_preferred, field_name="是否默认供应商"
            )
        except ValueError as exc:
            errors.append(f"第 {idx} 行：{str(exc)}")
            continue

        supplier = suppliers.get(supplier_name_s)
        if not supplier:
            supplier = {
                "base": {
                    "is_active": is_active,
                    "contact_name": contact_name_s,
                    "phone": phone_s,
                    "address": address_s,
                    "remark": supplier_remark_s,
                },
                "maps": [],
            }
            suppliers[supplier_name_s] = supplier
        else:
            # 后出现的行优先覆盖供应商基础字段
            supplier["base"] = {
                "is_active": is_active,
                "contact_name": contact_name_s,
                "phone": phone_s,
                "address": address_s,
                "remark": supplier_remark_s,
            }

        supplier["maps"].append(
            {
                "company_id": company_id,
                "material_id": material.id,
                "is_preferred": preferred_flag,  # None 表示不修改默认标记
                "remark": mapping_remark_s,
                "last_unit_price": price,
            }
        )

    return suppliers, errors


def _parse_requisition_lines(company_id: int) -> list[dict]:
    supplier_ids = request.form.getlist("line_supplier_id")
    material_ids = request.form.getlist("line_material_id")
    qtys = request.form.getlist("line_qty")
    units = request.form.getlist("line_unit")
    expected_dates = request.form.getlist("line_expected_date")
    remarks = request.form.getlist("line_remark")
    rows: list[dict] = []
    for idx, raw_supplier_id in enumerate(supplier_ids):
        raw_supplier_id = (raw_supplier_id or "").strip()
        raw_material_id = (material_ids[idx] if idx < len(material_ids) else "").strip()
        raw_qty = (qtys[idx] if idx < len(qtys) else "").strip()
        raw_unit = ((units[idx] if idx < len(units) else "") or "").strip()
        raw_expected = (
            expected_dates[idx] if idx < len(expected_dates) else ""
        ).strip()
        raw_remark = ((remarks[idx] if idx < len(remarks) else "") or "").strip()
        if not any(
            [
                raw_supplier_id,
                raw_material_id,
                raw_qty,
                raw_unit,
                raw_expected,
                raw_remark,
            ]
        ):
            continue
        if not raw_supplier_id or not raw_material_id:
            raise ValueError(f"请购第 {idx + 1} 行必须同时选择供应商和物料。")
        try:
            supplier_id = int(raw_supplier_id)
            material_id = int(raw_material_id)
        except ValueError as exc:
            raise ValueError(f"请购第 {idx + 1} 行存在无效供应商或物料。") from exc
        supplier, material, _mapping = _require_supplier_material_mapping(
            company_id,
            supplier_id,
            material_id,
            row_label=f"请购第 {idx + 1} 行",
        )
        qty = _parse_decimal(raw_qty, f"请购第 {idx + 1} 行数量", allow_zero=False)
        rows.append(
            {
                "company_id": company_id,
                "line_no": len(rows) + 1,
                "supplier_id": supplier_id,
                "material_id": material_id,
                "supplier_name": supplier.name,
                "item_name": material.name,
                "item_spec": material.spec,
                "qty": qty,
                "unit": raw_unit or material.base_unit or "pcs",
                "expected_date": _parse_optional_date(
                    raw_expected, f"请购第 {idx + 1} 行到货日期"
                ),
                "remark": raw_remark[:500] or None,
            }
        )
    if not rows:
        raise ValueError("请至少录入一行请购明细。")
    return rows


def create_purchase_orders_from_requisition(
    requisition: PurchaseRequisition,
    *,
    buyer_user_id: int,
    order_date: date | None = None,
) -> list[PurchaseOrder]:
    if not requisition.signed_at:
        raise ValueError("请购单签字后才能生成采购单。")
    company = requisition.company or Company.query.get(requisition.company_id)
    if not company:
        raise ValueError("请购单缺少有效主体。")
    created_orders: list[PurchaseOrder] = []
    for line in requisition.lines:
        _sync_requisition_line_status(line)
        if line.status == "ordered":
            continue
        supplier = (
            db.session.get(Supplier, line.supplier_id) if line.supplier_id else None
        )
        material = (
            db.session.get(SemiMaterial, line.material_id) if line.material_id else None
        )
        if not supplier or not material:
            raise ValueError("请购明细存在缺失的供应商或物料，无法生成采购单。")
        mapping = SupplierMaterialMap.query.filter_by(
            company_id=requisition.company_id,
            supplier_id=line.supplier_id,
            material_id=line.material_id,
        ).first()
        unit_price = (
            _as_decimal(mapping.last_unit_price)
            if mapping and mapping.last_unit_price is not None
            else Decimal("0")
        )
        order = PurchaseOrder(
            company_id=requisition.company_id,
            po_no=_next_purchase_order_no(company, order_date),
            requisition_id=requisition.id,
            requisition_line_id=line.id,
            buyer_user_id=buyer_user_id,
            supplier_id=line.supplier_id,
            material_id=line.material_id,
            supplier_name=line.supplier_name,
            supplier_contact_name=supplier.contact_name,
            supplier_phone=supplier.phone,
            supplier_address=supplier.address,
            item_name=line.item_name,
            item_spec=line.item_spec,
            qty=line.qty,
            unit=line.unit,
            unit_price=unit_price,
            amount=_as_decimal(line.qty) * unit_price,
            expected_date=line.expected_date,
            status="draft",
            reconcile_status="pending",
            remark=line.remark,
        )
        db.session.add(order)
        created_orders.append(order)
        _touch_supplier_material_map(
            company_id=requisition.company_id,
            supplier_id=line.supplier_id,
            material_id=line.material_id,
            unit_price=unit_price,
        )
    db.session.flush()
    for line in requisition.lines:
        _sync_requisition_line_status(line)
    _sync_requisition_summary(requisition)
    _sync_requisition_status(requisition)
    return created_orders


def register_procurement_routes(bp):
    @bp.route("/procurement/materials")
    @login_required
    @menu_required("procurement_material")
    def procurement_material_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        q = _material_search_query(
            keyword if current_user_can_cap("procurement_material.filter.keyword") else ""
        )
        page = request.args.get("page", 1, type=int)
        pagination = q.paginate(page=page, per_page=20)
        rows = [
            {
                "item": item,
                "default_mapping": _default_material_mapping(company_id, item.id),
            }
            for item in pagination.items
        ]
        return render_template(
            "procurement/material_list.html",
            rows=rows,
            pagination=pagination,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
        )

    @bp.route("/procurement/materials/new", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_material")
    @capability_required("procurement_material.action.create")
    def procurement_material_new():
        company_id, companies = _resolve_company_id()
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            spec = (request.form.get("spec") or "").strip() or None
            base_unit = (request.form.get("base_unit") or "").strip() or None
            remark = (request.form.get("remark") or "").strip() or None
            if not name:
                flash("名称为必填。", "danger")
                return render_template(
                    "procurement/material_form.html",
                    item=None,
                    companies=companies,
                    company_id=company_id,
                    supplier_options=[],
                    default_supplier_id=None,
                    default_supplier_price="",
                )
            item = SemiMaterial(kind="material")
            item.code = _next_material_code()
            item.name = name
            item.spec = spec
            item.base_unit = base_unit
            item.remark = remark
            max_tries = 3
            for attempt in range(max_tries):
                db.session.add(item)
                try:
                    db.session.commit()
                    flash("物料已保存。", "success")
                    return redirect(url_for("main.procurement_material_list"))
                except IntegrityError:
                    db.session.rollback()
                    db.session.expunge_all()
                    if item.code:
                        item.code = _bump_material_code(item.code)
                    db.session.add(item)
                    if attempt == max_tries - 1:
                        flash("保存失败：编码冲突，请稍后重试。", "danger")
        return render_template(
            "procurement/material_form.html",
            item=None,
            companies=companies,
            company_id=company_id,
            supplier_options=[],
            default_supplier_id=None,
            default_supplier_price="",
        )

    @bp.route("/procurement/materials/<int:item_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_material")
    @capability_required("procurement_material.action.edit")
    def procurement_material_edit(item_id: int):
        item = SemiMaterial.query.get_or_404(item_id)
        if item.kind != "material":
            abort(404)
        company_id, companies = _resolve_company_id()
        supplier_options, default_supplier_id, default_supplier_price = (
            _material_default_supplier_state(company_id, item.id)
        )
        if request.method == "POST":
            try:
                name = (request.form.get("name") or "").strip()
                spec = (request.form.get("spec") or "").strip() or None
                base_unit = (request.form.get("base_unit") or "").strip() or None
                remark = (request.form.get("remark") or "").strip() or None
                default_supplier_raw = (
                    request.form.get("default_supplier_id") or ""
                ).strip()
                selected_supplier_id = None
                if default_supplier_raw:
                    try:
                        selected_supplier_id = int(default_supplier_raw)
                    except ValueError as exc:
                        raise ValueError("默认供应商选择无效。") from exc
                item.name = name
                item.spec = spec
                item.base_unit = base_unit
                item.remark = remark
                if not item.name:
                    raise ValueError("名称为必填。")
                db.session.add(item)
                db.session.flush()
                if company_id:
                    _set_default_supplier_for_material(
                        company_id=company_id,
                        material_id=item.id,
                        supplier_id=selected_supplier_id,
                    )
                elif selected_supplier_id:
                    raise ValueError("请先设置默认经营主体后再指定默认供应商。")
                db.session.commit()
                flash("物料已更新。", "success")
                return redirect(url_for("main.procurement_material_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：数据冲突。", "danger")
            supplier_options, default_supplier_id, default_supplier_price = (
                _material_default_supplier_state(company_id, item.id)
            )
        return render_template(
            "procurement/material_form.html",
            item=item,
            companies=companies,
            company_id=company_id,
            supplier_options=supplier_options,
            default_supplier_id=default_supplier_id,
            default_supplier_price=default_supplier_price,
        )

    @bp.route("/procurement/materials/<int:item_id>/delete", methods=["POST"])
    @login_required
    @menu_required("procurement_material")
    @capability_required("procurement_material.action.delete")
    def procurement_material_delete(item_id: int):
        item = SemiMaterial.query.get_or_404(item_id)
        if item.kind != "material":
            abort(404)
        SupplierMaterialMap.query.filter_by(material_id=item.id).delete(
            synchronize_session=False
        )
        db.session.delete(item)
        db.session.commit()
        flash("物料已删除。", "success")
        return redirect(url_for("main.procurement_material_list"))

    @bp.route("/procurement/materials/export-import-template", methods=["GET"])
    @login_required
    @menu_required("procurement_material")
    @capability_required("procurement_material.action.import")
    def procurement_material_export_import_template():
        from io import BytesIO

        from openpyxl import Workbook

        headers = ["物料编号（可留空）", "名称", "规格", "基础单位", "备注"]
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col, header)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="采购物料导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/procurement/materials/import", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_material")
    @capability_required("procurement_material.action.import")
    def procurement_material_import():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("procurement/material_import.html", result=None)
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("procurement/material_import.html", result=None)
            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("procurement/material_import.html", result=None)

            success = 0
            errors = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                code, name, spec, base_unit, remark = (row + (None,) * 5)[:5]
                code = (code or "").strip() if isinstance(code, str) else (code or "")
                code = str(code).strip()
                name = (name or "").strip() if isinstance(name, str) else (name or "")
                name = str(name).strip()
                if not name:
                    if any(row):
                        errors.append(f"第 {idx} 行：名称为空")
                    continue
                spec = (spec or "").strip() if isinstance(spec, str) else (spec or "")
                spec = spec or None
                base_unit = (
                    (base_unit or "").strip()
                    if isinstance(base_unit, str)
                    else (base_unit or "")
                )
                base_unit = base_unit or None
                remark = (remark or "").strip() if isinstance(remark, str) else (remark or "")
                remark = remark or None
                if not code:
                    code = _next_material_code()
                    while SemiMaterial.query.filter_by(code=code).first():
                        code = _bump_material_code(code)
                existing = SemiMaterial.query.filter_by(code=code).first()
                if existing and existing.kind != "material":
                    errors.append(f"第 {idx} 行：编号已存在但类别不匹配")
                    continue
                if existing:
                    existing.kind = "material"
                    existing.name = name
                    existing.spec = spec
                    existing.base_unit = base_unit
                    existing.remark = remark
                    db.session.add(existing)
                else:
                    db.session.add(
                        SemiMaterial(
                            kind="material",
                            code=code,
                            name=name,
                            spec=spec,
                            base_unit=base_unit,
                            remark=remark,
                        )
                    )
                success += 1
                if success % 200 == 0:
                    db.session.flush()
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("导入失败：数据冲突。请重试或检查编码重复。", "danger")
                return render_template("procurement/material_import.html", result=None)
            if success:
                flash(f"成功导入/更新 {success} 条。", "success")
            if errors:
                flash(f"有 {len(errors)} 条记录导入失败，请查看错误列表。", "danger")
            return render_template(
                "procurement/material_import.html",
                result={"success": success, "errors": errors},
            )
        return render_template("procurement/material_import.html", result=None)

    @bp.route("/procurement/suppliers")
    @login_required
    @menu_required("procurement_supplier")
    def procurement_supplier_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        rows = (
            _supplier_search_query(company_id, keyword, active_only=False).all()
            if company_id
            else []
        )
        return render_template(
            "procurement/supplier_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
        )

    @bp.route("/procurement/suppliers/export-import-template", methods=["GET"])
    @login_required
    @menu_required("procurement_supplier")
    @capability_required("procurement_supplier.action.create")
    def procurement_supplier_export_import_template():
        from io import BytesIO

        from openpyxl import Workbook

        # 模板列顺序：第 1 行表头；第 2 行起为数据。
        # 物料匹配使用 SemiMaterial.code（kind=material）。
        headers = [
            "供应商名称",
            "联系人",
            "电话",
            "地址",
            "供应商状态（1启用/0停用）",
            "供应商备注",
            "物料编号",
            "最近单价",
            "物料-供应商备注",
            "是否默认供应商（1/0，可空=不修改）",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col, header)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="供应商导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/procurement/suppliers/import", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_supplier")
    @capability_required("procurement_supplier.action.create")
    def procurement_supplier_import():
        if request.method == "POST":
            company_id, _ = _resolve_company_id()
            if not company_id:
                flash("请先设置默认经营主体。", "danger")
                return render_template("procurement/supplier_import.html", result=None)

            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("procurement/supplier_import.html", result=None)

            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template(
                    "procurement/supplier_import.html", result=None
                )

            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("procurement/supplier_import.html", result=None)

            success = 0
            suppliers_data, errors = _parse_supplier_import_excel_ws(ws, company_id)

            # 一次性提交：允许文件里存在部分错误，但仍尽量导入可用数据。
            try:
                for _supplier_name, info in suppliers_data.items():
                    base = info.get("base") or {}
                    maps = info.get("maps") or []

                    supplier_name = (base.get("name") or "").strip() or _supplier_name
                    supplier = (
                        Supplier.query.filter_by(
                            company_id=company_id, name=supplier_name
                        ).first()
                    )
                    if not supplier:
                        supplier = Supplier(company_id=company_id, name=supplier_name)
                        db.session.add(supplier)
                        db.session.flush()

                    supplier.contact_name = base.get("contact_name")
                    supplier.phone = base.get("phone")
                    supplier.address = base.get("address")
                    supplier.is_active = bool(base.get("is_active", True))
                    supplier.remark = base.get("remark")

                    if maps:
                        _upsert_supplier_material_maps_no_delete(supplier, maps)
                        success += len(maps)

                db.session.commit()
            except ValueError as exc:
                db.session.rollback()
                errors.append(str(exc))
                flash("导入失败：请检查数据。", "danger")
                return render_template(
                    "procurement/supplier_import.html",
                    result={"success": success, "errors": errors},
                )
            except IntegrityError:
                db.session.rollback()
                flash("导入失败：数据冲突。请重试或检查唯一键。", "danger")
                return render_template("procurement/supplier_import.html", result=None)

            if success:
                flash(f"成功导入/更新 {success} 条。", "success")
            if errors:
                flash(f"有 {len(errors)} 条记录导入失败，请查看错误列表。", "danger")

            return render_template(
                "procurement/supplier_import.html",
                result={"success": success, "errors": errors},
            )

        return render_template("procurement/supplier_import.html", result=None)

    @bp.route("/procurement/suppliers/new", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_supplier")
    @capability_required("procurement_supplier.action.create")
    def procurement_supplier_new():
        company_id, companies = _resolve_company_id()
        if request.method == "POST":
            try:
                if not company_id:
                    raise ValueError("请先设置默认经营主体。")
                row = Supplier(
                    company_id=company_id,
                    name=((request.form.get("name") or "").strip()[:128]),
                    contact_name=(
                        (request.form.get("contact_name") or "").strip()[:64] or None
                    ),
                    phone=((request.form.get("phone") or "").strip()[:32] or None),
                    address=((request.form.get("address") or "").strip()[:255] or None),
                    is_active=request.form.get("is_active") == "1",
                    remark=((request.form.get("remark") or "").strip()[:500] or None),
                )
                if not row.name:
                    raise ValueError("供应商名称不能为空。")
                db.session.add(row)
                db.session.flush()
                _save_supplier_material_maps(row, _parse_supplier_map_rows(company_id))
                db.session.commit()
                flash("供应商已保存。", "success")
                return redirect(url_for("main.procurement_supplier_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("供应商名称或物料关联重复，请检查后重试。", "danger")
        return render_template(
            "procurement/supplier_form.html",
            row=None,
            company_id=company_id,
            companies=companies,
            map_rows=[],
            material_search_url=url_for("main.procurement_materials_search"),
        )

    @bp.route("/procurement/suppliers/<int:supplier_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_supplier")
    @capability_required("procurement_supplier.action.edit")
    def procurement_supplier_edit(supplier_id):
        row = Supplier.query.get_or_404(supplier_id)
        company_id, companies = _resolve_company_id()
        map_rows = (
            SupplierMaterialMap.query.options(
                selectinload(SupplierMaterialMap.material)
            )
            .filter_by(supplier_id=row.id)
            .order_by(SupplierMaterialMap.id.asc())
            .all()
        )
        if request.method == "POST":
            try:
                row.name = (request.form.get("name") or "").strip()[:128]
                row.contact_name = (request.form.get("contact_name") or "").strip()[
                    :64
                ] or None
                row.phone = (request.form.get("phone") or "").strip()[:32] or None
                row.address = (request.form.get("address") or "").strip()[:255] or None
                row.is_active = request.form.get("is_active") == "1"
                row.remark = (request.form.get("remark") or "").strip()[:500] or None
                if not row.name:
                    raise ValueError("供应商名称不能为空。")
                new_rows = _parse_supplier_map_rows(row.company_id)
                _save_supplier_material_maps(row, new_rows)
                db.session.commit()
                flash("供应商已更新。", "success")
                return redirect(url_for("main.procurement_supplier_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("供应商名称或物料关联重复，请检查后重试。", "danger")
        return render_template(
            "procurement/supplier_form.html",
            row=row,
            company_id=company_id,
            companies=companies,
            map_rows=map_rows,
            material_search_url=url_for("main.procurement_materials_search"),
        )

    @bp.route("/procurement/suppliers/<int:supplier_id>/delete", methods=["POST"])
    @login_required
    @menu_required("procurement_supplier")
    @capability_required("procurement_supplier.action.delete")
    def procurement_supplier_delete(supplier_id):
        row = Supplier.query.get_or_404(supplier_id)
        if PurchaseOrder.query.filter_by(supplier_id=row.id).first():
            flash("该供应商已被采购单使用，不能删除。", "danger")
            return redirect(url_for("main.procurement_supplier_list"))
        if PurchaseRequisitionLine.query.filter_by(supplier_id=row.id).first():
            flash("该供应商已被请购单使用，不能删除。", "danger")
            return redirect(url_for("main.procurement_supplier_list"))
        SupplierMaterialMap.query.filter_by(supplier_id=row.id).delete(
            synchronize_session=False
        )
        db.session.delete(row)
        db.session.commit()
        flash("供应商已删除。", "success")
        return redirect(url_for("main.procurement_supplier_list"))

    @bp.route("/api/procurement/suppliers-search")
    @login_required
    @menu_required(
        "procurement_supplier", "procurement_requisition", "procurement_order"
    )
    def procurement_suppliers_search():
        company_id, _ = _resolve_company_id()
        if not company_id:
            return jsonify({"items": []})
        keyword = (request.args.get("q") or "").strip()
        limit = max(1, min(request.args.get("limit", 20, type=int), 20))
        rows = (
            _supplier_search_query(company_id, keyword, active_only=True)
            .limit(limit)
            .all()
        )
        return jsonify({"items": [_supplier_payload(row) for row in rows]})

    @bp.route("/api/procurement/materials-search")
    @login_required
    @menu_required(
        "procurement_supplier", "procurement_requisition", "procurement_order"
    )
    def procurement_materials_search():
        company_id, _ = _resolve_company_id()
        keyword = (request.args.get("q") or "").strip()
        limit = max(1, min(request.args.get("limit", 20, type=int), 20))
        rows = _material_search_query(keyword).limit(limit).all()
        items = []
        for row in rows:
            best_mapping = _best_material_mapping(company_id, row.id)
            items.append(
                _material_payload(
                    row,
                    last_unit_price=(
                        _as_decimal(best_mapping.last_unit_price)
                        if best_mapping and best_mapping.last_unit_price is not None
                        else None
                    ),
                    is_default_supplier=bool(
                        best_mapping and best_mapping.is_preferred
                    ),
                    supplier_id=best_mapping.supplier_id
                    if best_mapping and best_mapping.is_preferred
                    else None,
                )
            )
        return jsonify({"items": items})

    @bp.route("/api/procurement/supplier-materials")
    @login_required
    @menu_required("procurement_requisition", "procurement_order")
    def procurement_supplier_materials_search():
        company_id, _ = _resolve_company_id()
        supplier_id = request.args.get("supplier_id", type=int)
        keyword = (request.args.get("q") or "").strip()
        limit = max(1, min(request.args.get("limit", 20, type=int), 20))
        if not company_id or not supplier_id:
            return jsonify({"items": []})
        return jsonify(
            {
                "items": _material_items_for_supplier(
                    company_id,
                    supplier_id,
                    keyword=keyword,
                    limit=limit,
                )
            }
        )

    @bp.route("/api/procurement/material-suppliers")
    @login_required
    @menu_required("procurement_requisition", "procurement_order")
    def procurement_material_suppliers_search():
        company_id, _ = _resolve_company_id()
        material_id = request.args.get("material_id", type=int)
        keyword = (request.args.get("q") or "").strip()
        limit = max(1, min(request.args.get("limit", 20, type=int), 20))
        if not company_id or not material_id:
            return jsonify({"items": []})
        return jsonify(
            {
                "items": _supplier_items_for_material(
                    company_id,
                    material_id,
                    keyword=keyword,
                    limit=limit,
                )
            }
        )

    @bp.route("/purchase-requisitions")
    @login_required
    @menu_required("procurement_requisition")
    def procurement_requisition_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        status = (request.args.get("status") or "").strip()
        q = PurchaseRequisition.query.options(
            selectinload(PurchaseRequisition.lines),
            selectinload(PurchaseRequisition.requester),
            selectinload(PurchaseRequisition.signer),
        )
        if company_id:
            q = q.filter(PurchaseRequisition.company_id == company_id)
        if current_user_can_cap("procurement_requisition.filter.keyword") and keyword:
            q = q.filter(
                or_(
                    PurchaseRequisition.req_no.contains(keyword),
                    PurchaseRequisition.supplier_name.contains(keyword),
                    PurchaseRequisition.item_name.contains(keyword),
                    PurchaseRequisition.lines.any(
                        PurchaseRequisitionLine.supplier_name.contains(keyword)
                    ),
                    PurchaseRequisition.lines.any(
                        PurchaseRequisitionLine.item_name.contains(keyword)
                    ),
                )
            )
        if status in REQUISITION_STATUS:
            q = q.filter(PurchaseRequisition.status == status)
        rows = q.order_by(PurchaseRequisition.id.desc()).all()
        return render_template(
            "procurement/requisition_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
            status=status,
            requisition_statuses=REQUISITION_STATUS,
        )

    @bp.route("/purchase-requisitions/new", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.create")
    def procurement_requisition_new():
        company_id, companies = _resolve_company_id()
        if request.method == "POST":
            try:
                if not company_id:
                    raise ValueError("请先设置默认经营主体。")
                row = PurchaseRequisition(
                    company_id=company_id,
                    req_no=_next_no("REQ", PurchaseRequisition, "req_no"),
                    requester_user_id=int(current_user.get_id()),
                    remark=((request.form.get("remark") or "").strip()[:500] or None),
                )
                db.session.add(row)
                db.session.flush()
                for item in _parse_requisition_lines(company_id):
                    row.lines.append(PurchaseRequisitionLine(**item))
                _sync_requisition_summary(row)
                _sync_requisition_status(row)
                db.session.commit()
                flash("请购单已保存。", "success")
                return redirect(url_for("main.procurement_requisition_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("请购单号冲突，请重试。", "danger")
        return render_template(
            "procurement/requisition_form.html",
            row=None,
            companies=companies,
            company_id=company_id,
            material_search_url=url_for("main.procurement_materials_search"),
            supplier_search_url=url_for("main.procurement_suppliers_search"),
            supplier_materials_url=url_for(
                "main.procurement_supplier_materials_search"
            ),
            material_suppliers_url=url_for(
                "main.procurement_material_suppliers_search"
            ),
        )

    @bp.route("/purchase-requisitions/<int:req_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.edit")
    def procurement_requisition_edit(req_id):
        row = PurchaseRequisition.query.options(
            selectinload(PurchaseRequisition.lines)
        ).get_or_404(req_id)
        company_id, companies = _resolve_company_id()
        if row.signed_at:
            flash("已签字请购单不能再修改明细。", "warning")
            return redirect(url_for("main.procurement_requisition_list"))
        if row.purchase_orders.filter(PurchaseOrder.status != "cancelled").first():
            flash("请购单已生成采购单，不能再修改。", "warning")
            return redirect(url_for("main.procurement_requisition_list"))
        if request.method == "POST":
            try:
                row.remark = (request.form.get("remark") or "").strip()[:500] or None
                row.lines[:] = []
                db.session.flush()
                for item in _parse_requisition_lines(row.company_id):
                    row.lines.append(PurchaseRequisitionLine(**item))
                _sync_requisition_summary(row)
                _sync_requisition_status(row)
                db.session.commit()
                flash("请购单已更新。", "success")
                return redirect(url_for("main.procurement_requisition_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
        return render_template(
            "procurement/requisition_form.html",
            row=row,
            companies=companies,
            company_id=company_id,
            material_search_url=url_for("main.procurement_materials_search"),
            supplier_search_url=url_for("main.procurement_suppliers_search"),
            supplier_materials_url=url_for(
                "main.procurement_supplier_materials_search"
            ),
            material_suppliers_url=url_for(
                "main.procurement_material_suppliers_search"
            ),
        )

    @bp.route("/purchase-requisitions/<int:req_id>/delete", methods=["POST"])
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.delete")
    def procurement_requisition_delete(req_id):
        row = PurchaseRequisition.query.get_or_404(req_id)
        if row.purchase_orders.filter(PurchaseOrder.status != "cancelled").first():
            flash("该请购单已生成采购单，不能删除。", "danger")
            return redirect(url_for("main.procurement_requisition_list"))
        db.session.delete(row)
        db.session.commit()
        flash("请购单已删除。", "success")
        return redirect(url_for("main.procurement_requisition_list"))

    @bp.route("/purchase-requisitions/<int:req_id>/print")
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.print")
    def procurement_requisition_print(req_id):
        row = PurchaseRequisition.query.options(
            selectinload(PurchaseRequisition.lines),
            selectinload(PurchaseRequisition.requester),
            selectinload(PurchaseRequisition.signer),
            selectinload(PurchaseRequisition.company),
        ).get_or_404(req_id)
        row.printed_at = datetime.now()
        db.session.commit()
        return render_template("procurement/requisition_print.html", row=row)

    @bp.route("/purchase-requisitions/<int:req_id>/mark-signed", methods=["POST"])
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.mark_signed")
    def procurement_requisition_mark_signed(req_id):
        row = PurchaseRequisition.query.options(
            selectinload(PurchaseRequisition.lines)
        ).get_or_404(req_id)
        if not row.lines:
            flash("请购单没有明细，不能签字。", "danger")
            return redirect(url_for("main.procurement_requisition_list"))
        row.signed_at = datetime.now()
        row.signed_by = int(current_user.get_id())
        _sync_requisition_status(row)
        db.session.commit()
        flash("请购单已标记签字。", "success")
        return redirect(url_for("main.procurement_requisition_list"))

    @bp.route("/purchase-requisitions/<int:req_id>/generate-orders", methods=["POST"])
    @login_required
    @menu_required("procurement_requisition")
    @capability_required("procurement_requisition.action.generate_orders")
    def procurement_requisition_generate_orders(req_id):
        row = PurchaseRequisition.query.options(
            selectinload(PurchaseRequisition.lines),
            selectinload(PurchaseRequisition.company),
        ).get_or_404(req_id)
        try:
            created_orders = create_purchase_orders_from_requisition(
                row,
                buyer_user_id=int(current_user.get_id()),
            )
            db.session.commit()
            if created_orders:
                flash(f"已生成 {len(created_orders)} 张采购单。", "success")
            else:
                flash("没有可生成的采购单明细。", "warning")
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        except IntegrityError:
            db.session.rollback()
            flash("采购单编号冲突，请重试。", "danger")
        return redirect(url_for("main.procurement_requisition_list"))

    @bp.route("/purchase-orders")
    @login_required
    @menu_required("procurement_order")
    def procurement_order_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        status = (request.args.get("status") or "").strip()
        q = PurchaseOrder.query.options(
            selectinload(PurchaseOrder.requisition),
            selectinload(PurchaseOrder.requisition_line),
            selectinload(PurchaseOrder.supplier),
            selectinload(PurchaseOrder.material),
        )
        if company_id:
            q = q.filter(PurchaseOrder.company_id == company_id)
        if current_user_can_cap("procurement_order.filter.keyword") and keyword:
            q = q.filter(
                or_(
                    PurchaseOrder.po_no.contains(keyword),
                    PurchaseOrder.supplier_name.contains(keyword),
                    PurchaseOrder.item_name.contains(keyword),
                )
            )
        if status in PO_STATUS:
            q = q.filter(PurchaseOrder.status == status)
        rows = q.order_by(PurchaseOrder.id.desc()).all()
        return render_template(
            "procurement/order_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
            status=status,
            po_statuses=PO_STATUS,
        )

    @bp.route("/purchase-orders/new", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.create")
    def procurement_order_new():
        company_id, companies = _resolve_company_id()
        requisition_lines = (
            PurchaseRequisitionLine.query.options(
                selectinload(PurchaseRequisitionLine.requisition)
            )
            .filter(
                PurchaseRequisitionLine.company_id == company_id,
                PurchaseRequisitionLine.status != "ordered",
            )
            .order_by(PurchaseRequisitionLine.id.desc())
            .limit(200)
            .all()
            if company_id
            else []
        )
        if request.method == "POST":
            try:
                if not company_id:
                    raise ValueError("请先设置默认经营主体。")
                supplier_id = request.form.get("supplier_id", type=int)
                material_id = request.form.get("material_id", type=int)
                requisition_line_id = request.form.get("requisition_line_id", type=int)
                requisition_line = (
                    PurchaseRequisitionLine.query.get(requisition_line_id)
                    if requisition_line_id
                    else None
                )
                if requisition_line:
                    supplier_id = requisition_line.supplier_id
                    material_id = requisition_line.material_id
                if not supplier_id or not material_id:
                    raise ValueError("请选择有效供应商和物料。")
                supplier, material, _mapping = _require_supplier_material_mapping(
                    company_id,
                    supplier_id,
                    material_id,
                    row_label="采购单",
                )
                qty = _parse_decimal(
                    request.form.get("qty")
                    or (str(requisition_line.qty) if requisition_line else ""),
                    "采购数量",
                    allow_zero=False,
                )
                unit_price = _parse_decimal(request.form.get("unit_price"), "单价")
                row = PurchaseOrder(
                    company_id=company_id,
                    po_no=_next_purchase_order_no(Company.query.get(company_id)),
                    requisition_id=requisition_line.requisition_id
                    if requisition_line
                    else None,
                    requisition_line_id=requisition_line.id
                    if requisition_line
                    else None,
                    buyer_user_id=int(current_user.get_id()),
                    supplier_id=supplier.id,
                    material_id=material.id,
                    supplier_name=supplier.name,
                    supplier_contact_name=supplier.contact_name,
                    supplier_phone=supplier.phone,
                    supplier_address=supplier.address,
                    item_name=material.name,
                    item_spec=material.spec,
                    qty=qty,
                    unit=(
                        (request.form.get("unit") or "").strip()[:16]
                        or material.base_unit
                        or "pcs"
                    ),
                    unit_price=unit_price,
                    amount=qty * unit_price,
                    expected_date=_parse_optional_date(
                        request.form.get("expected_date"), "到货日期"
                    ),
                    status=(request.form.get("status") or "draft").strip()
                    if (request.form.get("status") or "").strip() in PO_STATUS
                    else "draft",
                    remark=((request.form.get("remark") or "").strip()[:500] or None),
                    reconcile_status="pending",
                )
                if row.status == "ordered":
                    row.ordered_at = datetime.now()
                    row.ordered_by = int(current_user.get_id())
                db.session.add(row)
                _touch_supplier_material_map(
                    company_id=company_id,
                    supplier_id=supplier.id,
                    material_id=material.id,
                    unit_price=unit_price,
                )
                db.session.flush()
                _update_requisition_from_order(row)
                db.session.commit()
                flash("采购单已保存。", "success")
                return redirect(url_for("main.procurement_order_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("采购单编号冲突，请重试。", "danger")
        return render_template(
            "procurement/order_form.html",
            row=None,
            companies=companies,
            company_id=company_id,
            requisition_lines=requisition_lines,
            po_statuses=PO_STATUS,
            material_search_url=url_for("main.procurement_materials_search"),
            supplier_search_url=url_for("main.procurement_suppliers_search"),
            supplier_materials_url=url_for(
                "main.procurement_supplier_materials_search"
            ),
            material_suppliers_url=url_for(
                "main.procurement_material_suppliers_search"
            ),
        )

    @bp.route("/purchase-orders/<int:po_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.edit")
    def procurement_order_edit(po_id):
        row = PurchaseOrder.query.get_or_404(po_id)
        company_id, companies = _resolve_company_id()
        requisition_lines = (
            PurchaseRequisitionLine.query.options(
                selectinload(PurchaseRequisitionLine.requisition)
            )
            .filter(PurchaseRequisitionLine.company_id == row.company_id)
            .order_by(PurchaseRequisitionLine.id.desc())
            .limit(200)
            .all()
        )
        if row.receipts.count():
            flash("采购单已有收货记录，不能再编辑。", "warning")
            return redirect(url_for("main.procurement_order_detail", po_id=row.id))
        if request.method == "POST":
            try:
                supplier_id = request.form.get("supplier_id", type=int)
                material_id = request.form.get("material_id", type=int)
                requisition_line_id = request.form.get("requisition_line_id", type=int)
                requisition_line = (
                    PurchaseRequisitionLine.query.get(requisition_line_id)
                    if requisition_line_id
                    else None
                )
                if requisition_line:
                    supplier_id = requisition_line.supplier_id
                    material_id = requisition_line.material_id
                if not supplier_id or not material_id:
                    raise ValueError("请选择有效供应商和物料。")
                supplier, material, _mapping = _require_supplier_material_mapping(
                    row.company_id,
                    supplier_id,
                    material_id,
                    row_label="采购单",
                )
                qty = _parse_decimal(
                    request.form.get("qty")
                    or (
                        str(requisition_line.qty) if requisition_line else str(row.qty)
                    ),
                    "采购数量",
                    allow_zero=False,
                )
                unit_price = _parse_decimal(request.form.get("unit_price"), "单价")
                row.requisition_id = (
                    requisition_line.requisition_id if requisition_line else None
                )
                row.requisition_line_id = (
                    requisition_line.id if requisition_line else None
                )
                row.supplier_id = supplier.id
                row.material_id = material.id
                row.supplier_name = supplier.name
                row.supplier_contact_name = supplier.contact_name
                row.supplier_phone = supplier.phone
                row.supplier_address = supplier.address
                row.item_name = material.name
                row.item_spec = material.spec
                row.qty = qty
                row.unit = (
                    (request.form.get("unit") or "").strip()[:16]
                    or material.base_unit
                    or "pcs"
                )
                row.unit_price = unit_price
                row.amount = qty * unit_price
                row.expected_date = _parse_optional_date(
                    request.form.get("expected_date"), "到货日期"
                )
                row.remark = (request.form.get("remark") or "").strip()[:500] or None
                if row.status != "received":
                    row.status = (
                        (request.form.get("status") or row.status).strip()
                        if (request.form.get("status") or row.status).strip()
                        in PO_STATUS
                        else row.status
                    )
                if row.status == "ordered" and not row.ordered_at:
                    row.ordered_at = datetime.now()
                    row.ordered_by = int(current_user.get_id())
                _touch_supplier_material_map(
                    company_id=row.company_id,
                    supplier_id=supplier.id,
                    material_id=material.id,
                    unit_price=unit_price,
                )
                _update_requisition_from_order(row)
                db.session.commit()
                flash("采购单已更新。", "success")
                return redirect(url_for("main.procurement_order_detail", po_id=row.id))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
        return render_template(
            "procurement/order_form.html",
            row=row,
            companies=companies,
            company_id=company_id,
            requisition_lines=requisition_lines,
            po_statuses=PO_STATUS,
            material_search_url=url_for("main.procurement_materials_search"),
            supplier_search_url=url_for("main.procurement_suppliers_search"),
            supplier_materials_url=url_for(
                "main.procurement_supplier_materials_search"
            ),
            material_suppliers_url=url_for(
                "main.procurement_material_suppliers_search"
            ),
        )

    @bp.route("/purchase-orders/<int:po_id>/delete", methods=["POST"])
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.delete")
    def procurement_order_delete(po_id):
        row = PurchaseOrder.query.get_or_404(po_id)
        if row.receipts.count():
            flash("采购单已有收货记录，不能删除。", "danger")
            return redirect(url_for("main.procurement_order_list"))
        requisition = row.requisition
        requisition_line = row.requisition_line
        db.session.delete(row)
        db.session.flush()
        if requisition_line:
            _sync_requisition_line_status(requisition_line)
        if requisition:
            _sync_requisition_summary(requisition)
            _sync_requisition_status(requisition)
        db.session.commit()
        flash("采购单已删除。", "success")
        return redirect(url_for("main.procurement_order_list"))

    @bp.route("/purchase-orders/<int:po_id>")
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.detail")
    def procurement_order_detail(po_id):
        row = PurchaseOrder.query.options(
            selectinload(PurchaseOrder.requisition),
            selectinload(PurchaseOrder.requisition_line),
            selectinload(PurchaseOrder.company),
        ).get_or_404(po_id)
        receipts = (
            PurchaseReceipt.query.filter_by(purchase_order_id=row.id)
            .order_by(PurchaseReceipt.id.desc())
            .all()
        )
        compare_summary = {
            "ordered_qty": _as_decimal(row.qty),
            "received_qty": _sum_posted_receipt_qty(row),
            "warehouse_qty": _sum_warehouse_qty(row),
            "reconcile_status": row.reconcile_status,
        }
        return render_template(
            "procurement/order_detail.html",
            row=row,
            receipts=receipts,
            compare_summary=compare_summary,
        )

    @bp.route("/purchase-orders/<int:po_id>/print")
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.print")
    def procurement_order_print(po_id):
        row = PurchaseOrder.query.options(
            selectinload(PurchaseOrder.company),
            selectinload(PurchaseOrder.supplier),
            selectinload(PurchaseOrder.material),
        ).get_or_404(po_id)
        row.printed_at = datetime.now()
        workbook = build_purchase_order_workbook(row)
        db.session.commit()
        return send_file(
            workbook,
            as_attachment=True,
            download_name=f"{row.po_no}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/purchase-orders/<int:po_id>/mark-ordered", methods=["POST"])
    @login_required
    @menu_required("procurement_order")
    @capability_required("procurement_order.action.mark_ordered")
    def procurement_order_mark_ordered(po_id):
        row = PurchaseOrder.query.get_or_404(po_id)
        if row.status != "cancelled":
            row.status = "ordered" if row.status == "draft" else row.status
            row.ordered_at = row.ordered_at or datetime.now()
            row.ordered_by = row.ordered_by or int(current_user.get_id())
            db.session.commit()
            flash("采购单已标记为已下单。", "success")
        return redirect(url_for("main.procurement_order_detail", po_id=row.id))

    @bp.route("/purchase-receipts")
    @login_required
    @menu_required("procurement_receipt")
    def procurement_receipt_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        status = (request.args.get("status") or "").strip()
        q = PurchaseReceipt.query.options(
            selectinload(PurchaseReceipt.purchase_order),
            selectinload(PurchaseReceipt.receiver),
            selectinload(PurchaseReceipt.reconciler),
        )
        if company_id:
            q = q.filter(PurchaseReceipt.company_id == company_id)
        if current_user_can_cap("procurement_receipt.filter.keyword") and keyword:
            q = q.join(
                PurchaseOrder, PurchaseOrder.id == PurchaseReceipt.purchase_order_id
            ).filter(
                or_(
                    PurchaseReceipt.receipt_no.contains(keyword),
                    PurchaseOrder.po_no.contains(keyword),
                    PurchaseOrder.supplier_name.contains(keyword),
                )
            )
        if status in RECEIPT_STATUS:
            q = q.filter(PurchaseReceipt.status == status)
        rows = q.order_by(PurchaseReceipt.id.desc()).all()
        return render_template(
            "procurement/receipt_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
            status=status,
            receipt_statuses=RECEIPT_STATUS,
        )

    @bp.route("/purchase-receipts/new", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_receipt")
    @capability_required("procurement_receipt.action.create")
    def procurement_receipt_new():
        company_id, companies = _resolve_company_id()
        orders = (
            PurchaseOrder.query.filter(
                PurchaseOrder.company_id == company_id,
                PurchaseOrder.status.in_(
                    ("draft", "ordered", "partially_received", "received")
                ),
            )
            .order_by(PurchaseOrder.id.desc())
            .limit(200)
            .all()
            if company_id
            else []
        )
        if request.method == "POST":
            try:
                purchase_order_id = request.form.get("purchase_order_id", type=int)
                po = (
                    PurchaseOrder.query.get(purchase_order_id)
                    if purchase_order_id
                    else None
                )
                if not po or po.company_id != company_id:
                    raise ValueError("请选择有效采购单。")
                row = PurchaseReceipt(
                    company_id=company_id,
                    receipt_no=_next_no("RCV", PurchaseReceipt, "receipt_no"),
                    purchase_order_id=po.id,
                    receiver_user_id=int(current_user.get_id()),
                    received_qty=_parse_decimal(
                        request.form.get("received_qty"), "收货数量", allow_zero=False
                    ),
                    received_at=_parse_datetime_local(
                        request.form.get("received_at"), "收货时间"
                    ),
                    status=(request.form.get("status") or "draft").strip()
                    if (request.form.get("status") or "").strip() in RECEIPT_STATUS
                    else "draft",
                    reconcile_status="pending",
                    remark=((request.form.get("remark") or "").strip()[:500] or None),
                )
                db.session.add(row)
                db.session.flush()
                _sync_purchase_order_status(po)
                _sync_purchase_order_reconcile_status(po)
                db.session.commit()
                flash("收货单已保存。", "success")
                return redirect(url_for("main.procurement_receipt_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("收货单号冲突，请重试。", "danger")
        return render_template(
            "procurement/receipt_form.html",
            row=None,
            companies=companies,
            company_id=company_id,
            orders=orders,
            receipt_statuses=RECEIPT_STATUS,
        )

    @bp.route("/purchase-receipts/<int:receipt_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("procurement_receipt")
    @capability_required("procurement_receipt.action.edit")
    def procurement_receipt_edit(receipt_id):
        row = PurchaseReceipt.query.get_or_404(receipt_id)
        company_id, companies = _resolve_company_id()
        if PurchaseStockIn.query.filter_by(receipt_id=row.id).first():
            flash("该收货单已完成采购确认，不能再编辑。", "warning")
            return redirect(
                url_for("main.procurement_receipt_compare", receipt_id=row.id)
            )
        orders = (
            PurchaseOrder.query.filter(
                PurchaseOrder.company_id == row.company_id,
                PurchaseOrder.status.in_(
                    ("draft", "ordered", "partially_received", "received")
                ),
            )
            .order_by(PurchaseOrder.id.desc())
            .limit(200)
            .all()
        )
        if request.method == "POST":
            try:
                purchase_order_id = request.form.get("purchase_order_id", type=int)
                po = (
                    PurchaseOrder.query.get(purchase_order_id)
                    if purchase_order_id
                    else None
                )
                if not po or po.company_id != row.company_id:
                    raise ValueError("请选择有效采购单。")
                old_po = row.purchase_order
                row.purchase_order_id = po.id
                row.received_qty = _parse_decimal(
                    request.form.get("received_qty"), "收货数量", allow_zero=False
                )
                row.received_at = _parse_datetime_local(
                    request.form.get("received_at"), "收货时间"
                )
                row.status = (
                    (request.form.get("status") or row.status).strip()
                    if (request.form.get("status") or row.status).strip()
                    in RECEIPT_STATUS
                    else row.status
                )
                row.remark = (request.form.get("remark") or "").strip()[:500] or None
                row.reconcile_status = "pending"
                row.reconcile_note = None
                row.reconciled_at = None
                row.reconciled_by = None
                if old_po and old_po.id != po.id:
                    _sync_purchase_order_status(old_po)
                    _sync_purchase_order_reconcile_status(old_po)
                _sync_purchase_order_status(po)
                _sync_purchase_order_reconcile_status(po)
                db.session.commit()
                flash("收货单已更新。", "success")
                return redirect(url_for("main.procurement_receipt_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
        return render_template(
            "procurement/receipt_form.html",
            row=row,
            companies=companies,
            company_id=company_id,
            orders=orders,
            receipt_statuses=RECEIPT_STATUS,
        )

    @bp.route("/purchase-receipts/<int:receipt_id>/delete", methods=["POST"])
    @login_required
    @menu_required("procurement_receipt")
    @capability_required("procurement_receipt.action.delete")
    def procurement_receipt_delete(receipt_id):
        row = PurchaseReceipt.query.get_or_404(receipt_id)
        if PurchaseStockIn.query.filter_by(receipt_id=row.id).first():
            flash("该收货单已完成采购确认，不能删除。", "danger")
            return redirect(url_for("main.procurement_receipt_list"))
        po = row.purchase_order
        db.session.delete(row)
        db.session.flush()
        if po:
            _sync_purchase_order_status(po)
            _sync_purchase_order_reconcile_status(po)
        db.session.commit()
        flash("收货单已删除。", "success")
        return redirect(url_for("main.procurement_receipt_list"))

    @bp.route("/purchase-receipts/<int:receipt_id>/compare")
    @login_required
    @menu_required("procurement_receipt")
    @capability_required("procurement_receipt.action.compare")
    def procurement_receipt_compare(receipt_id):
        row = PurchaseReceipt.query.options(
            selectinload(PurchaseReceipt.purchase_order).selectinload(
                PurchaseOrder.material
            ),
            selectinload(PurchaseReceipt.purchase_order).selectinload(
                PurchaseOrder.company
            ),
            selectinload(PurchaseReceipt.receiver),
            selectinload(PurchaseReceipt.reconciler),
        ).get_or_404(receipt_id)
        if not row.purchase_order:
            abort(404)
        summary = _build_compare_summary(row)
        stockin = PurchaseStockIn.query.filter_by(receipt_id=row.id).first()
        inventory_entry_url = url_for(
            "main.inventory_movement_new",
            category=summary["category"],
            purchase_order_id=row.purchase_order_id,
            purchase_receipt_id=row.id,
        )
        return render_template(
            "procurement/stockin_compare.html",
            row=row,
            summary=summary,
            stockin=stockin,
            inventory_entry_url=inventory_entry_url,
        )

    @bp.route("/purchase-receipts/<int:receipt_id>/approve-stockin", methods=["POST"])
    @login_required
    @menu_required("procurement_receipt")
    @capability_required("procurement_receipt.action.approve_stockin")
    def procurement_receipt_approve_stockin(receipt_id):
        row = PurchaseReceipt.query.options(
            selectinload(PurchaseReceipt.purchase_order).selectinload(
                PurchaseOrder.material
            ),
        ).get_or_404(receipt_id)
        if row.status != "posted":
            flash("只有已过账收货单才能做采购确认。", "danger")
            return redirect(
                url_for("main.procurement_receipt_compare", receipt_id=row.id)
            )
        summary = _build_compare_summary(row)
        note = (request.form.get("reconcile_note") or "").strip()[:500] or None
        approval_status = "matched" if summary["is_matched"] else "exception"
        if approval_status == "exception" and not note:
            flash("数量不一致时必须填写差异说明。", "danger")
            return redirect(
                url_for("main.procurement_receipt_compare", receipt_id=row.id)
            )
        stockin = PurchaseStockIn.query.filter_by(receipt_id=row.id).first()
        now = datetime.now()
        if not stockin:
            stockin = PurchaseStockIn(
                company_id=row.company_id,
                stock_in_no=_next_no("SIN", PurchaseStockIn, "stock_in_no"),
                receipt_id=row.id,
                created_by=int(current_user.get_id()),
            )
            db.session.add(stockin)
        stockin.purchase_order_id = row.purchase_order_id
        stockin.qty = summary["ordered_qty"]
        stockin.received_qty = summary["received_qty"]
        stockin.warehouse_qty = summary["warehouse_qty"]
        stockin.variance_qty = summary["variance_qty"]
        stockin.approval_status = approval_status
        stockin.storage_area = _latest_storage_area(row.purchase_order, row)
        stockin.stock_in_at = now
        stockin.approved_by = int(current_user.get_id())
        stockin.approved_at = now
        stockin.remark = note or (
            "系统自动对比一致，直接确认。" if approval_status == "matched" else None
        )
        row.reconcile_status = approval_status
        row.reconcile_note = note
        row.reconciled_at = now
        row.reconciled_by = int(current_user.get_id())
        _sync_purchase_order_reconcile_status(row.purchase_order)
        db.session.commit()
        flash("采购确认已完成。", "success")
        return redirect(url_for("main.procurement_receipt_compare", receipt_id=row.id))

    @bp.route("/purchase-stock-ins")
    @login_required
    @menu_required("procurement_stockin")
    def procurement_stockin_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        q = PurchaseStockIn.query.options(
            selectinload(PurchaseStockIn.receipt),
            selectinload(PurchaseStockIn.purchase_order),
            selectinload(PurchaseStockIn.creator),
            selectinload(PurchaseStockIn.approver),
        )
        if company_id:
            q = q.filter(PurchaseStockIn.company_id == company_id)
        if current_user_can_cap("procurement_stockin.filter.keyword") and keyword:
            q = (
                q.join(
                    PurchaseReceipt, PurchaseReceipt.id == PurchaseStockIn.receipt_id
                )
                .join(
                    PurchaseOrder, PurchaseOrder.id == PurchaseStockIn.purchase_order_id
                )
                .filter(
                    or_(
                        PurchaseStockIn.stock_in_no.contains(keyword),
                        PurchaseReceipt.receipt_no.contains(keyword),
                        PurchaseOrder.po_no.contains(keyword),
                    )
                )
            )
        rows = q.order_by(PurchaseStockIn.id.desc()).all()
        return render_template(
            "procurement/stockin_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
        )
