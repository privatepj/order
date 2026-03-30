from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from flask import flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from app import db
from app.auth.capabilities import current_user_can_cap
from app.auth.decorators import capability_required, menu_required
from app.models import BomHeader, BomLine, Product, SemiMaterial
from app.services import bom_svc
from app.utils.query import keyword_like_or


def _parse_decimal(val: Any, *, default: Optional[Decimal] = None) -> Optional[Decimal]:
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    try:
        return Decimal(s)
    except InvalidOperation:
        return default


def _parse_lines_from_form(form) -> List[Dict[str, Any]]:
    child_kinds = form.getlist("line_child_kind")
    child_ids = form.getlist("line_child_material_id")
    qtys = form.getlist("line_quantity")
    units = form.getlist("line_unit")
    remarks = form.getlist("line_remark")

    out: List[Dict[str, Any]] = []
    line_no = 1
    for i, cid in enumerate(child_ids):
        cid_s = (cid or "").strip()
        if not cid_s:
            continue
        try:
            child_material_id = int(cid_s)
        except ValueError:
            continue

        child_kind = (child_kinds[i] if i < len(child_kinds) else "").strip() or ""
        if child_kind not in (bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
            raise ValueError("请选择 BOM 子项类别（semi/material）。")
        qty = _parse_decimal(qtys[i] if i < len(qtys) else None)
        if qty is None:
            raise ValueError("用量数量格式不正确。")
        if qty <= 0:
            raise ValueError("用量数量必须大于 0。")

        unit = (units[i] if i < len(units) else None) or None
        unit = unit.strip()[:16] if isinstance(unit, str) and unit.strip() else None

        rmk = (remarks[i] if i < len(remarks) else None) or None
        rmk = rmk.strip()[:255] if isinstance(rmk, str) and rmk.strip() else None

        out.append(
            {
                "line_no": line_no,
                "child_kind": child_kind,
                "child_material_id": child_material_id,
                "quantity": qty,
                "unit": unit,
                "remark": rmk,
            }
        )
        line_no += 1

    return out


def _parent_from_form() -> Tuple[str, int]:
    parent_kind = (request.form.get("parent_kind") or "").strip()
    if parent_kind not in (bom_svc.PARENT_FINISHED, bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
        parent_kind = bom_svc.PARENT_FINISHED
    parent_id_raw = (request.form.get("parent_id") or "").strip()
    try:
        parent_id = int(parent_id_raw)
    except ValueError:
        parent_id = 0
    if not parent_id:
        raise ValueError("请选择父项。")
    return parent_kind, parent_id


def _bom_parent_ids(parent_kind: str, parent_id: int) -> Tuple[int, int]:
    if parent_kind == bom_svc.PARENT_FINISHED:
        return parent_id, 0
    return 0, parent_id


def register_bom_routes(bp):
    # ----- 列表 -----
    @bp.route("/boms")
    @login_required
    @menu_required("bom")
    def bom_list():
        parent_kind = (request.args.get("parent_kind") or "finished").strip()
        if parent_kind not in (bom_svc.PARENT_FINISHED, bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
            parent_kind = bom_svc.PARENT_FINISHED

        keyword = (request.args.get("keyword") or "").strip()
        if not current_user_can_cap("bom.filter.keyword"):
            keyword = ""

        page = request.args.get("page", 1, type=int)
        per_page = 20

        q = BomHeader.query.filter(
            BomHeader.is_active.is_(True),
            BomHeader.parent_kind == parent_kind,
        )

        if parent_kind == bom_svc.PARENT_FINISHED:
            q = q.join(Product, BomHeader.parent_product_id == Product.id)
            q = q.options(selectinload(BomHeader.parent_product))
            if keyword:
                q = q.filter(
                    keyword_like_or(
                        keyword,
                        Product.product_code,
                        Product.name,
                        Product.spec,
                    )
                )
        else:
            q = q.join(SemiMaterial, BomHeader.parent_material_id == SemiMaterial.id)
            q = q.options(selectinload(BomHeader.parent_material))
            if keyword:
                q = q.filter(
                    keyword_like_or(
                        keyword,
                        SemiMaterial.code,
                        SemiMaterial.name,
                        SemiMaterial.spec,
                    )
                )

        q = q.order_by(BomHeader.id.desc())
        pagination = q.paginate(page=page, per_page=per_page)

        return render_template(
            "bom/list.html",
            pagination=pagination,
            parent_kind=parent_kind,
            keyword=keyword,
        )

    # ----- 新建 -----
    @bp.route("/boms/new", methods=["GET", "POST"])
    @login_required
    @menu_required("bom")
    @capability_required("bom.action.create")
    def bom_new():
        if request.method == "POST":
            try:
                parent_kind, parent_id = _parent_from_form()
                lines = _parse_lines_from_form(request.form)
                if not lines:
                    raise ValueError("请至少录入一行 BOM 子项。")

                # 补齐子项 kind（来自表单）并做校验
                bom_svc.validate_bom_lines(parent_kind=parent_kind, parent_id=parent_id, lines=lines)

                # 反激活旧版本（仅允许一个 active）
                parent_product_id, parent_material_id = _bom_parent_ids(parent_kind, parent_id)
                BomHeader.query.filter(
                    BomHeader.is_active.is_(True),
                    BomHeader.parent_kind == parent_kind,
                    BomHeader.parent_product_id == parent_product_id,
                    BomHeader.parent_material_id == parent_material_id,
                ).update({"is_active": False}, synchronize_session=False)

                max_ver = (
                    BomHeader.query.filter(
                        BomHeader.parent_kind == parent_kind,
                        BomHeader.parent_product_id == parent_product_id,
                        BomHeader.parent_material_id == parent_material_id,
                    )
                    .with_entities(func.max(BomHeader.version_no))
                    .scalar()
                )
                next_ver = int(max_ver or 0) + 1

                header = BomHeader(
                    parent_kind=parent_kind,
                    parent_product_id=parent_product_id,
                    parent_material_id=parent_material_id,
                    version_no=next_ver,
                    is_active=True,
                    remark=(request.form.get("remark") or "").strip()[:255] or None,
                )
                db.session.add(header)
                db.session.flush()

                for ln in lines:
                    db.session.add(
                        BomLine(
                            bom_header_id=header.id,
                            line_no=ln["line_no"],
                            child_kind=ln["child_kind"],
                            child_material_id=ln["child_material_id"],
                            quantity=ln["quantity"],
                            unit=ln.get("unit"),
                            remark=ln.get("remark"),
                        )
                    )
                db.session.commit()
                flash("BOM 已保存并生效。", "success")
                return redirect(url_for("main.bom_list", parent_kind=parent_kind))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：数据冲突。请重试。", "danger")

        parent_kind = request.args.get("parent_kind") or "finished"
        return render_template(
            "bom/form.html",
            header=None,
            parent_kind=parent_kind,
            parent_id=None,
            parent_label=None,
            lines=[],
        )

    # ----- 编辑（创建新版本并设置生效） -----
    @bp.route("/boms/<int:header_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("bom")
    @capability_required("bom.action.edit")
    def bom_edit(header_id: int):
        header = BomHeader.query.get_or_404(header_id)

        if request.method == "POST":
            try:
                parent_kind = header.parent_kind
                parent_id = header.parent_product_id if parent_kind == bom_svc.PARENT_FINISHED else header.parent_material_id

                lines = _parse_lines_from_form(request.form)
                if not lines:
                    raise ValueError("请至少录入一行 BOM 子项。")

                bom_svc.validate_bom_lines(parent_kind=parent_kind, parent_id=parent_id, lines=lines)

                # 反激活旧版本（仅允许一个 active）
                parent_product_id, parent_material_id = _bom_parent_ids(parent_kind, parent_id)
                BomHeader.query.filter(
                    BomHeader.is_active.is_(True),
                    BomHeader.parent_kind == parent_kind,
                    BomHeader.parent_product_id == parent_product_id,
                    BomHeader.parent_material_id == parent_material_id,
                ).update({"is_active": False}, synchronize_session=False)

                max_ver = (
                    BomHeader.query.filter(
                        BomHeader.parent_kind == parent_kind,
                        BomHeader.parent_product_id == parent_product_id,
                        BomHeader.parent_material_id == parent_material_id,
                    )
                    .with_entities(func.max(BomHeader.version_no))
                    .scalar()
                )
                next_ver = int(max_ver or 0) + 1

                new_header = BomHeader(
                    parent_kind=parent_kind,
                    parent_product_id=parent_product_id,
                    parent_material_id=parent_material_id,
                    version_no=next_ver,
                    is_active=True,
                    remark=(request.form.get("remark") or "").strip()[:255] or None,
                )
                db.session.add(new_header)
                db.session.flush()

                for ln in lines:
                    db.session.add(
                        BomLine(
                            bom_header_id=new_header.id,
                            line_no=ln["line_no"],
                            child_kind=ln["child_kind"],
                            child_material_id=ln["child_material_id"],
                            quantity=ln["quantity"],
                            unit=ln.get("unit"),
                            remark=ln.get("remark"),
                        )
                    )

                db.session.commit()
                flash("BOM 已更新并生效（新版本）。", "success")
                return redirect(url_for("main.bom_list", parent_kind=parent_kind))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：数据冲突。请重试。", "danger")

        # GET：回显当前版本明细
        lines_ctx: List[Dict[str, Any]] = []
        child_ids = [ln.child_material_id for ln in list(header.lines or [])]
        sm_map = {s.id: s for s in SemiMaterial.query.filter(SemiMaterial.id.in_(child_ids)).all()} if child_ids else {}

        for ln in sorted(list(header.lines or []), key=lambda x: x.line_no):
            sm = sm_map.get(ln.child_material_id)
            lines_ctx.append(
                {
                    "line_no": ln.line_no,
                    "child_kind": ln.child_kind,
                    "child_material_id": ln.child_material_id,
                    "quantity": ln.quantity,
                    "unit": ln.unit,
                    "remark": ln.remark,
                    "child_code": sm.code if sm else None,
                    "child_name": sm.name if sm else None,
                    "child_spec": sm.spec if sm else None,
                    "child_base_unit": sm.base_unit if sm else None,
                }
            )
        parent_kind = header.parent_kind
        parent_id = header.parent_product_id if parent_kind == bom_svc.PARENT_FINISHED else header.parent_material_id
        parent_label = None
        if parent_kind == bom_svc.PARENT_FINISHED:
            p = header.parent_product
            parent_label = f"{p.product_code} — {p.name}" if p else ""
            if p and p.spec:
                parent_label += f"（{p.spec}）"
        else:
            sm = header.parent_material
            parent_label = f"{sm.code} — {sm.name}" if sm else ""
            if sm and sm.spec:
                parent_label += f"（{sm.spec}）"

        return render_template(
            "bom/form.html",
            header=header,
            parent_kind=parent_kind,
            parent_id=parent_id,
            parent_label=parent_label,
            lines=lines_ctx,
        )

    # ----- 删除（删除某个版本；若删除 active 则激活最新版本） -----
    @bp.route("/boms/<int:header_id>/delete", methods=["POST"])
    @login_required
    @menu_required("bom")
    @capability_required("bom.action.delete")
    def bom_delete(header_id: int):
        header = BomHeader.query.get_or_404(header_id)
        parent_kind = header.parent_kind
        parent_product_id = header.parent_product_id
        parent_material_id = header.parent_material_id
        was_active = bool(header.is_active)

        db.session.delete(header)
        db.session.flush()

        if was_active:
            next_header = (
                BomHeader.query.filter(
                    BomHeader.parent_kind == parent_kind,
                    BomHeader.parent_product_id == parent_product_id,
                    BomHeader.parent_material_id == parent_material_id,
                )
                .order_by(BomHeader.version_no.desc())
                .first()
            )
            if next_header:
                next_header.is_active = True

        db.session.commit()
        flash("BOM 版本已删除。", "success")
        return redirect(url_for("main.bom_list", parent_kind=parent_kind))

    # ----- Excel：导入模板下载 -----
    @bp.route("/boms/export-import-template", methods=["GET"])
    @login_required
    @menu_required("bom")
    @capability_required("bom.action.import")
    def bom_export_import_template():
        from openpyxl import Workbook

        headers = [
            "父项类别（finished/semi/material）",
            "父项编号（product_code 或 semi_material.code）",
            "子项类别（semi/material）",
            "子项编号（semi_material.code）",
            "用量数量（>0）",
            "单位（可空）",
            "子项备注（可空）",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM 导入模板"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="BOM导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ----- Excel：导入 -----
    @bp.route("/boms/import", methods=["GET", "POST"])
    @login_required
    @menu_required("bom")
    @capability_required("bom.action.import")
    def bom_import():
        result: Optional[Dict[str, Any]] = None

        if request.method == "POST":
            file = request.files.get("file")
            if not file or not (file.filename or "").strip():
                flash("请先选择要上传的 Excel 文件（.xlsx）。", "danger")
                return render_template("bom/import.html", result=None)

            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("bom/import.html", result=None)

            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("bom/import.html", result=None)

            errors: List[str] = []
            groups: Dict[Tuple[str, int], List[Dict[str, Any]]] = {}

            product_cache: Dict[str, Optional[int]] = {}
            semi_cache: Dict[Tuple[str, str], Optional[int]] = {}

            def resolve_product_code(code_s: str) -> Optional[int]:
                if code_s in product_cache:
                    return product_cache[code_s]
                p = Product.query.filter_by(product_code=code_s).first()
                product_cache[code_s] = p.id if p else None
                return product_cache[code_s]

            def resolve_semi_kind_code(kind_s: str, code_s: str) -> Optional[int]:
                key = (kind_s, code_s)
                if key in semi_cache:
                    return semi_cache[key]
                sm = SemiMaterial.query.filter_by(kind=kind_s, code=code_s).first()
                semi_cache[key] = sm.id if sm else None
                return semi_cache[key]

            # Excel 从第 2 行开始
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                (
                    parent_kind_raw,
                    parent_code_raw,
                    child_kind_raw,
                    child_code_raw,
                    qty_raw,
                    unit_raw,
                    remark_raw,
                ) = (tuple(row) + (None,) * 7)[:7]

                parent_kind = (str(parent_kind_raw).strip() if parent_kind_raw is not None else "").strip()
                parent_code = (str(parent_code_raw).strip() if parent_code_raw is not None else "").strip()
                child_kind = (str(child_kind_raw).strip() if child_kind_raw is not None else "").strip()
                child_code = (str(child_code_raw).strip() if child_code_raw is not None else "").strip()

                if not any([parent_kind, parent_code, child_kind, child_code, qty_raw, unit_raw, remark_raw]):
                    continue

                if parent_kind not in (bom_svc.PARENT_FINISHED, bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
                    errors.append(f"第 {idx} 行：父项类别无效。")
                    continue
                if child_kind not in (bom_svc.PARENT_SEMI, bom_svc.PARENT_MATERIAL):
                    errors.append(f"第 {idx} 行：子项类别无效。")
                    continue
                if not parent_code:
                    errors.append(f"第 {idx} 行：父项编号不能为空。")
                    continue
                if not child_code:
                    errors.append(f"第 {idx} 行：子项编号不能为空。")
                    continue

                qty = None
                try:
                    qty = Decimal(str(qty_raw)) if qty_raw is not None else None
                except Exception:
                    qty = None
                if qty is None or qty <= 0:
                    errors.append(f"第 {idx} 行：用量数量须大于 0。")
                    continue

                unit = (str(unit_raw).strip() if unit_raw is not None else "").strip() or None
                if unit:
                    unit = unit[:16]

                remark = (str(remark_raw).strip() if remark_raw is not None else "").strip() or None
                if remark:
                    remark = remark[:255]

                # resolve parent id
                parent_id = None
                if parent_kind == bom_svc.PARENT_FINISHED:
                    parent_id = resolve_product_code(parent_code)
                else:
                    parent_id = resolve_semi_kind_code(parent_kind, parent_code)
                if not parent_id:
                    errors.append(f"第 {idx} 行：未找到父项（{parent_kind} / {parent_code}）。")
                    continue

                # resolve child id
                child_id = resolve_semi_kind_code(child_kind, child_code)
                if not child_id:
                    errors.append(f"第 {idx} 行：未找到子项（{child_kind} / {child_code}）。")
                    continue

                key = (parent_kind, parent_id)
                line_no = len(groups.get(key, [])) + 1
                groups.setdefault(key, []).append(
                    {
                        "line_no": line_no,
                        "child_kind": child_kind,
                        "child_material_id": child_id,
                        "quantity": qty,
                        "unit": unit,
                        "remark": remark,
                    }
                )

            # 逐组校验并写入
            success_cnt = 0
            try:
                for (pk, pid), glines in groups.items():
                    bom_svc.validate_bom_lines(parent_kind=pk, parent_id=pid, lines=glines)

                    parent_product_id, parent_material_id = _bom_parent_ids(pk, pid)
                    BomHeader.query.filter(
                        BomHeader.is_active.is_(True),
                        BomHeader.parent_kind == pk,
                        BomHeader.parent_product_id == parent_product_id,
                        BomHeader.parent_material_id == parent_material_id,
                    ).update({"is_active": False}, synchronize_session=False)

                    max_ver = (
                        BomHeader.query.filter(
                            BomHeader.parent_kind == pk,
                            BomHeader.parent_product_id == parent_product_id,
                            BomHeader.parent_material_id == parent_material_id,
                        )
                        .with_entities(func.max(BomHeader.version_no))
                        .scalar()
                    )
                    next_ver = int(max_ver or 0) + 1

                    header = BomHeader(
                        parent_kind=pk,
                        parent_product_id=parent_product_id,
                        parent_material_id=parent_material_id,
                        version_no=next_ver,
                        is_active=True,
                        remark=None,
                    )
                    db.session.add(header)
                    db.session.flush()

                    for ln in glines:
                        db.session.add(
                            BomLine(
                                bom_header_id=header.id,
                                line_no=ln["line_no"],
                                child_kind=ln["child_kind"],
                                child_material_id=ln["child_material_id"],
                                quantity=ln["quantity"],
                                unit=ln.get("unit"),
                                remark=ln.get("remark"),
                            )
                        )

                    success_cnt += 1

                if success_cnt:
                    db.session.commit()
                else:
                    db.session.rollback()
            except ValueError as e:
                db.session.rollback()
                errors.append(f"导入校验失败：{e}")
            except IntegrityError:
                db.session.rollback()
                errors.append("导入失败：数据冲突。")

            result = {"success": success_cnt, "errors": errors}

        return render_template("bom/import.html", result=result)

