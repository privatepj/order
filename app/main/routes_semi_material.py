from __future__ import annotations

from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy.exc import IntegrityError

from app import db
from app.auth.capabilities import current_user_can_cap
from app.auth.decorators import capability_required, menu_required
from app.models import SemiMaterial
from app.utils.form_display import clean_optional_text
from app.utils.query import keyword_like_or


def _next_code_for_kind(kind: str) -> str:
    prefix = "SM" if kind == "semi" else "MT"
    m = 0
    for (c,) in (
        db.session.query(SemiMaterial.code)
        .filter(SemiMaterial.code.like(f"{prefix}%"))
        .all()
    ):
        if not c:
            continue
        c = str(c)
        if not c.startswith(prefix):
            continue
        tail = c[len(prefix) :]
        if tail.isdigit():
            try:
                m = max(m, int(tail))
            except ValueError:
                pass
    return f"{prefix}{m + 1:04d}"


def _bump_code(code: str) -> str:
    # 简单兜底：尝试把最后一段数字加 1，不行就拼 N
    import re

    s = (code or "").strip()
    m = re.match(r"^(SM|MT)(\d+)$", s)
    if not m:
        return s + "N"
    prefix = m.group(1)
    tail = m.group(2)
    try:
        return f"{prefix}{int(tail) + 1:04d}"
    except ValueError:
        return s + "N"


def register_semi_material_routes(bp):
    # ----- 列表 -----
    @bp.route("/semi-materials")
    @login_required
    @menu_required("semi_material")
    def semi_material_list():
        kind = (request.args.get("kind") or "semi").strip()
        if kind == "material":
            return redirect(
                url_for(
                    "main.procurement_material_list",
                    keyword=(request.args.get("keyword") or "").strip(),
                    page=request.args.get("page", 1, type=int),
                )
            )
        if kind not in ("semi", "material"):
            kind = "semi"
        keyword = (request.args.get("keyword") or "").strip()
        if not current_user_can_cap("semi_material.filter.keyword"):
            keyword = ""

        q = SemiMaterial.query.filter(SemiMaterial.kind == kind).order_by(SemiMaterial.code)
        cond = keyword_like_or(
            keyword,
            SemiMaterial.code,
            SemiMaterial.name,
            SemiMaterial.spec,
            SemiMaterial.nav_type,
            SemiMaterial.base_unit,
            SemiMaterial.remark,
        )
        if cond is not None:
            q = q.filter(cond)

        page = request.args.get("page", 1, type=int)
        pagination = q.paginate(page=page, per_page=20)
        return render_template(
            "semi_material/list.html",
            pagination=pagination,
            kind=kind,
            keyword=keyword,
        )

    # ----- 新增 -----
    @bp.route("/semi-materials/new", methods=["GET", "POST"])
    @login_required
    @menu_required("semi_material")
    @capability_required("semi_material.action.create")
    def semi_material_new():
        if request.method == "GET" and (request.args.get("kind") or "").strip() == "material":
            return redirect(url_for("main.procurement_material_new"))
        if request.method == "POST":
            kind = (request.form.get("kind") or "").strip()
            if kind not in ("semi", "material"):
                kind = "semi"

            name = (request.form.get("name") or "").strip()
            spec = (request.form.get("spec") or "").strip() or None
            base_unit = (request.form.get("base_unit") or "").strip() or None
            remark = (request.form.get("remark") or "").strip() or None

            if not name:
                flash("名称为必填。", "danger")
                return render_template("semi_material/form.html", item=None, kind=kind)

            item = SemiMaterial(kind=kind)
            item.code = _next_code_for_kind(kind)
            item.name = name
            item.spec = spec
            if kind == "semi":
                item.series = clean_optional_text(request.form.get("series"), max_len=64)
            item.base_unit = base_unit
            item.remark = remark
            item.nav_type = clean_optional_text(request.form.get("nav_type"), max_len=64)

            max_tries = 3
            db.session.add(item)
            for attempt in range(max_tries):
                try:
                    db.session.commit()
                    flash("主数据已保存。", "success")
                    return redirect(url_for("main.semi_material_list", kind=kind))
                except IntegrityError:
                    db.session.rollback()
                    db.session.expunge_all()
                    if item.code:
                        item.code = _bump_code(item.code)
                    db.session.add(item)
                    if attempt == max_tries - 1:
                        flash("保存失败：编码冲突，请稍后重试。", "danger")
                        return render_template("semi_material/form.html", item=None, kind=kind)

        kind = request.args.get("kind") or "semi"
        if kind not in ("semi", "material"):
            kind = "semi"
        return render_template("semi_material/form.html", item=None, kind=kind)

    # ----- 编辑 -----
    @bp.route("/semi-materials/<int:item_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("semi_material")
    @capability_required("semi_material.action.edit")
    def semi_material_edit(item_id: int):
        item = SemiMaterial.query.get_or_404(item_id)
        if item.kind == "material" and request.method == "GET":
            return redirect(url_for("main.procurement_material_edit", item_id=item.id))
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            spec = (request.form.get("spec") or "").strip() or None
            base_unit = (request.form.get("base_unit") or "").strip() or None
            remark = (request.form.get("remark") or "").strip() or None
            is_admin = getattr(current_user, "role_code", None) == "admin"
            if not name:
                flash("名称为必填。", "danger")
                return render_template("semi_material/form.html", item=item, kind=item.kind)
            item.name = name
            item.spec = spec
            if item.kind == "semi":
                item.series = clean_optional_text(request.form.get("series"), max_len=64)
            item.nav_type = clean_optional_text(request.form.get("nav_type"), max_len=64)
            item.base_unit = base_unit
            item.remark = remark
            if is_admin:
                su_raw = (request.form.get("standard_unit_cost") or "").strip()
                if su_raw == "":
                    item.standard_unit_cost = None
                else:
                    try:
                        v = float(su_raw)
                        if v < 0:
                            raise ValueError()
                        item.standard_unit_cost = v
                    except ValueError:
                        flash("标准单位成本格式不正确。", "danger")
                        return render_template("semi_material/form.html", item=item, kind=item.kind)
            db.session.add(item)
            try:
                db.session.commit()
                flash("主数据已更新。", "success")
                return redirect(url_for("main.semi_material_list", kind=item.kind))
            except IntegrityError:
                db.session.rollback()
                flash("保存失败：数据冲突。", "danger")

        return render_template("semi_material/form.html", item=item, kind=item.kind)

    # ----- 删除 -----
    @bp.route("/semi-materials/<int:item_id>/delete", methods=["POST"])
    @login_required
    @menu_required("semi_material")
    @capability_required("semi_material.action.delete")
    def semi_material_delete(item_id: int):
        item = SemiMaterial.query.get_or_404(item_id)
        kind = item.kind
        db.session.delete(item)
        db.session.commit()
        flash("主数据已删除。", "success")
        return redirect(url_for("main.semi_material_list", kind=kind))

    # ----- Excel：导入模板 -----
    @bp.route("/semi-materials/export-import-template", methods=["GET"])
    @login_required
    @menu_required("semi_material")
    @capability_required("semi_material.action.import")
    def semi_material_export_import_template():
        kind = (request.args.get("kind") or "semi").strip()
        if kind == "material":
            return redirect(url_for("main.procurement_material_export_import_template"))
        if kind not in ("semi", "material"):
            kind = "semi"

        from openpyxl import Workbook

        headers = ["物料编号（可留空）", "名称", "规格", "基础单位", "备注", "系列", "类型"]
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"半成品物料导入模板_{kind}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ----- Excel：导入 -----
    @bp.route("/semi-materials/import", methods=["GET", "POST"])
    @login_required
    @menu_required("semi_material")
    @capability_required("semi_material.action.import")
    def semi_material_import():
        kind = (request.args.get("kind") or "semi").strip()
        if kind == "material" and request.method == "GET":
            return redirect(url_for("main.procurement_material_import"))
        if kind not in ("semi", "material"):
            kind = "semi"

        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择要上传的 Excel 文件。", "danger")
                return render_template("semi_material/import.html", kind=kind, result=None)

            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务器缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("semi_material/import.html", kind=kind, result=None)

            try:
                wb = load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Excel 文件无法读取，请确认格式为 .xlsx。", "danger")
                return render_template("semi_material/import.html", kind=kind, result=None)

            success = 0
            errors = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                code, name, spec, base_unit, remark, series_col, nav_type_col = (
                    row + (None,) * 7
                )[:7]
                code = (code or "").strip() if isinstance(code, str) else (code or "")
                code = str(code).strip()
                name = (name or "").strip() if isinstance(name, str) else (name or "")
                name = str(name).strip()
                if not name:
                    if any(row):
                        errors.append(f"第 {idx} 行：名称为空")
                    continue

                spec = (spec or "").strip() if isinstance(spec, str) else (spec or "")
                spec = spec or None
                base_unit = (base_unit or "").strip() if isinstance(base_unit, str) else (base_unit or "")
                base_unit = base_unit or None
                remark = (remark or "").strip() if isinstance(remark, str) else (remark or "")
                remark = remark or None
                nav_val = clean_optional_text(nav_type_col, max_len=64)

                # code 可留空：系统自动生成
                if not code:
                    code = _next_code_for_kind(kind)
                    while SemiMaterial.query.filter_by(code=code).first():
                        code = _bump_code(code)

                existing = SemiMaterial.query.filter_by(code=code).first()
                if existing and existing.kind != kind:
                    errors.append(f"第 {idx} 行：编号已存在但类别不匹配")
                    continue

                if existing:
                    existing.kind = kind
                    existing.name = name
                    existing.spec = spec
                    existing.base_unit = base_unit
                    existing.remark = remark
                    existing.nav_type = nav_val
                    if kind == "semi":
                        existing.series = clean_optional_text(series_col, max_len=64)
                    else:
                        existing.series = None
                    db.session.add(existing)
                else:
                    db.session.add(
                        SemiMaterial(
                            kind=kind,
                            code=code,
                            name=name,
                            spec=spec,
                            base_unit=base_unit,
                            remark=remark,
                            nav_type=nav_val,
                            series=clean_optional_text(series_col, max_len=64)
                            if kind == "semi"
                            else None,
                        )
                    )

                success += 1

                # 避免大文件一次 flush 太大
                if success % 200 == 0:
                    db.session.flush()

            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("导入失败：数据冲突。请重试或检查编码重复。", "danger")
                return render_template("semi_material/import.html", kind=kind, result=None)

            if success:
                flash(f"成功导入/更新 {success} 条。", "success")
            if errors:
                flash(f"有 {len(errors)} 条记录导入失败，请查看错误列表。", "danger")
            result = {"success": success, "errors": errors}
            return render_template("semi_material/import.html", kind=kind, result=result)

        return render_template("semi_material/import.html", kind=kind, result=None)

