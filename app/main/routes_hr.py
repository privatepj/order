from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from flask import flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError

from app import db
from app.auth.capabilities import current_user_can_cap
from app.auth.decorators import capability_required, menu_required
from app.auth.menus import first_landing_url
from app.models import (
    AuditLog,
    Company,
    HrDepartment,
    HrDepartmentWorkTypeMap,
    HrEmployee,
    HrEmployeeCapability,
    HrEmployeeScheduleBooking,
    HrEmployeeWorkType,
    HrPayrollLine,
    HrPerformanceReview,
    HrWorkType,
    HrWorkTypePieceRate,
    User,
)
from app.services.hr_work_type_svc import (
    allowed_work_type_ids_for_department,
    find_work_type_by_name,
    list_work_types,
    secondary_work_type_ids,
    sync_employee_work_types,
)
from app.utils.hr_privacy import display_id_card, display_phone


def _trunc(s, n):
    if s is None:
        return None
    s = str(s)
    if len(s) <= n:
        return s
    return s[: n - 3] + "..."


def _hr_audit(action: str, payload: dict):
    from flask import has_request_context, request

    if not has_request_context():
        return
    uid = int(current_user.get_id()) if current_user.is_authenticated else None
    row = AuditLog(
        event_type="hr",
        method=_trunc(request.method, 8),
        path=_trunc(request.path, 512),
        query_string=_trunc(
            request.query_string.decode("utf-8", errors="replace") if request.query_string else None,
            2048,
        ),
        status_code=None,
        duration_ms=None,
        user_id=uid,
        auth_type="session" if uid else "anonymous",
        ip=_trunc(request.remote_addr or "", 45) or "-",
        user_agent=_trunc(request.headers.get("User-Agent"), 512),
        endpoint=_trunc(request.endpoint or "", 128) if request.endpoint else None,
        extra={"action": action, **payload},
    )
    try:
        db.session.add(row)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _companies():
    return Company.query.order_by(Company.id).all()


def _resolve_company_id():
    companies = _companies()
    return Company.get_default_id(), companies


def _period_first_day(period: str) -> date | None:
    try:
        year, month = (period or "").strip().split("-")
        return date(int(year), int(month), 1)
    except (TypeError, ValueError):
        return None


def _departments_for_company(company_id: int | None):
    if not company_id:
        return []
    return (
        HrDepartment.query.filter_by(company_id=company_id)
        .order_by(HrDepartment.sort_order.asc(), HrDepartment.id.asc())
        .all()
    )


def _work_types_for_company(company_id: int | None, *, active_only: bool = False):
    if not company_id:
        return []
    return list_work_types(company_id=company_id, active_only=active_only)


def _validate_work_type_ids(company_id: int, main_work_type_id: int | None, secondary_ids_raw):
    work_type_map = {int(row.id): row for row in _work_types_for_company(company_id, active_only=False)}
    if main_work_type_id and int(main_work_type_id) not in work_type_map:
        raise ValueError("主工种不存在或不属于当前主体。")

    secondary_ids: list[int] = []
    for raw in secondary_ids_raw or []:
        try:
            work_type_id = int(raw)
        except (TypeError, ValueError):
            continue
        if work_type_id <= 0:
            continue
        if work_type_id not in work_type_map:
            raise ValueError("辅工种不存在或不属于当前主体。")
        if main_work_type_id and int(main_work_type_id) == work_type_id:
            continue
        if work_type_id not in secondary_ids:
            secondary_ids.append(work_type_id)
    return secondary_ids


def register_hr_routes(bp):
    @bp.route("/hr-departments")
    @login_required
    @menu_required("hr_department")
    def hr_department_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        if not current_user_can_cap("hr_department.filter.keyword"):
            keyword = ""
        q = HrDepartment.query
        if company_id:
            q = q.filter(HrDepartment.company_id == company_id)
        if keyword:
            q = q.filter(HrDepartment.name.contains(keyword))
        rows = q.order_by(HrDepartment.sort_order.asc(), HrDepartment.id.asc()).all()
        return render_template(
            "hr/department_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
        )

    @bp.route("/hr-departments/new", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_department")
    @capability_required("hr_department.action.create")
    def hr_department_new():
        companies = _companies()
        if not companies:
            flash("请先建立经营主体。", "danger")
            return redirect(url_for("main.hr_department_list"))
        if request.method == "POST":
            company_id = Company.get_default_id()
            name = (request.form.get("name") or "").strip()
            sort_order = request.form.get("sort_order", type=int) or 0
            if not company_id:
                flash("请先设置默认经营主体。", "danger")
                return render_template("hr/department_form.html", dept=None, companies=companies, company_id=None)
            if not name:
                flash("部门名称为必填。", "danger")
                return render_template(
                    "hr/department_form.html",
                    dept=None,
                    companies=companies,
                    company_id=company_id,
                )
            row = HrDepartment(company_id=company_id, name=name, sort_order=sort_order)
            db.session.add(row)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("同一主体下部门名称已存在。", "danger")
                return render_template(
                    "hr/department_form.html",
                    dept=None,
                    companies=companies,
                    company_id=company_id,
                )
            flash("部门已新增。", "success")
            return redirect(url_for("main.hr_department_list", company_id=company_id))
        return render_template(
            "hr/department_form.html",
            dept=None,
            companies=companies,
            company_id=Company.get_default_id(),
        )

    @bp.route("/hr-departments/<int:dept_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_department")
    @capability_required("hr_department.action.edit")
    def hr_department_edit(dept_id):
        row = HrDepartment.query.get_or_404(dept_id)
        companies = _companies()
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            sort_order = request.form.get("sort_order", type=int) or 0
            if not name:
                flash("部门名称为必填。", "danger")
                return render_template(
                    "hr/department_form.html",
                    dept=row,
                    companies=companies,
                    company_id=row.company_id,
                )
            row.name = name
            row.sort_order = sort_order
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("同一主体下部门名称已存在。", "danger")
                return render_template(
                    "hr/department_form.html",
                    dept=row,
                    companies=companies,
                    company_id=row.company_id,
                )
            flash("已保存。", "success")
            return redirect(url_for("main.hr_department_list", company_id=row.company_id))
        return render_template(
            "hr/department_form.html",
            dept=row,
            companies=companies,
            company_id=row.company_id,
        )

    @bp.route("/hr-departments/<int:dept_id>/delete", methods=["POST"])
    @login_required
    @menu_required("hr_department")
    @capability_required("hr_department.action.delete")
    def hr_department_delete(dept_id):
        row = HrDepartment.query.get_or_404(dept_id)
        if HrEmployee.query.filter(HrEmployee.department_id == dept_id).first():
            flash("该部门下仍有人员，无法删除。", "danger")
            return redirect(url_for("main.hr_department_list", company_id=row.company_id))
        db.session.delete(row)
        db.session.commit()
        flash("部门已删除。", "success")
        return redirect(url_for("main.hr_department_list", company_id=row.company_id))

    @bp.route("/hr-work-types")
    @login_required
    @menu_required("hr_work_type")
    def hr_work_type_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        status = (request.args.get("status") or "").strip()
        q = HrWorkType.query
        if company_id:
            q = q.filter(HrWorkType.company_id == company_id)
        if keyword:
            q = q.filter(HrWorkType.name.contains(keyword))
        if status == "active":
            q = q.filter(HrWorkType.is_active.is_(True))
        elif status == "inactive":
            q = q.filter(HrWorkType.is_active.is_(False))
        rows = q.order_by(HrWorkType.sort_order.asc(), HrWorkType.id.asc()).all()
        return render_template(
            "hr/work_type_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
            status=status,
        )

    @bp.route("/hr-work-types/new", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_work_type")
    @capability_required("hr_work_type.action.create")
    def hr_work_type_new():
        companies = _companies()
        company_id = Company.get_default_id()
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            sort_order = request.form.get("sort_order", type=int) or 0
            is_active = (request.form.get("is_active") or "1") == "1"
            remark = (request.form.get("remark") or "").strip() or None
            if not company_id:
                flash("请先设置默认经营主体。", "danger")
                return render_template("hr/work_type_form.html", row=None, companies=companies, company_id=None)
            if not name:
                flash("工种名称为必填。", "danger")
                return render_template(
                    "hr/work_type_form.html",
                    row=None,
                    companies=companies,
                    company_id=company_id,
                )
            row = HrWorkType(
                company_id=company_id,
                name=name,
                sort_order=sort_order,
                is_active=is_active,
                remark=remark,
            )
            db.session.add(row)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("同一主体下工种名称已存在。", "danger")
                return render_template(
                    "hr/work_type_form.html",
                    row=None,
                    companies=companies,
                    company_id=company_id,
                )
            flash("工种已新增。", "success")
            return redirect(url_for("main.hr_work_type_list", company_id=company_id))
        return render_template(
            "hr/work_type_form.html",
            row=None,
            companies=companies,
            company_id=company_id,
        )

    @bp.route("/hr-work-types/<int:work_type_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_work_type")
    @capability_required("hr_work_type.action.edit")
    def hr_work_type_edit(work_type_id):
        row = HrWorkType.query.get_or_404(work_type_id)
        companies = _companies()
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            sort_order = request.form.get("sort_order", type=int) or 0
            is_active = (request.form.get("is_active") or "1") == "1"
            remark = (request.form.get("remark") or "").strip() or None
            if not name:
                flash("工种名称为必填。", "danger")
                return render_template(
                    "hr/work_type_form.html",
                    row=row,
                    companies=companies,
                    company_id=row.company_id,
                )
            row.name = name
            row.sort_order = sort_order
            row.is_active = is_active
            row.remark = remark
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("同一主体下工种名称已存在。", "danger")
                return render_template(
                    "hr/work_type_form.html",
                    row=row,
                    companies=companies,
                    company_id=row.company_id,
                )
            flash("已保存。", "success")
            return redirect(url_for("main.hr_work_type_list", company_id=row.company_id))
        return render_template(
            "hr/work_type_form.html",
            row=row,
            companies=companies,
            company_id=row.company_id,
        )

    @bp.route("/hr-work-types/<int:work_type_id>/delete", methods=["POST"])
    @login_required
    @menu_required("hr_work_type")
    @capability_required("hr_work_type.action.delete")
    def hr_work_type_delete(work_type_id):
        row = HrWorkType.query.get_or_404(work_type_id)
        company_id = int(row.company_id)
        used = any(
            [
                HrEmployee.query.filter(HrEmployee.main_work_type_id == work_type_id).first(),
                HrEmployeeWorkType.query.filter_by(work_type_id=work_type_id).first(),
                HrDepartmentWorkTypeMap.query.filter_by(work_type_id=work_type_id).first(),
                HrEmployeeCapability.query.filter_by(work_type_id=work_type_id).first(),
                HrEmployeeScheduleBooking.query.filter_by(work_type_id=work_type_id).first(),
                HrWorkTypePieceRate.query.filter_by(work_type_id=work_type_id).first(),
            ]
        )
        if used:
            flash("该工种已被引用，无法删除。", "danger")
            return redirect(url_for("main.hr_work_type_list", company_id=company_id))
        db.session.delete(row)
        db.session.commit()
        flash("工种已删除。", "success")
        return redirect(url_for("main.hr_work_type_list", company_id=company_id))

    @bp.route("/hr-employees")
    @login_required
    @menu_required("hr_employee")
    def hr_employee_list():
        company_id, companies = _resolve_company_id()
        keyword = (request.args.get("keyword") or "").strip()
        department_id = request.args.get("department_id", type=int) or 0
        work_type_id = request.args.get("work_type_id", type=int) or 0
        if not current_user_can_cap("hr_employee.filter.keyword"):
            keyword = ""
        q = HrEmployee.query
        if company_id:
            q = q.filter(HrEmployee.company_id == company_id)
        if department_id:
            q = q.filter(HrEmployee.department_id == department_id)
        if keyword:
            q = q.filter(
                or_(
                    HrEmployee.name.contains(keyword),
                    HrEmployee.employee_no.contains(keyword),
                    HrEmployee.phone.contains(keyword),
                )
            )
        if work_type_id:
            q = (
                q.outerjoin(HrEmployeeWorkType, HrEmployeeWorkType.employee_id == HrEmployee.id)
                .filter(
                    or_(
                        HrEmployee.main_work_type_id == work_type_id,
                        HrEmployeeWorkType.work_type_id == work_type_id,
                    )
                )
                .distinct()
            )
        rows = q.order_by(HrEmployee.company_id.asc(), HrEmployee.employee_no.asc()).all()
        return render_template(
            "hr/employee_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            keyword=keyword,
            department_id=department_id,
            work_type_id=work_type_id,
            departments=_departments_for_company(company_id),
            work_types=_work_types_for_company(company_id, active_only=False),
            show_sensitive=current_user_can_cap("hr_employee.view_sensitive"),
            display_phone=display_phone,
            display_id_card=display_id_card,
        )

    @bp.route("/hr-employees/new", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_employee")
    @capability_required("hr_employee.action.create")
    def hr_employee_new():
        companies = _companies()
        if not companies:
            flash("请先建立经营主体。", "danger")
            return redirect(url_for("main.hr_employee_list"))
        users = User.query.order_by(User.id.asc()).all()
        if request.method == "POST":
            return _hr_employee_save(None, companies, users)
        return _hr_employee_form(None, companies, users)

    @bp.route("/hr-employees/<int:emp_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_employee")
    @capability_required("hr_employee.action.edit")
    def hr_employee_edit(emp_id):
        emp = HrEmployee.query.get_or_404(emp_id)
        companies = _companies()
        users = User.query.order_by(User.id.asc()).all()
        if request.method == "POST":
            return _hr_employee_save(emp, companies, users)
        return _hr_employee_form(emp, companies, users)

    @bp.route("/hr-employees/<int:emp_id>/delete", methods=["POST"])
    @login_required
    @menu_required("hr_employee")
    @capability_required("hr_employee.action.delete")
    def hr_employee_delete(emp_id):
        emp = HrEmployee.query.get_or_404(emp_id)
        if HrPayrollLine.query.filter_by(employee_id=emp.id).first():
            flash("该人员已有工资记录，请先清理相关工资数据。", "danger")
            return redirect(url_for("main.hr_employee_list", company_id=emp.company_id))
        if HrPerformanceReview.query.filter_by(employee_id=emp.id).first():
            flash("该人员已有绩效记录，请先清理相关绩效数据。", "danger")
            return redirect(url_for("main.hr_employee_list", company_id=emp.company_id))
        db.session.delete(emp)
        db.session.commit()
        flash("人员档案已删除。", "success")
        return redirect(url_for("main.hr_employee_list", company_id=emp.company_id))

    @bp.route("/hr-employees/export-import-template", methods=["GET"])
    @login_required
    @menu_required("hr_employee")
    @capability_required("hr_employee.action.export_template")
    def hr_employee_export_import_template():
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "人员导入"
        ws.append(
            [
                "经营主体短码",
                "工号",
                "姓名",
                "部门名称",
                "手机",
                "身份证",
                "主工种",
                "辅工种列表(逗号分隔)",
                "入职日期",
                "状态",
            ]
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="人员导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/hr-employees/import", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_employee")
    @capability_required("hr_employee.action.import")
    def hr_employee_import():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("请先选择 Excel 文件。", "danger")
                return render_template("hr/employee_import.html", result=None)

            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("服务端缺少 openpyxl 依赖，无法导入。", "danger")
                return render_template("hr/employee_import.html", result=None)

            try:
                ws = load_workbook(file, data_only=True).active
            except Exception:
                flash("Excel 无法读取。", "danger")
                return render_template("hr/employee_import.html", result=None)

            errors: list[str] = []
            success = 0
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                (
                    company_code,
                    employee_no,
                    name,
                    department_name,
                    phone,
                    id_card,
                    main_work_type_name,
                    secondary_work_types_raw,
                    hire_raw,
                    status,
                ) = (row + (None,) * 10)[:10]

                company_code = (str(company_code).strip() if company_code is not None else "") or ""
                employee_no = (str(employee_no).strip() if employee_no is not None else "") or ""
                name = (str(name).strip() if name is not None else "") or ""
                if not company_code or not employee_no or not name:
                    errors.append(f"第 {idx} 行：经营主体短码、工号、姓名为必填")
                    continue

                company = Company.query.filter_by(code=company_code).first()
                if not company:
                    errors.append(f"第 {idx} 行：经营主体短码 {company_code} 不存在")
                    continue

                department_id = None
                if department_name:
                    department = HrDepartment.query.filter_by(
                        company_id=company.id,
                        name=str(department_name).strip(),
                    ).first()
                    if not department:
                        errors.append(f"第 {idx} 行：部门“{department_name}”不存在，请先维护部门")
                        continue
                    department_id = int(department.id)

                main_work_type = None
                if main_work_type_name:
                    main_work_type = find_work_type_by_name(
                        company_id=company.id,
                        name=str(main_work_type_name).strip(),
                    )
                    if not main_work_type:
                        errors.append(f"第 {idx} 行：主工种“{main_work_type_name}”不存在，请先维护工种")
                        continue

                secondary_work_type_ids: list[int] = []
                missing_names: list[str] = []
                secondary_names = [
                    item.strip()
                    for item in str(secondary_work_types_raw or "").replace("，", ",").split(",")
                    if item and str(item).strip()
                ]
                for name_item in secondary_names:
                    wt = find_work_type_by_name(company_id=company.id, name=name_item)
                    if not wt:
                        missing_names.append(name_item)
                        continue
                    if int(wt.id) not in secondary_work_type_ids:
                        secondary_work_type_ids.append(int(wt.id))
                if missing_names:
                    errors.append(
                        f"第 {idx} 行：辅工种不存在：{', '.join(missing_names)}，请先维护工种"
                    )
                    continue

                hire_date = None
                if hire_raw:
                    if isinstance(hire_raw, datetime):
                        hire_date = hire_raw.date()
                    else:
                        try:
                            hire_date = datetime.strptime(str(hire_raw).strip()[:10], "%Y-%m-%d").date()
                        except ValueError:
                            errors.append(f"第 {idx} 行：入职日期格式应为 YYYY-MM-DD")
                            continue

                status_value = (str(status).strip().lower() if status else "active") or "active"
                if status_value not in ("active", "left"):
                    status_value = "active"

                emp = HrEmployee.query.filter_by(
                    company_id=company.id,
                    employee_no=employee_no,
                ).first()
                if not emp:
                    emp = HrEmployee(company_id=company.id, employee_no=employee_no)
                    db.session.add(emp)
                emp.name = name
                emp.department_id = department_id
                emp.phone = (str(phone).strip() if phone else None) or None
                emp.id_card = (str(id_card).strip() if id_card else None) or None
                emp.hire_date = hire_date
                emp.status = status_value
                if status_value == "left" and not emp.leave_date:
                    emp.leave_date = date.today()
                db.session.flush()
                sync_employee_work_types(
                    employee=emp,
                    main_work_type_id=int(main_work_type.id) if main_work_type else None,
                    secondary_work_type_ids=secondary_work_type_ids,
                )
                success += 1

            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("导入失败：数据冲突（工号重复等）。", "danger")
                return render_template("hr/employee_import.html", result=None)

            _hr_audit("employee_import", {"success": success, "errors": len(errors)})
            return render_template("hr/employee_import.html", result={"success": success, "errors": errors})

        return render_template("hr/employee_import.html", result=None)

    def _hr_employee_form(emp, companies, users):
        company_id = emp.company_id if emp else Company.get_default_id()
        return render_template(
            "hr/employee_form.html",
            emp=emp,
            companies=companies,
            users=users,
            departments=_departments_for_company(company_id),
            work_types=_work_types_for_company(company_id, active_only=False),
            selected_secondary_work_type_ids=secondary_work_type_ids(emp),
            show_sensitive=current_user_can_cap("hr_employee.view_sensitive"),
            company_id=company_id,
            default_company_id=company_id,
        )

    def _hr_employee_save(emp, companies, users):
        company_id = emp.company_id if emp else Company.get_default_id()
        employee_no = (request.form.get("employee_no") or "").strip()
        name = (request.form.get("name") or "").strip()
        department_id = request.form.get("department_id", type=int) or None
        user_id = request.form.get("user_id", type=int) or None
        phone = (request.form.get("phone") or "").strip() or None
        id_card = (request.form.get("id_card") or "").strip() or None
        main_work_type_id = request.form.get("main_work_type_id", type=int) or None
        secondary_work_type_ids_raw = request.form.getlist("secondary_work_type_ids")
        status = (request.form.get("status") or "active").strip()
        hire_raw = (request.form.get("hire_date") or "").strip()
        leave_raw = (request.form.get("leave_date") or "").strip()
        remark = (request.form.get("remark") or "").strip() or None

        if status not in ("active", "left"):
            status = "active"
        if not company_id:
            flash("请先设置默认经营主体。", "danger")
            return _hr_employee_form(emp, companies, users)
        if not employee_no or not name:
            flash("工号与姓名为必填。", "danger")
            return _hr_employee_form(emp, companies, users)
        if department_id:
            department = HrDepartment.query.get(department_id)
            if not department or int(department.company_id) != int(company_id):
                flash("部门与经营主体不匹配。", "danger")
                return _hr_employee_form(emp, companies, users)
        if user_id and not User.query.get(user_id):
            flash("所选登录用户不存在。", "danger")
            return _hr_employee_form(emp, companies, users)
        try:
            secondary_work_type_ids_clean = _validate_work_type_ids(
                company_id,
                main_work_type_id,
                secondary_work_type_ids_raw,
            )
        except ValueError as exc:
            flash(str(exc), "danger")
            return _hr_employee_form(emp, companies, users)

        hire_date = None
        if hire_raw:
            try:
                hire_date = datetime.strptime(hire_raw[:10], "%Y-%m-%d").date()
            except ValueError:
                flash("入职日期格式错误。", "danger")
                return _hr_employee_form(emp, companies, users)
        leave_date = None
        if leave_raw:
            try:
                leave_date = datetime.strptime(leave_raw[:10], "%Y-%m-%d").date()
            except ValueError:
                flash("离职日期格式错误。", "danger")
                return _hr_employee_form(emp, companies, users)

        if not emp:
            emp = HrEmployee(company_id=company_id, employee_no=employee_no)
            db.session.add(emp)
        else:
            emp.company_id = company_id
            emp.employee_no = employee_no
        emp.name = name
        emp.department_id = department_id
        emp.user_id = user_id
        if current_user_can_cap("hr_employee.view_sensitive"):
            emp.phone = phone
            emp.id_card = id_card
        elif phone or id_card:
            flash("当前账号无敏感字段编辑权限，已保留原值。", "warning")
        emp.status = status
        emp.hire_date = hire_date
        emp.leave_date = leave_date
        emp.remark = remark
        db.session.flush()
        sync_employee_work_types(
            employee=emp,
            main_work_type_id=main_work_type_id,
            secondary_work_type_ids=secondary_work_type_ids_clean,
        )
        allowed_ids = set(
            allowed_work_type_ids_for_department(company_id=company_id, department_id=department_id)
        )
        selected_ids = set(([main_work_type_id] if main_work_type_id else []) + secondary_work_type_ids_clean)
        if allowed_ids and selected_ids - allowed_ids:
            flash("当前部门未配置部分工种允许关系，已保存，但建议检查“部门-工种允许关系”。", "warning")
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("同一主体下工号已存在。", "danger")
            return _hr_employee_form(emp, companies, users)
        flash("已保存。", "success")
        return redirect(url_for("main.hr_employee_list", company_id=company_id))

    @bp.route("/hr-payroll")
    @login_required
    @menu_required("hr_payroll")
    def hr_payroll_list():
        if not current_user_can_cap("hr_payroll.view"):
            flash("无权限查看工资。", "danger")
            return redirect(first_landing_url())
        company_id, companies = _resolve_company_id()
        period = (request.args.get("period") or "").strip() or date.today().strftime("%Y-%m")
        q = HrPayrollLine.query
        if company_id:
            q = q.filter(HrPayrollLine.company_id == company_id)
        rows = q.filter(HrPayrollLine.period == period).order_by(HrPayrollLine.id.asc()).all()
        return render_template(
            "hr/payroll_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            period=period,
        )

    @bp.route("/hr-payroll/new", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_payroll")
    @capability_required("hr_payroll.edit")
    def hr_payroll_new():
        companies = _companies()
        company_id, _ = _resolve_company_id()
        if request.method == "POST":
            employee_id = request.form.get("employee_id", type=int)
            period = (request.form.get("period") or "").strip()
            wage_kind = (request.form.get("wage_kind") or "monthly").strip().lower()
            work_hours_raw = (request.form.get("work_hours") or "").strip()
            hourly_rate_raw = (request.form.get("hourly_rate") or "").strip()
            base = Decimal(request.form.get("base_salary") or "0")
            allowance = Decimal(request.form.get("allowance") or "0")
            deduction = Decimal(request.form.get("deduction") or "0")
            remark = (request.form.get("remark") or "").strip() or None
            if not company_id or not employee_id or len(period) != 7:
                flash("请填写经营主体、员工与账期（YYYY-MM）。", "danger")
                return _payroll_form(None, companies, default_company_id=company_id)
            if wage_kind not in ("monthly", "hourly", "piece_rate"):
                flash("工资口径不合法。", "danger")
                return _payroll_form(None, companies, default_company_id=company_id)
            emp = HrEmployee.query.get(employee_id)
            if not emp or int(emp.company_id) != int(company_id):
                flash("员工与主体不匹配。", "danger")
                return _payroll_form(None, companies, default_company_id=company_id)
            if emp.status == "left" and emp.leave_date:
                period_day = _period_first_day(period)
                if period_day and period_day > emp.leave_date:
                    flash("该员工已离职，账期晚于离职日期，不能录入工资。", "danger")
                    return _payroll_form(None, companies, default_company_id=company_id)
            try:
                work_hours = Decimal(work_hours_raw or "0")
            except Exception:
                work_hours = Decimal("0")
            try:
                hourly_rate = Decimal(hourly_rate_raw or "0")
            except Exception:
                hourly_rate = Decimal("0")
            if wage_kind != "piece_rate" and work_hours <= 0:
                flash("请填写工时（小时，>0）。", "danger")
                return _payroll_form(None, companies, default_company_id=company_id)
            if wage_kind == "hourly":
                if hourly_rate <= 0:
                    flash("时薪必须大于 0。", "danger")
                    return _payroll_form(None, companies, default_company_id=company_id)
                net_pay = hourly_rate * work_hours
            elif wage_kind == "piece_rate":
                net_pay = Decimal("0")
            else:
                net_pay = base + allowance - deduction
            row = HrPayrollLine(
                company_id=company_id,
                employee_id=employee_id,
                period=period,
                wage_kind=wage_kind,
                work_hours=work_hours,
                hourly_rate=hourly_rate if wage_kind == "hourly" else Decimal("0"),
                base_salary=base,
                allowance=allowance,
                deduction=deduction,
                net_pay=net_pay,
                remark=remark,
                created_by=int(current_user.get_id()),
            )
            db.session.add(row)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("该员工该账期已有工资记录，请编辑原记录。", "danger")
                return _payroll_form(None, companies, default_company_id=company_id)
            _hr_audit("payroll_create", {"company_id": company_id, "employee_id": employee_id, "period": period})
            flash("工资已保存。", "success")
            return redirect(url_for("main.hr_payroll_list", company_id=company_id, period=period))
        return _payroll_form(
            None,
            companies,
            default_company_id=company_id,
            default_period=date.today().strftime("%Y-%m"),
        )

    @bp.route("/hr-payroll/<int:line_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_payroll")
    @capability_required("hr_payroll.edit")
    def hr_payroll_edit(line_id):
        line = HrPayrollLine.query.get_or_404(line_id)
        companies = _companies()
        if request.method == "POST":
            wage_kind = (request.form.get("wage_kind") or "monthly").strip().lower()
            work_hours_raw = (request.form.get("work_hours") or "").strip()
            hourly_rate_raw = (request.form.get("hourly_rate") or "").strip()
            base = Decimal(request.form.get("base_salary") or "0")
            allowance = Decimal(request.form.get("allowance") or "0")
            deduction = Decimal(request.form.get("deduction") or "0")
            remark = (request.form.get("remark") or "").strip() or None
            if wage_kind not in ("monthly", "hourly", "piece_rate"):
                flash("工资口径不合法。", "danger")
                return _payroll_form(line, companies)
            try:
                work_hours = Decimal(work_hours_raw or "0")
            except Exception:
                work_hours = Decimal("0")
            try:
                hourly_rate = Decimal(hourly_rate_raw or "0")
            except Exception:
                hourly_rate = Decimal("0")
            if wage_kind != "piece_rate" and work_hours <= 0:
                flash("请填写工时（小时，>0）。", "danger")
                return _payroll_form(line, companies)
            if wage_kind == "hourly":
                if hourly_rate <= 0:
                    flash("时薪必须大于 0。", "danger")
                    return _payroll_form(line, companies)
                net_pay = hourly_rate * work_hours
            elif wage_kind == "piece_rate":
                net_pay = Decimal("0")
            else:
                net_pay = base + allowance - deduction
            line.wage_kind = wage_kind
            line.work_hours = work_hours
            line.hourly_rate = hourly_rate if wage_kind == "hourly" else Decimal("0")
            line.base_salary = base
            line.allowance = allowance
            line.deduction = deduction
            line.net_pay = net_pay
            line.remark = remark
            db.session.commit()
            _hr_audit("payroll_edit", {"line_id": line_id, "period": line.period})
            flash("已保存。", "success")
            return redirect(url_for("main.hr_payroll_list", company_id=line.company_id, period=line.period))
        return _payroll_form(line, companies)

    def _payroll_form(line, companies, default_company_id=None, default_period=None):
        company_id = line.company_id if line else (default_company_id or (companies[0].id if companies else None))
        employees = []
        if company_id:
            employees = (
                HrEmployee.query.filter_by(company_id=company_id)
                .order_by(HrEmployee.employee_no.asc())
                .all()
            )
        return render_template(
            "hr/payroll_form.html",
            line=line,
            companies=companies,
            employees=employees,
            default_company_id=company_id,
            default_period=default_period or date.today().strftime("%Y-%m"),
        )

    @bp.route("/hr/dept-piece-rate", methods=["GET"])
    @login_required
    @menu_required("dept_piece_rate")
    @capability_required("dept_piece_rate.view")
    def dept_piece_rate_list():
        company_id, companies = _resolve_company_id()
        period = (request.args.get("period") or "").strip() or date.today().strftime("%Y-%m")
        q = HrWorkTypePieceRate.query
        if company_id:
            q = q.filter(HrWorkTypePieceRate.company_id == company_id)
        if period:
            q = q.filter(HrWorkTypePieceRate.period == period)
        rows = q.order_by(HrWorkTypePieceRate.id.asc()).all()
        return render_template(
            "hr/dept_piece_rate_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            period=period,
        )

    @bp.route("/hr/dept-piece-rate/new", methods=["GET", "POST"])
    @login_required
    @menu_required("dept_piece_rate")
    @capability_required("dept_piece_rate.edit")
    def dept_piece_rate_new():
        companies = _companies()
        company_id, _ = _resolve_company_id()
        if request.method == "POST":
            work_type_id = request.form.get("work_type_id", type=int)
            period = (request.form.get("period") or "").strip()
            rate_raw = (request.form.get("rate_per_unit") or "").strip()
            remark = (request.form.get("remark") or "").strip() or None
            if not company_id or not work_type_id or len(period) != 7:
                flash("请填写经营主体、工种与账期（YYYY-MM）。", "danger")
                return _dept_piece_rate_form(None, companies, default_company_id=company_id)
            if not HrWorkType.query.filter_by(company_id=company_id, id=work_type_id).first():
                flash("工种不存在或不属于当前主体。", "danger")
                return _dept_piece_rate_form(None, companies, default_company_id=company_id)
            try:
                rate = Decimal(rate_raw or "0")
            except Exception:
                rate = Decimal("0")
            if rate <= 0:
                flash("计件单价必须大于 0。", "danger")
                return _dept_piece_rate_form(None, companies, default_company_id=company_id)
            row = HrWorkTypePieceRate(
                company_id=company_id,
                work_type_id=work_type_id,
                period=period,
                rate_per_unit=rate,
                remark=remark,
                created_by=int(current_user.get_id()),
            )
            db.session.add(row)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("该工种该账期已有单价记录，请编辑原记录。", "danger")
                return _dept_piece_rate_form(None, companies, default_company_id=company_id)
            _hr_audit("work_type_piece_rate_create", {"company_id": company_id, "work_type_id": work_type_id, "period": period})
            flash("工种计件单价已保存。", "success")
            return redirect(url_for("main.dept_piece_rate_list", company_id=company_id, period=period))
        return _dept_piece_rate_form(
            None,
            companies,
            default_company_id=company_id,
            default_period=date.today().strftime("%Y-%m"),
        )

    @bp.route("/hr/dept-piece-rate/<int:rate_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("dept_piece_rate")
    @capability_required("dept_piece_rate.edit")
    def dept_piece_rate_edit(rate_id):
        row = HrWorkTypePieceRate.query.get_or_404(rate_id)
        companies = _companies()
        if request.method == "POST":
            rate_raw = (request.form.get("rate_per_unit") or "").strip()
            remark = (request.form.get("remark") or "").strip() or None
            try:
                rate = Decimal(rate_raw or "0")
            except Exception:
                rate = Decimal("0")
            if rate <= 0:
                flash("计件单价必须大于 0。", "danger")
                return _dept_piece_rate_form(row, companies)
            row.rate_per_unit = rate
            row.remark = remark
            db.session.commit()
            _hr_audit("work_type_piece_rate_edit", {"rate_id": rate_id, "period": row.period})
            flash("已保存。", "success")
            return redirect(url_for("main.dept_piece_rate_list", company_id=row.company_id, period=row.period))
        return _dept_piece_rate_form(row, companies)

    def _dept_piece_rate_form(row, companies, default_company_id=None, default_period=None):
        company_id = row.company_id if row else (default_company_id or (companies[0].id if companies else None))
        return render_template(
            "hr/dept_piece_rate_form.html",
            row=row,
            companies=companies,
            work_types=_work_types_for_company(company_id, active_only=False),
            default_company_id=company_id,
            default_period=default_period or date.today().strftime("%Y-%m"),
        )

    @bp.route("/hr/dept-piece-rate/<int:rate_id>/delete", methods=["POST"])
    @login_required
    @menu_required("dept_piece_rate")
    @capability_required("dept_piece_rate.edit")
    def dept_piece_rate_delete(rate_id):
        row = HrWorkTypePieceRate.query.get_or_404(rate_id)
        company_id = row.company_id
        period = row.period
        db.session.delete(row)
        db.session.commit()
        _hr_audit("work_type_piece_rate_delete", {"rate_id": rate_id, "period": period})
        flash("已删除。", "success")
        return redirect(url_for("main.dept_piece_rate_list", company_id=company_id, period=period))

    @bp.route("/hr-payroll/export", methods=["GET"])
    @login_required
    @menu_required("hr_payroll")
    @capability_required("hr_payroll.export")
    def hr_payroll_export():
        from openpyxl import Workbook

        company_id, _ = _resolve_company_id()
        period = (request.args.get("period") or "").strip() or date.today().strftime("%Y-%m")
        q = HrPayrollLine.query.filter(HrPayrollLine.period == period)
        if company_id:
            q = q.filter(HrPayrollLine.company_id == company_id)
        rows = q.order_by(HrPayrollLine.id.asc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "工资"
        ws.append(["账期", "主体ID", "工号", "姓名", "工资口径", "工时(小时)", "时薪", "基本工资", "补贴", "扣款", "实发", "备注"])
        for row in rows:
            emp = row.employee
            ws.append(
                [
                    row.period,
                    row.company_id,
                    emp.employee_no if emp else "",
                    emp.name if emp else "",
                    row.wage_kind,
                    float(row.work_hours or 0),
                    float(row.hourly_rate or 0),
                    float(row.base_salary or 0),
                    float(row.allowance or 0),
                    float(row.deduction or 0),
                    float(row.net_pay or 0),
                    row.remark or "",
                ]
            )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"工资导出_{period}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @bp.route("/hr-performance")
    @login_required
    @menu_required("hr_performance")
    def hr_performance_list():
        company_id, companies = _resolve_company_id()
        cycle = (request.args.get("cycle") or "").strip()
        q = HrPerformanceReview.query
        if company_id:
            q = q.filter(HrPerformanceReview.company_id == company_id)
        if cycle:
            q = q.filter(HrPerformanceReview.cycle == cycle)
        rows = q.order_by(HrPerformanceReview.cycle.desc(), HrPerformanceReview.id.asc()).all()
        return render_template(
            "hr/performance_list.html",
            rows=rows,
            companies=companies,
            company_id=company_id,
            cycle=cycle,
        )

    @bp.route("/hr-performance/new", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_performance")
    @capability_required("hr_performance.action.create")
    def hr_performance_new():
        companies = _companies()
        company_id, _ = _resolve_company_id()
        if request.method == "POST":
            employee_id = request.form.get("employee_id", type=int)
            cycle = (request.form.get("cycle") or "").strip()
            score_raw = request.form.get("score")
            comment = (request.form.get("comment") or "").strip() or None
            if not company_id or not employee_id or not cycle:
                flash("请填写经营主体、员工与考核周期。", "danger")
                return _perf_form(None, companies)
            emp = HrEmployee.query.get(employee_id)
            if not emp or int(emp.company_id) != int(company_id):
                flash("员工与主体不匹配。", "danger")
                return _perf_form(None, companies)
            score = None
            if score_raw not in (None, ""):
                try:
                    score = Decimal(str(score_raw))
                except Exception:
                    flash("评分格式错误。", "danger")
                    return _perf_form(None, companies)
            row = HrPerformanceReview(
                company_id=company_id,
                employee_id=employee_id,
                cycle=cycle,
                score=score,
                comment=comment,
                reviewer_user_id=int(current_user.get_id()),
                status="draft",
            )
            db.session.add(row)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("该员工该周期已有绩效记录。", "danger")
                return _perf_form(None, companies)
            flash("已保存。", "success")
            return redirect(url_for("main.hr_performance_list", company_id=company_id, cycle=cycle))
        return _perf_form(None, companies, default_company_id=company_id)

    @bp.route("/hr-performance/<int:pr_id>/edit", methods=["GET", "POST"])
    @login_required
    @menu_required("hr_performance")
    def hr_performance_edit(pr_id):
        row = HrPerformanceReview.query.get_or_404(pr_id)
        role_code = getattr(current_user, "role_code", None)
        if row.status == "finalized" and not current_user_can_cap("hr_performance.action.override") and role_code != "admin":
            flash("该绩效已定稿，无权限修改。", "danger")
            return redirect(url_for("main.hr_performance_list", company_id=row.company_id))
        if row.status != "finalized" and not current_user_can_cap("hr_performance.action.edit") and role_code != "admin":
            flash("无权限编辑。", "danger")
            return redirect(url_for("main.hr_performance_list", company_id=row.company_id))
        companies = _companies()
        if request.method == "POST":
            score_raw = request.form.get("score")
            comment = (request.form.get("comment") or "").strip() or None
            score = None
            if score_raw not in (None, ""):
                try:
                    score = Decimal(str(score_raw))
                except Exception:
                    flash("评分格式错误。", "danger")
                    return _perf_form(row, companies)
            row.score = score
            row.comment = comment
            row.reviewer_user_id = int(current_user.get_id())
            db.session.commit()
            _hr_audit("performance_edit", {"id": pr_id, "cycle": row.cycle})
            flash("已保存。", "success")
            return redirect(url_for("main.hr_performance_list", company_id=row.company_id, cycle=row.cycle))
        return _perf_form(row, companies)

    @bp.route("/hr-performance/<int:pr_id>/delete", methods=["POST"])
    @login_required
    @menu_required("hr_performance")
    @capability_required("hr_performance.action.delete")
    def hr_performance_delete(pr_id):
        row = HrPerformanceReview.query.get_or_404(pr_id)
        company_id = row.company_id
        cycle = row.cycle
        db.session.delete(row)
        db.session.commit()
        flash("已删除。", "success")
        return redirect(url_for("main.hr_performance_list", company_id=company_id, cycle=cycle))

    @bp.route("/hr-performance/<int:pr_id>/finalize", methods=["POST"])
    @login_required
    @menu_required("hr_performance")
    @capability_required("hr_performance.action.finalize")
    def hr_performance_finalize(pr_id):
        row = HrPerformanceReview.query.get_or_404(pr_id)
        if row.status == "finalized":
            flash("当前记录已定稿。", "info")
            return redirect(url_for("main.hr_performance_list", company_id=row.company_id))
        row.status = "finalized"
        row.reviewer_user_id = int(current_user.get_id())
        db.session.commit()
        _hr_audit("performance_finalize", {"id": pr_id, "cycle": row.cycle})
        flash("已定稿。", "success")
        return redirect(url_for("main.hr_performance_list", company_id=row.company_id, cycle=row.cycle))

    def _perf_form(pr, companies, default_company_id=None):
        company_id = pr.company_id if pr else (default_company_id or (companies[0].id if companies else None))
        employees = []
        if company_id:
            employees = (
                HrEmployee.query.filter_by(company_id=company_id)
                .order_by(HrEmployee.employee_no.asc())
                .all()
            )
        return render_template(
            "hr/performance_form.html",
            pr=pr,
            companies=companies,
            employees=employees,
            default_company_id=company_id,
        )

    @bp.route("/hr/department-capability-map")
    @login_required
    @menu_required("hr_department_capability_map")
    @capability_required("hr_department_capability_map.view")
    def hr_department_capability_map_list():
        company_id, companies = _resolve_company_id()
        rows = (
            HrDepartmentWorkTypeMap.query.filter_by(company_id=company_id)
            .order_by(HrDepartmentWorkTypeMap.id.desc())
            .all()
        )
        return render_template(
            "hr/dept_capability_map.html",
            rows=rows,
            depts=_departments_for_company(company_id),
            work_types=_work_types_for_company(company_id, active_only=False),
            company_id=company_id,
            companies=companies,
        )

    @bp.route("/hr/department-capability-map/add", methods=["POST"])
    @login_required
    @menu_required("hr_department_capability_map")
    @capability_required("hr_department_capability_map.edit")
    def hr_department_capability_map_add():
        company_id, _ = _resolve_company_id()
        department_id = request.form.get("department_id", type=int)
        work_type_id = request.form.get("work_type_id", type=int)
        remark = (request.form.get("remark") or "").strip() or None
        if not company_id or not department_id or not work_type_id:
            flash("请选择部门与工种。", "danger")
            return redirect(url_for("main.hr_department_capability_map_list"))
        row = HrDepartmentWorkTypeMap.query.filter_by(
            company_id=company_id,
            department_id=department_id,
            work_type_id=work_type_id,
        ).first()
        if not row:
            row = HrDepartmentWorkTypeMap(
                company_id=company_id,
                department_id=department_id,
                work_type_id=work_type_id,
            )
            db.session.add(row)
        row.is_active = True
        row.remark = remark
        db.session.commit()
        flash("部门-工种允许关系已保存。", "success")
        return redirect(url_for("main.hr_department_capability_map_list"))

    @bp.route("/hr/department-capability-map/<int:row_id>/delete", methods=["POST"])
    @login_required
    @menu_required("hr_department_capability_map")
    @capability_required("hr_department_capability_map.edit")
    def hr_department_capability_map_delete(row_id: int):
        row = HrDepartmentWorkTypeMap.query.get_or_404(row_id)
        company_id, _ = _resolve_company_id()
        if int(row.company_id) != int(company_id or 0):
            flash("无权限删除。", "danger")
            return redirect(url_for("main.hr_department_capability_map_list"))
        db.session.delete(row)
        db.session.commit()
        flash("已删除。", "success")
        return redirect(url_for("main.hr_department_capability_map_list"))

    @bp.route("/hr/api/departments-by-company", methods=["GET"])
    @login_required
    @menu_required("hr_employee", "hr_payroll", "hr_performance", "hr_work_type")
    def hr_api_departments_by_company():
        company_id = request.args.get("company_id", type=int)
        if not company_id:
            return jsonify({"departments": []})
        rows = _departments_for_company(company_id)
        return jsonify({"departments": [{"id": row.id, "name": row.name} for row in rows]})

    @bp.route("/hr/api/work-types-by-company", methods=["GET"])
    @login_required
    @menu_required("hr_employee", "hr_payroll", "hr_performance", "hr_work_type")
    def hr_api_work_types_by_company():
        company_id = request.args.get("company_id", type=int)
        if not company_id:
            return jsonify({"work_types": []})
        rows = _work_types_for_company(company_id, active_only=False)
        return jsonify({"work_types": [{"id": row.id, "name": row.name, "is_active": bool(row.is_active)} for row in rows]})

    @bp.route("/hr/api/work-types-by-department", methods=["GET"])
    @login_required
    @menu_required("hr_employee", "hr_payroll", "hr_performance", "hr_work_type")
    def hr_api_work_types_by_department():
        company_id = request.args.get("company_id", type=int)
        department_id = request.args.get("department_id", type=int)
        if not company_id:
            return jsonify({"work_types": []})
        rows = list_work_types(
            company_id=company_id,
            department_id=department_id,
            active_only=False,
        )
        return jsonify(
            {
                "work_types": [
                    {"id": row.id, "name": row.name, "is_active": bool(row.is_active)}
                    for row in rows
                ],
                "allowed_work_type_ids": allowed_work_type_ids_for_department(
                    company_id=company_id,
                    department_id=department_id,
                ),
            }
        )
